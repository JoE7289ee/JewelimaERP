# Copyright (c) 2026, efeone and contributors
# For license information, please see license.txt
"""Whitelisted server endpoints the Jewelima pages call over AJAX."""

import re
import json

import frappe
from frappe.utils import cint, flt


def _company():
	return frappe.defaults.get_defaults().get("company") or frappe.db.get_single_value(
		"Global Defaults", "default_company"
	)


def _abbr():
	return frappe.db.get_value("Company", _company(), "abbr")


def _wh(name):
	"""Full '<name> - <abbr>' warehouse, or None if it doesn't exist."""
	full = "{0} - {1}".format(name, _abbr())
	return full if frappe.db.exists("Warehouse", full) else None


def _stock_move(item, qty, source, target):
	"""Submit a Material Transfer of `qty` (stock UOM) of `item` from source ->
	target warehouse. No-op when qty<=0 or either warehouse is missing."""
	qty = flt(qty)
	if qty <= 0 or not source or not target:
		return None
	se = frappe.get_doc({
		"doctype": "Stock Entry",
		"stock_entry_type": "Material Transfer",
		"company": _company(),
		"items": [{
			"item_code": item,
			"qty": qty,
			"uom": frappe.db.get_value("Item", item, "stock_uom") or "Gram",
			"s_warehouse": source,
			"t_warehouse": target,
			# we track grams here; valuation flows in automatically once gold is
			# actually purchased (until then a zero rate is fine).
			"allow_zero_valuation_rate": 1,
		}],
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	return se.name


@frappe.whitelist()
def melt_gold(warehouse, output_item, output_weight, inputs, send_to_casting=0):
	"""Convert fine gold + alloy into a karat gold (e.g. 999 -> 18KPG) in ONE warehouse.
	Repack Stock Entry: consume each input (gold, alloy) and produce `output_weight` grams
	of `output_item` in the same warehouse. inputs = [{item, weight}, ...]. Output weight
	may be less than total input (the difference is melt loss). With `send_to_casting`
	the repack fires first, then the fresh karat gold transfers straight on to the
	Casting warehouse (rule 1: casting gold lives there)."""
	if isinstance(inputs, str):
		inputs = json.loads(inputs or "[]")
	if not warehouse or not frappe.db.exists("Warehouse", warehouse):
		frappe.throw(frappe._("Pick a valid warehouse."))
	out_w = flt(output_weight)
	if not output_item or out_w <= 0:
		frappe.throw(frappe._("Choose what to create and a positive output weight."))

	src, total_in = [], 0.0
	for i in inputs or []:
		item, w = i.get("item"), flt(i.get("weight"))
		if not item or w <= 0:
			continue
		src.append({
			"item_code": item, "qty": w,
			"uom": frappe.db.get_value("Item", item, "stock_uom") or "Gram",
			"s_warehouse": warehouse, "allow_zero_valuation_rate": 1,
		})
		total_in += w
	if not src:
		frappe.throw(frappe._("Add at least one input (gold + alloy)."))
	if out_w > total_in + 0.0005:
		frappe.throw(frappe._("Output ({0} g) cannot exceed the total input ({1} g).").format(out_w, round(total_in, 3)))

	# every input must actually be in stock in this warehouse
	for s in src:
		avail = flt(frappe.db.get_value("Bin", {"item_code": s["item_code"], "warehouse": warehouse}, "actual_qty"))
		if s["qty"] > avail + 0.0005:
			frappe.throw(frappe._("Only {0} of {1} in stock here — can't melt {2}.").format(round(avail, 3), s["item_code"], s["qty"]))

	se = frappe.get_doc({
		"doctype": "Stock Entry",
		"stock_entry_type": "Repack",
		"company": _company(),
		"items": src + [{
			"item_code": output_item, "qty": out_w,
			"uom": frappe.db.get_value("Item", output_item, "stock_uom") or "Gram",
			"t_warehouse": warehouse, "allow_zero_valuation_rate": 1,
		}],
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	out = {"name": se.name, "total_in": round(total_in, 3), "output": round(out_w, 3), "loss": round(total_in - out_w, 3)}
	if cint(send_to_casting):
		out["casting_transfer"] = _stock_move(output_item, out_w, warehouse, _wh("Casting"))
		out["casting_warehouse"] = _wh("Casting")
	frappe.db.commit()
	return out


@frappe.whitelist()
def get_melt_stock(warehouse):
	"""Non-empty stock in a warehouse for the Melting screen: item, group, purity, weight
	(balance) and pure gold grams (weight x purity% for gram-weighted metals; 0 for stones).
	Sorted by pure desc. Returns rows + gross/pure totals (grams only)."""
	if not warehouse or not frappe.db.exists("Warehouse", warehouse):
		return {"rows": [], "total_weight": 0, "total_pure": 0}
	rows, tw, tp = [], 0.0, 0.0
	for b in frappe.get_all("Bin", filters={"warehouse": warehouse, "actual_qty": [">", 0]}, fields=["item_code", "actual_qty", "stock_uom"]):
		m = frappe.db.get_value("Item", b.item_code, ["item_name", "item_group", "purity_percentage", "stone_type"], as_dict=True) or {}
		wt = flt(b.actual_qty)
		pure = wt * flt(m.get("purity_percentage")) / 100.0 if not m.get("stone_type") else 0.0
		rows.append({
			"item": b.item_code, "item_name": m.get("item_name") or b.item_code,
			"item_group": m.get("item_group") or "", "purity": flt(m.get("purity_percentage")),
			"weight": round(wt, 3), "uom": b.stock_uom or "Gram", "pure": round(pure, 3),
		})
		if (b.stock_uom or "Gram") == "Gram":
			tw += wt
		tp += pure
	rows.sort(key=lambda r: -r["pure"])
	return {"rows": rows, "total_weight": round(tw, 3), "total_pure": round(tp, 3)}


@frappe.whitelist()
def get_warehouse_flags():
	"""All toggleable Warehouse flags (custom Check fields) + every leaf warehouse's current
	values. Powers the Warehouse Management page (flags only — no warehouse create/edit)."""
	flags = frappe.get_all(
		"Custom Field", filters={"dt": "Warehouse", "fieldtype": "Check"},
		fields=["fieldname", "label"], order_by="idx",
	)
	fieldnames = [f.fieldname for f in flags]
	warehouses = frappe.get_all(
		"Warehouse", filters={"is_group": 0},
		fields=["name", "warehouse_name"] + fieldnames, order_by="warehouse_name",
	)
	return {"flags": flags, "warehouses": warehouses}


@frappe.whitelist()
def set_warehouse_flag(warehouse, flag, value):
	"""Toggle ONE flag (a custom Check field) on a warehouse. Flags only — nothing else."""
	frappe.only_for(["System Manager", "Stock Manager"])
	if not frappe.get_all("Custom Field", filters={"dt": "Warehouse", "fieldtype": "Check", "fieldname": flag}, limit=1):
		frappe.throw(frappe._("Unknown warehouse flag: {0}").format(flag))
	if not warehouse or not frappe.db.exists("Warehouse", warehouse):
		frappe.throw(frappe._("No such warehouse."))
	val = 1 if frappe.utils.cint(value) else 0
	frappe.db.set_value("Warehouse", warehouse, flag, val)
	frappe.db.commit()
	return {"ok": 1, "warehouse": warehouse, "flag": flag, "value": val}


@frappe.whitelist()
def transfer_stock(from_warehouse, to_warehouse, items):
	"""Move stock between two warehouses (Material Transfer). items = [{item, weight}].
	Refuses to transfer more than is available in the source for any item."""
	if isinstance(items, str):
		items = json.loads(items or "[]")
	if not from_warehouse or not frappe.db.exists("Warehouse", from_warehouse):
		frappe.throw(frappe._("Pick a valid source warehouse."))
	if not to_warehouse or not frappe.db.exists("Warehouse", to_warehouse):
		frappe.throw(frappe._("Pick a valid destination warehouse."))
	if from_warehouse == to_warehouse:
		frappe.throw(frappe._("Source and destination warehouses must be different."))

	rows = []
	for i in items or []:
		item, wt = i.get("item"), flt(i.get("weight"))
		if not item or wt <= 0:
			continue
		avail = flt(frappe.db.get_value("Bin", {"item_code": item, "warehouse": from_warehouse}, "actual_qty"))
		if wt > avail + 0.0005:
			frappe.throw(frappe._("Only {0} of {1} in {2} — can't transfer {3}.").format(round(avail, 3), item, from_warehouse, wt))
		rows.append({
			"item_code": item, "qty": wt,
			"uom": frappe.db.get_value("Item", item, "stock_uom") or "Gram",
			"s_warehouse": from_warehouse, "t_warehouse": to_warehouse, "allow_zero_valuation_rate": 1,
		})
	if not rows:
		frappe.throw(frappe._("Add at least one item with a weight to transfer."))

	se = frappe.get_doc({
		"doctype": "Stock Entry", "stock_entry_type": "Material Transfer",
		"company": _company(), "items": rows,
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	frappe.db.commit()
	return {"name": se.name, "count": len(rows)}


@frappe.whitelist()
def post_raw_material_purchase(supplier, warehouse, posting_date=None, items=None, voucher_type=None):
	"""Create + submit a Purchase Receipt for raw materials (from the Purchase
	Raw Material page) AND write Jewelima's own Purchase Record — named by the
	voucher type's code series (SIN-0001) — storing what/how much/who/total."""
	from frappe.utils import today

	if not voucher_type or not frappe.db.exists("Voucher Type", voucher_type):
		frappe.throw(frappe._("Pick the voucher type."))

	if isinstance(items, str):
		items = json.loads(items or "[]")
	items = items or []

	# stock qty is the WEIGHT (carats for stones, grams for metal). Stones also carry a
	# piece count, stored on the PR line for reference (it does not affect stock).
	codes = list({i.get("item") for i in items if i.get("item")})
	stype = {}
	if codes:
		for it in frappe.get_all("Item", filters={"name": ["in", codes]}, fields=["name", "stone_type"]):
			stype[it.name] = it.stone_type or ""

	rows = []
	for i in items:
		item = i.get("item")
		weight = flt(i.get("weight"))
		if not item or weight <= 0:
			continue
		count = cint(i.get("count"))
		if stype.get(item) and count <= 0:
			frappe.throw(frappe._("{0} is a stone — enter the piece count (Qty).").format(item))
		row = {"item_code": item, "qty": weight, "rate": flt(i.get("rate")), "warehouse": warehouse}
		if stype.get(item):
			row["custom_stone_count"] = count
		else:
			row["custom_purity"] = flt(i.get("purity"))
		rows.append(row)
	if not rows:
		frappe.throw(frappe._("Add at least one item with a weight."))
	if not supplier:
		frappe.throw(frappe._("Select a supplier."))
	if not warehouse:
		frappe.throw(frappe._("Select a warehouse."))

	pr = frappe.get_doc(
		{
			"doctype": "Purchase Receipt",
			"supplier": supplier,
			"company": _company(),
			"posting_date": posting_date or today(),
			"set_warehouse": warehouse,
			"items": rows,
		}
	)
	pr.insert(ignore_permissions=True)
	pr.submit()

	rec = frappe.get_doc({
		"doctype": "Purchase Record", "voucher_type": voucher_type, "supplier": supplier,
		"purchase_date": posting_date or today(), "warehouse": warehouse, "purchase_receipt": pr.name,
		"items": [{"item": i.get("item"), "weight": flt(i.get("weight")), "count": cint(i.get("count")),
			"purity": flt(i.get("purity")), "rate": flt(i.get("rate"))} for i in items
			if i.get("item") and flt(i.get("weight")) > 0],
	})
	rec.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": pr.name, "record": rec.name, "total": rec.total_amount}


@frappe.whitelist()
def get_order_defaults():
	"""Global Place Order defaults from the Jewelima Order Settings single."""
	s = frappe.get_cached_doc("Jewelima Order Settings")
	return {
		"days": frappe.utils.cint(s.default_days) or 0,
		"order_type": s.default_order_type or None,
		"salesman": s.default_salesman or None,
	}


# --- Order dropdown masters (Setup -> Types & Salesmen) ----------------------------
# kind "type" = Order Type (jewelima master, retire via `disabled`);
# kind "salesman" = Sales Person (ERPNext tree master, retire via `enabled`).

def _order_master_conf(kind):
	if kind == "type":
		return {"doctype": "Order Type", "bag_field": "order_type", "off_field": "disabled", "off_value": 1}
	if kind == "salesman":
		return {"doctype": "Sales Person", "bag_field": "salesman", "off_field": "enabled", "off_value": 0}
	frappe.throw(frappe._("Unknown master kind: {0}").format(kind))


def _active_bags_using(bag_field, value):
	"""Order Bags in an ACTIVE state (not Cancelled/Sold) carrying this value."""
	return frappe.db.sql(
		f"""SELECT COUNT(*) FROM `tabOrder Bag`
		    WHERE `{bag_field}` = %s AND IFNULL(stock_status, '') NOT IN ('Cancelled', 'Sold')""",
		(value,),
	)[0][0]


@frappe.whitelist()
def get_order_masters():
	"""Both dropdown masters with usage + retired state, for the Types & Salesmen page."""
	out = {}
	for kind in ("type", "salesman"):
		c = _order_master_conf(kind)
		filters = {"is_group": 0} if kind == "salesman" else {}
		rows = []
		for d in frappe.get_all(c["doctype"], filters=filters, fields=["name", c["off_field"]], order_by="name"):
			retired = bool(d[c["off_field"]]) if kind == "type" else not d[c["off_field"]]
			rows.append({
				"name": d.name,
				"retired": retired,
				"active_bags": int(_active_bags_using(c["bag_field"], d.name)),
				"total_bags": int(frappe.db.count("Order Bag", {c["bag_field"]: d.name})),
			})
		out[kind] = rows
	return out


@frappe.whitelist()
def add_order_master(kind, name):
	frappe.only_for(["System Manager", "Stock Manager"])
	c = _order_master_conf(kind)
	name = (name or "").strip().upper()
	if not name:
		frappe.throw(frappe._("Name is required"))
	if frappe.db.exists(c["doctype"], name):
		frappe.throw(frappe._("{0} '{1}' already exists").format(c["doctype"], name))
	if kind == "type":
		frappe.get_doc({"doctype": "Order Type", "order_type_name": name}).insert(ignore_permissions=True)
	else:
		root = frappe.db.get_value("Sales Person", {"is_group": 1, "parent_sales_person": ["in", ["", None]]}, "name")
		frappe.get_doc({
			"doctype": "Sales Person", "sales_person_name": name,
			"parent_sales_person": root, "is_group": 0, "enabled": 1,
		}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": name}


@frappe.whitelist()
def retire_order_master(kind, name, restore=0):
	"""Retire (hide from dropdowns) or restore a Type / Salesman. Retiring is blocked while
	ACTIVE Order Bags use it (Cancelled/Sold don't count); a never-used value is deleted."""
	frappe.only_for(["System Manager", "Stock Manager"])
	c = _order_master_conf(kind)
	if not frappe.db.exists(c["doctype"], name):
		frappe.throw(frappe._("Not found"))
	if frappe.utils.cint(restore):
		frappe.db.set_value(c["doctype"], name, c["off_field"], 0 if kind == "type" else 1)
		frappe.db.commit()
		return {"name": name, "state": "restored"}
	active = _active_bags_using(c["bag_field"], name)
	if active:
		frappe.throw(frappe._("Cannot retire '{0}' — {1} active Order Bag(s) use it (only Cancelled/Sold usage may be retired).").format(name, active))
	if not frappe.db.count("Order Bag", {c["bag_field"]: name}) and not frappe.db.count("Job Order", {c["bag_field"]: name}):
		frappe.delete_doc(c["doctype"], name, ignore_permissions=True)
		frappe.db.commit()
		return {"name": name, "state": "deleted"}
	frappe.db.set_value(c["doctype"], name, c["off_field"], c["off_value"])
	frappe.db.commit()
	return {"name": name, "state": "retired"}


# --- Tree Making: per-karat queues -> one wax tree -> off to CASTING ---------------

def _karat_suffix(item):
	"""18KPG -> 18P, 22KYG -> 22Y (same convention as design purity variants)."""
	import re

	m = re.match(r"^(\d+)K([A-Z])G$", item or "")
	return f"{m.group(1)}{m.group(2)}" if m else "MIX"


def _bag_karat(bag_name, karat_names):
	"""The karat gold in the bag's OWN BOM (its casting metal); None if it has none."""
	for it in frappe.get_all("Order Bag BOM Item", filters={"parent": bag_name, "parenttype": "Order Bag"}, pluck="item"):
		if it in karat_names:
			return it
	return None


@frappe.whitelist()
def get_tree_queues():
	"""Cards waiting at TREE MAKING grouped by their casting karat — one queue (table)
	per purity. Bags whose BOM has no karat gold fall into 'OTHER'."""
	karat_names = set(frappe.get_all("Item", filters={"material_group": "GOLD", "metal_purity": ["!=", ""]}, pluck="name"))
	bags = frappe.get_all(
		"Order Bag",
		filters={"location": "TREE MAKING", "is_finished": 0, "stock_status": ["not in", ["Cancelled", "Sold"]], "tree": ["in", ["", None]]},
		fields=["name", "design", "qty", "size", "due_date", "customer", "nett_weight"],
		order_by="name",
	)
	queues = {}
	for b in bags:
		karat = _bag_karat(b.name, karat_names)
		key = karat or "OTHER"
		b["karat"] = karat
		queues.setdefault(key, []).append(b)
	return [{"karat": k, "suffix": _karat_suffix(k if k != "OTHER" else None), "cards": rows} for k, rows in sorted(queues.items())]


@frappe.whitelist()
def get_order_stock_analysis():
	"""Stock Analysis (Reports > Orders): per raw material, what the ACTIVE order
	book still needs vs what the issue warehouses hold.
	  plan        = Σ over active bags (not finished / Cancelled / Sold) of bag-BOM weight × qty
	  issued      = what those bags already received (ledger In − Out)
	  outstanding = max(0, plan − issued)
	  available   = stock across the issue-side warehouses (Raw Materials Store,
	                Gold Issue, Stone Issue)
	Balance = available − outstanding; anything negative is called out SHORT."""
	bags = frappe.get_all(
		"Order Bag",
		filters={"is_finished": 0, "stock_status": ["not in", ["Cancelled", "Sold"]]},
		fields=["name", "qty"],
	)
	if not bags:
		return {"rows": [], "bags": 0, "short": 0}
	qty_of = {b.name: max(cint(b.qty) or 1, 1) for b in bags}
	names = list(qty_of)

	plan = {}
	for r in frappe.get_all("Order Bag BOM Item", filters={"parent": ["in", names]},
	                        fields=["parent", "item", "weight"]):
		if r.item and flt(r.weight) > 0:
			plan[r.item] = plan.get(r.item, 0.0) + flt(r.weight) * qty_of[r.parent]

	issued = {}
	for r in frappe.db.sql(
		"""select item, sum(case when direction = 'In' then qty else -qty end)
		   from `tabBag Material Ledger` where order_bag in %(bags)s group by item""",
		{"bags": names},
	):
		issued[r[0]] = flt(r[1])

	issue_whs = [w for w in (_wh("Raw Materials Store"), _wh("Gold Issue"), _wh("Stone Issue"))
	             if w and frappe.db.exists("Warehouse", w)]
	rows, short = [], 0
	for item in sorted(plan):
		p = plan[item]
		iss = max(0.0, issued.get(item, 0.0))
		outstanding = max(0.0, p - iss)
		available = sum(flt(b.actual_qty) for b in frappe.get_all(
			"Bin", filters={"item_code": item, "warehouse": ["in", issue_whs]}, fields=["actual_qty"]))
		balance = available - outstanding
		meta = frappe.db.get_value("Item", item, ["item_group", "stock_uom", "stone_type"], as_dict=True) or {}
		is_short = outstanding > 0 and balance < -0.0005
		short += 1 if is_short else 0
		rows.append({
			"item": item, "group": meta.get("item_group") or "", "uom": meta.get("stock_uom") or "",
			"is_stone": bool(meta.get("stone_type")),
			"plan": round(p, 3), "issued": round(iss, 3), "outstanding": round(outstanding, 3),
			"available": round(available, 3), "balance": round(balance, 3), "short": is_short,
		})
	# the ones in trouble first, biggest outstanding next
	rows.sort(key=lambda r: (not r["short"], -r["outstanding"]))
	return {"rows": rows, "bags": len(names), "short": short, "issue_warehouses": issue_whs}


@frappe.whitelist()
def get_gold_casting_report():
	"""The Gold Casting report (Reports > Casting): every tree awaiting cast (its
	bags sit at CASTING), the karat-gold requirement per gold item, what the
	Casting warehouse already holds (RULE: casting gold ALWAYS comes from there —
	stems recycle back), the shortfall to melt, and whether the melt warehouses
	hold enough Standard Gold (pure-equivalent) + Alloy to cover it."""
	casting_wh = _wh("Casting")

	tree_names = [r[0] for r in frappe.db.sql(
		"""select distinct ob.tree from `tabOrder Bag` ob
		   join `tabWax Tree` wt on wt.name = ob.tree
		   where ob.location = 'CASTING' and ifnull(ob.tree, '') != '' and ifnull(wt.cast, 0) = 0""")]
	trees, per_karat = [], {}
	for t in sorted(tree_names):
		doc = frappe.get_doc("Wax Tree", t)
		sw, gr, pg = flt(doc.stone_weight), flt(doc.gold_required), flt(doc.pure_gold_needed)
		if not gr and flt(doc.wax_weight) and doc.karat:
			# tree made before the casting fields existed — compute on the fly
			n = _tree_casting_numbers(doc.karat, doc.wax_weight, [c.order_bag for c in doc.cards])
			sw, gr, pg = n["stone_weight"], n["gold_required"], n["pure_gold_needed"]
		mp = frappe.db.get_value("Item", doc.karat, "metal_purity") if doc.karat else ""
		trees.append({
			"tree": t, "karat": doc.karat or "", "metal_purity": mp or "",
			"cards": len(doc.cards),
			"employee": (frappe.db.get_value("Employee", doc.employee, "employee_name") if doc.employee else "") or "",
			"made_on": doc.made_on, "wax_weight": flt(doc.wax_weight),
			"stone_weight": sw, "gold_required": gr, "pure_gold_needed": pg,
		})
		if doc.karat:
			agg = per_karat.setdefault(doc.karat, {"trees": 0, "required": 0.0})
			agg["trees"] += 1
			agg["required"] += gr

	karats, pure_needed_total, alloy_needed_total = [], 0.0, 0.0
	for item in sorted(per_karat):
		agg = per_karat[item]
		purity = flt(frappe.db.get_value("Item", item, "purity_percentage"))
		available = flt(frappe.db.get_value("Bin", {"item_code": item, "warehouse": casting_wh}, "actual_qty"))
		short = max(0.0, agg["required"] - available)
		pure = short * purity / 100.0
		alloy = short - pure
		pure_needed_total += pure
		alloy_needed_total += alloy
		karats.append({
			"item": item, "purity": purity, "trees": agg["trees"],
			"required": round(agg["required"], 3), "available": round(available, 3),
			"shortfall": round(short, 3), "pure_needed": round(pure, 3), "alloy_needed": round(alloy, 3),
		})

	# melt-side stock: Standard Gold as PURE-equivalent + Alloy, in the melt warehouses
	melt_whs = frappe.get_all("Warehouse", filters={"is_melt_warehouse": 1}, pluck="name") or [_wh("Raw Materials Store")]
	pure_available = 0.0
	for it in frappe.get_all("Item", filters={"item_group": "GOLD STANDARD"}, fields=["name", "purity_percentage"]):
		qty = sum(flt(b.actual_qty) for b in frappe.get_all(
			"Bin", filters={"item_code": it.name, "warehouse": ["in", melt_whs]}, fields=["actual_qty"]))
		pure_available += qty * flt(it.purity_percentage) / 100.0
	alloy_available = 0.0
	for it in frappe.get_all("Item", filters={"item_group": "ALLOY"}, pluck="name"):
		alloy_available += sum(flt(b.actual_qty) for b in frappe.get_all(
			"Bin", filters={"item_code": it, "warehouse": ["in", melt_whs]}, fields=["actual_qty"]))

	return {
		"casting_warehouse": casting_wh,
		"trees": trees,
		"karats": karats,
		"melt": {
			"pure_needed": round(pure_needed_total, 3),
			"pure_available": round(pure_available, 3),
			"pure_short": round(max(0.0, pure_needed_total - pure_available), 3),
			"alloy_needed": round(alloy_needed_total, 3),
			"alloy_available": round(alloy_available, 3),
			"alloy_short": round(max(0.0, alloy_needed_total - alloy_available), 3),
			"melt_warehouses": melt_whs,
		},
	}


# --- Casting bench (Manufacturing > Casting) ---------------------------------------
def _bag_gold_and_stone(order_bag):
	"""(gold grams held, stone grams held = issued carats x 0.2) for one bag."""
	gold = stone = 0.0
	for it in get_bag_contents(order_bag)["items"]:
		if it["qty"] <= 0:
			continue
		if frappe.db.get_value("Item", it["item"], "stone_type"):
			stone += flt(it["qty"]) * 0.2
		else:
			gold += flt(it["qty"])
	return round(gold, 3), round(stone, 3)


@frappe.whitelist()
def get_casting_queue():
	"""Trees awaiting cast (cast=0, still holding cards at CASTING) + the stock the
	Casting warehouse holds — the Casting bench page."""
	tree_names = [r[0] for r in frappe.db.sql(
		"""select distinct ob.tree from `tabOrder Bag` ob
		   join `tabWax Tree` wt on wt.name = ob.tree
		   where ob.location = 'CASTING' and ifnull(ob.tree, '') != '' and ifnull(wt.cast, 0) = 0""")]
	trees = []
	for t in sorted(tree_names):
		doc = frappe.get_doc("Wax Tree", t)
		weighted = sum(1 for c in doc.cards if _bag_gold_and_stone(c.order_bag)[0] > 0.0005)
		trees.append({
			"tree": t, "karat": doc.karat or "", "cards": len(doc.cards), "weighted": weighted,
			"wax_weight": flt(doc.wax_weight), "gold_required": flt(doc.gold_required),
			"casting_date": doc.casting_date,
			"employee": (frappe.db.get_value("Employee", doc.employee, "employee_name") if doc.employee else "") or "",
			"made_on": doc.made_on,
		})
	stock = [{"item": r[0], "purity": flt(r[1]), "qty": flt(r[2])} for r in frappe.db.sql(
		"""select b.item_code, i.purity_percentage, b.actual_qty from tabBin b
		   join tabItem i on i.name = b.item_code
		   where b.warehouse = %s and abs(b.actual_qty) > 0.0005 order by b.item_code""", _wh("Casting"))]
	return {"trees": trees, "stock": stock, "casting_warehouse": _wh("Casting")}


@frappe.whitelist()
def set_tree_casting_date(tree, date=None):
	"""Plan (or clear) a tree's casting date from the queue."""
	frappe.db.set_value("Wax Tree", tree, "casting_date", date or None)
	frappe.db.commit()


@frappe.whitelist()
def get_tree_for_weighing(card=None, tree=None):
	"""The weigh page's loader: scan a CARD (must sit at CASTING, on a tree) or
	come straight from the queue with a tree. Returns the tree + every card with
	its issued-stone grams and gold already held."""
	if card:
		bag = frappe.db.get_value("Order Bag", card,
		                          ["name", "location", "tree", "is_finished", "stock_status"], as_dict=True)
		if not bag:
			frappe.throw(frappe._("No card {0}.").format(card))
		if bag.is_finished or bag.stock_status in ("Cancelled", "Sold"):
			frappe.throw(frappe._("{0} is {1} — nothing to cast.").format(card, "finished" if bag.is_finished else bag.stock_status))
		if not bag.tree:
			frappe.throw(frappe._("{0} is not on any wax tree — make the tree first.").format(card))
		if bag.location != "CASTING":
			frappe.throw(frappe._("{0} is at {1}, not CASTING — it can't be cast-weighed from there.").format(card, bag.location or "—"))
		tree = bag.tree
	if not tree or not frappe.db.exists("Wax Tree", tree):
		frappe.throw(frappe._("Scan a card (or pick a tree from the queue)."))
	doc = frappe.get_doc("Wax Tree", tree)
	cards = []
	for c in doc.cards:
		bagv = frappe.db.get_value("Order Bag", c.order_bag, ["location", "nett_weight"], as_dict=True) or {}
		gold, stone = _bag_gold_and_stone(c.order_bag)
		cards.append({
			"order_bag": c.order_bag, "design": c.design or "", "qty": c.qty or 1,
			"location": bagv.get("location") or "—", "stone_g": stone, "gold_held": gold,
			"plan_gold": round(flt(bagv.get("nett_weight")), 3),  # the design BOM's gold, qty-scaled
			"weighable": bagv.get("location") == "CASTING",
		})
	karat_stock = flt(frappe.db.get_value("Bin", {"item_code": doc.karat, "warehouse": _wh("Casting")}, "actual_qty"))
	return {
		"tree": tree, "karat": doc.karat or "", "cast": cint(doc.cast),
		"casting_date": doc.casting_date, "wax_weight": flt(doc.wax_weight),
		"gold_required": flt(doc.gold_required), "karat_stock": round(karat_stock, 3),
		"scanned": card or "", "cards": cards,
	}


@frappe.whitelist()
def cast_weigh(tree, entries):
	"""Book cast weights onto a tree's cards. Each entry = {order_bag, gross}: the
	machine reads GROSS; gold booked = gross − issued stones (ct x 0.2), pulled
	from the CASTING warehouse (one move per card). Cards stay at CASTING. When
	every card on the tree holds gold, the tree is marked Cast and the casting
	date stamps itself."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE

	if isinstance(entries, str):
		entries = json.loads(entries or "[]")
	doc = frappe.get_doc("Wax Tree", tree)
	if not doc.karat:
		frappe.throw(frappe._("{0} has no karat — can't weigh.").format(tree))
	tree_cards = {c.order_bag for c in doc.cards}
	cast_wh, in_bags = _wh("Casting"), _wh(IN_PRODUCTION_WAREHOUSE)

	lines, total_gold = [], 0.0
	for e in entries or []:
		bag, gross = e.get("order_bag"), flt(e.get("gross"))
		if not bag or gross <= 0:
			continue
		if bag not in tree_cards:
			frappe.throw(frappe._("{0} is not on tree {1}.").format(bag, tree))
		loc = frappe.db.get_value("Order Bag", bag, "location")
		if loc != "CASTING":
			frappe.throw(frappe._("{0} is at {1}, not CASTING.").format(bag, loc or "—"))
		stone = _bag_gold_and_stone(bag)[1]
		gold = round(gross - stone, 3)
		if gold <= 0:
			frappe.throw(frappe._("{0}: gross {1} g ≤ its stones ({2} g) — nothing left as gold.").format(bag, gross, stone))
		lines.append({"bag": bag, "gross": gross, "stone": stone, "gold": gold})
		total_gold += gold
	if not lines:
		frappe.throw(frappe._("Enter at least one gross weight."))

	avail = flt(frappe.db.get_value("Bin", {"item_code": doc.karat, "warehouse": cast_wh}, "actual_qty"))
	if total_gold > avail + 0.0005:
		frappe.throw(frappe._("Only {0} g of {1} in the Casting warehouse — these weights need {2} g. Melt first.")
		             .format(round(avail, 3), doc.karat, round(total_gold, 3)))

	for ln in lines:
		_bag_ledger(ln["bag"], doc.karat, "In", ln["gold"], "Casting",
		            remarks=f"cast gross {ln['gross']} g − stones {ln['stone']} g")
		_stock_move(doc.karat, ln["gold"], cast_wh, in_bags)  # reduced one by one
		_recompute_bag_from_contents(ln["bag"])

	tree_cast = all(_bag_gold_and_stone(c.order_bag)[0] > 0.0005 for c in doc.cards)
	if tree_cast and not cint(doc.cast):
		frappe.db.set_value("Wax Tree", tree, {
			"cast": 1, "casting_date": doc.casting_date or frappe.utils.today()})
	frappe.db.commit()
	remaining = flt(frappe.db.get_value("Bin", {"item_code": doc.karat, "warehouse": cast_wh}, "actual_qty"))
	return {"booked": lines, "total_gold": round(total_gold, 3),
	        "tree_cast": tree_cast, "remaining_stock": round(remaining, 3)}


# wax -> metal casting conversion, per KARAT (same for all colors); stem is fixed
CASTING_MULTIPLIER = {"14K": 13.5, "18K": 15.5, "22K": 18.5}
CASTING_STEM_G = 3.0


def _tree_casting_numbers(karat_item, wax_weight, bag_names):
	"""stone_weight (issued stones only, ct x 0.2), gold_required, pure_gold_needed."""
	stone_g = 0.0
	for nm in bag_names or []:
		for it in get_bag_contents(nm)["items"]:
			if it["qty"] > 0 and frappe.db.get_value("Item", it["item"], "stone_type"):
				stone_g += flt(it["qty"]) * 0.2
	mp, purity = frappe.db.get_value("Item", karat_item, ["metal_purity", "purity_percentage"]) or (None, 0)
	mult = CASTING_MULTIPLIER.get(mp or "", 0)
	gold = max(0.0, flt(wax_weight) - CASTING_STEM_G - stone_g) * mult if flt(wax_weight) else 0.0
	return {
		"stone_weight": round(stone_g, 3),
		"gold_required": round(gold, 3),
		"pure_gold_needed": round(gold * flt(purity) / 100.0, 3),
	}


@frappe.whitelist()
def make_tree(karat, names, employee=None, wax_weight=None):
	"""Mount the selected TREE MAKING cards onto ONE wax tree (a Wax Tree record,
	numbered T-<karat>-###), stamp the tree + employee on their bench records, and
	transfer every card to CASTING."""
	if isinstance(names, str):
		names = json.loads(names or "[]")
	names = [n for n in names if n]
	if not names:
		frappe.throw(frappe._("Select at least one card."))

	karat_names = set(frappe.get_all("Item", filters={"material_group": "GOLD", "metal_purity": ["!=", ""]}, pluck="name"))
	karat_val = karat if karat and karat != "OTHER" else None
	cards = []
	for nm in names:
		bag = frappe.db.get_value("Order Bag", nm, ["name", "design", "qty", "location", "tree"], as_dict=True)
		if not bag:
			frappe.throw(frappe._("No Order Bag {0}").format(nm))
		if bag.location != "TREE MAKING":
			frappe.throw(frappe._("{0} is at {1}, not TREE MAKING.").format(nm, bag.location))
		if bag.tree:
			frappe.throw(frappe._("{0} is already on tree {1}.").format(nm, bag.tree))
		bk = _bag_karat(nm, karat_names)
		if (bk or None) != karat_val:
			frappe.throw(frappe._("{0} is {1} — this tree is {2}. One purity per tree.").format(nm, bk or "OTHER", karat or "OTHER"))
		cards.append(bag)

	# next number in this karat's series: T-18P-001, T-18P-002, …
	suffix = _karat_suffix(karat_val)
	prefix = f"T-{suffix}-"
	last = frappe.db.sql("SELECT name FROM `tabWax Tree` WHERE name LIKE %s ORDER BY name DESC LIMIT 1", (prefix + "%",))
	nxt = (int(last[0][0].rsplit("-", 1)[1]) + 1) if last else 1
	nums = _tree_casting_numbers(karat_val, wax_weight, [c.name for c in cards]) if karat_val else {}
	tree = frappe.get_doc({
		"doctype": "Wax Tree",
		"tree_no": f"{prefix}{nxt:03d}",
		"karat": karat_val,
		"employee": employee or None,
		"made_on": frappe.utils.now_datetime(),
		"wax_weight": flt(wax_weight) or None,
		"cards": [{"order_bag": c.name, "design": c.design, "qty": c.qty} for c in cards],
		**nums,
	}).insert(ignore_permissions=True)

	# stamp the bags + their TREE MAKING records, then move everything to CASTING
	errors = []
	for c in cards:
		frappe.db.set_value("Order Bag", c.name, "tree", tree.name)
		rec = _current_bench_record("Tree Making", c.name)
		if rec:
			vals = {"tree": tree.name}
			if employee:
				vals["employee"] = employee
			frappe.db.set_value("Tree Making", rec, vals)
		try:
			transfer_order_bag(c.name, "CASTING")
		except Exception as e:
			errors.append({"name": c.name, "error": str(e)})
	frappe.db.commit()
	return {"tree": tree.name, "karat": karat_val, "count": len(cards) - len(errors), "errors": errors}


# --- CAD jobs: targets live on the bag until the real design is finalized ----------

def _cad_siblings(bag):
	"""Other is_cad bags of the same Job Order with the SAME targets (split twins etc.)."""
	return frappe.get_all(
		"Order Bag",
		filters={
			"job_order": bag.job_order, "is_cad": 1, "name": ["!=", bag.name],
			"cad_design_type": bag.cad_design_type or "", "cad_karat": bag.cad_karat or "",
			"cad_gold_weight": bag.cad_gold_weight or "", "cad_diamond_weight": flt(bag.cad_diamond_weight),
		},
		pluck="name",
	)


@frappe.whitelist()
def get_cad_bag_info(order_bag):
	"""A CAD bag's targets + its same-target siblings — feeds the Finalize dialog."""
	bag = frappe.db.get_value(
		"Order Bag", order_bag,
		["name", "job_order", "is_cad", "qty", "size", "location", "cad_design_type", "cad_karat",
		 "cad_gold_weight", "cad_diamond_weight", "cad_stone_no", "cad_reference", "cad_remarks"],
		as_dict=True,
	)
	if not bag:
		frappe.throw(frappe._("No Order Bag {0}").format(order_bag))
	bag["siblings"] = _cad_siblings(bag) if bag.is_cad else []
	return bag


@frappe.whitelist()
def finalize_cad_design(order_bag, design_name, design_style=None, image=None, materials=None, apply_to_siblings=1):
	"""CAD done: create the REAL design and attach it to the bag (and its same-target
	siblings) — design set, bag BOM seeded, plan recomputed, image pulled, is_cad cleared.
	From here the bag collects and routes like any designed order."""
	bag = frappe.get_doc("Order Bag", order_bag)
	if not bag.is_cad:
		frappe.throw(frappe._("{0} is not awaiting a CAD design.").format(order_bag))
	if not bag.cad_design_type:
		frappe.throw(frappe._("{0} has no CAD design type recorded.").format(order_bag))

	res = create_design(design_name, bag.cad_design_type, design_style, image, materials)
	design = frappe.get_doc("Design", res["name"])

	targets = [bag.name] + (_cad_siblings(bag) if frappe.utils.cint(apply_to_siblings) else [])
	for nm in targets:
		b = frappe.get_doc("Order Bag", nm)
		b.design = design.name
		b.image = design.image or b.image  # keep the CAD reference photo if the design has none
		b.is_cad = 0
		b.set("bag_bom", [])
		for m in design.materials:
			b.append("bag_bom", {"item": m.item, "qty": m.qty, "weight": m.weight})
		b.save(ignore_permissions=True)  # validate() recomputes the plan from the new BOM
	frappe.db.commit()
	return {"design": design.name, "bags": targets, **res}


@frappe.whitelist()
def get_cad_jobs():
	"""All bags still awaiting a CAD design — the CAD Jobs page."""
	return frappe.get_all(
		"Order Bag", filters={"is_cad": 1},
		fields=["name", "job_order", "qty", "size", "location", "customer", "due_date",
			"cad_design_type", "cad_karat", "cad_gold_weight", "cad_diamond_weight", "cad_stone_no", "cad_reference", "cad_remarks"],
		order_by="creation",
	)


@frappe.whitelist()
def get_design_variants(design):
	"""Purity-variant options for a design (the Place Order 'New' button).

	Finds the design BOM's gold row, then for every OTHER karat gold (22KPG, 18KYG, …)
	returns the would-be variant name (base + karat suffix, e.g. 'A 2849 PB' + 18KPG ->
	'A 2849 PB 18P') and whether that Design already exists."""
	import re

	if not design or not frappe.db.exists("Design", design):
		frappe.throw(frappe._("No such design"))
	doc = frappe.get_doc("Design", design)

	karats = frappe.get_all(
		"Item", filters={"material_group": "GOLD", "metal_purity": ["!=", ""]},
		fields=["name", "purity_percentage"], order_by="name desc",
	)
	karat_names = {k.name for k in karats}

	def suffix(item):  # 18KPG -> 18P, 14KWG -> 14W
		m = re.match(r"^(\d+)K([A-Z])G$", item)
		return f"{m.group(1)}{m.group(2)}" if m else item

	# the BOM's gold row = a karat-gold item (fall back to any metal row)
	current_gold = next((m.item for m in doc.materials if m.item in karat_names), None)
	if not current_gold:
		stones = {i.name for i in frappe.get_all("Item", filters={"name": ["in", [m.item for m in doc.materials]], "stone_type": ["!=", ""]})}
		current_gold = next((m.item for m in doc.materials if m.item not in stones), None)
	if not current_gold:
		frappe.throw(frappe._("{0} has no gold row in its BOM to swap.").format(design))

	# strip an existing karat suffix so variants of variants share one base name
	base = doc.design_name
	m = re.match(r"^(.*)\s\d+[A-Z]$", base)
	root = m.group(1) if m and any(base.endswith(" " + suffix(k.name)) for k in karats) else base

	# the design FAMILY = the root-named design + every 'root <suffix>' design.
	# Map each member's actual BOM gold -> its name, so a karat counts as "exists" when ANY
	# family member carries it (e.g. from 'A 2849 PB 18Y', 22KYG = the original 'A 2849 PB' —
	# no duplicate 'A 2849 PB 22Y' can be created).
	family = {}
	for nm in [root] + [f"{root} {suffix(k.name)}" for k in karats]:
		if not frappe.db.exists("Design", nm):
			continue
		mats = frappe.get_all("Design BOM Item", filters={"parent": nm, "parenttype": "Design"}, pluck="item")
		g = next((i for i in mats if i in karat_names), None)
		if g and g not in family:
			family[g] = nm

	out = []
	for k in karats:
		if k.name == current_gold:
			continue
		existing = family.get(k.name)
		out.append({
			"karat": k.name, "purity": flt(k.purity_percentage), "suffix": suffix(k.name),
			"variant_name": existing or f"{root} {suffix(k.name)}", "exists": bool(existing),
		})
	return {
		"design": design, "base": root, "current_gold": current_gold,
		"current_purity": flt(frappe.db.get_value("Item", current_gold, "purity_percentage")),
		"design_type": doc.design_type, "design_style": doc.design_style, "image": doc.image,
		"variants": out,
	}


@frappe.whitelist()
def get_allowed_quick_pages(routes):
	"""Which of these desk-page routes may the current user open? (Ctrl+Space quick
	menu — items the user has no role for are dropped, not shown dead.)"""
	import json

	if isinstance(routes, str):
		routes = json.loads(routes)
	out = []
	for r in routes or []:
		if not frappe.db.exists("Page", r):
			continue
		if frappe.get_cached_doc("Page", r).is_permitted():
			out.append(r)
	return out


@frappe.whitelist()
def get_design_types_with_sizes():
	"""All Design Types with their size lists — the Setup page grid + the Place Order
	Size dropdown (sizes follow the picked design's type)."""
	out = []
	used = {r.design_type: r.cnt for r in frappe.db.sql(
		"SELECT design_type, COUNT(*) cnt FROM `tabDesign` GROUP BY design_type", as_dict=True)}
	for t in frappe.get_all("Design Type", order_by="name", pluck="name"):
		rows = frappe.get_all(
			"Design Type Size", filters={"parent": t, "parenttype": "Design Type"},
			fields=["size", "is_default"], order_by="idx",
		)
		default = next((r.size for r in rows if r.is_default), None)
		out.append({
			"design_type": t, "sizes": [r.size for r in rows],
			"default": default, "used_by": int(used.get(t, 0)),
		})
	return out


@frappe.whitelist()
def set_design_type_default(design_type, size=None):
	"""Mark one size as the type's default (pre-selected on Place Order). Pass no size
	(or the current default) to clear it."""
	frappe.only_for(["System Manager", "Stock Manager"])
	doc = frappe.get_doc("Design Type", design_type)
	size = (size or "").strip()
	if size and size not in [r.size for r in doc.sizes]:
		frappe.throw(frappe._("'{0}' is not a size of {1}").format(size, design_type))
	for r in doc.sizes:
		r.is_default = 1 if (size and r.size == size) else 0
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"design_type": design_type, "default": size or None}


@frappe.whitelist()
def add_design_type(name):
	frappe.only_for(["System Manager", "Stock Manager"])
	name = (name or "").strip().upper()
	if not name:
		frappe.throw(frappe._("Type name is required"))
	if frappe.db.exists("Design Type", name):
		frappe.throw(frappe._("Design Type '{0}' already exists").format(name))
	frappe.get_doc({"doctype": "Design Type", "design_type_name": name}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"design_type": name}


@frappe.whitelist()
def delete_design_type(name):
	frappe.only_for(["System Manager", "Stock Manager"])
	used = frappe.db.count("Design", {"design_type": name})
	if used:
		frappe.throw(frappe._("Cannot delete — {0} Design(s) use this type.").format(used))
	frappe.delete_doc("Design Type", name, ignore_permissions=True)
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def set_design_type_sizes(design_type, sizes):
	"""Replace a Design Type's size list (the Setup page's add/remove)."""
	frappe.only_for(["System Manager", "Stock Manager"])
	sizes = frappe.parse_json(sizes) if isinstance(sizes, str) else (sizes or [])
	seen, clean = set(), []
	for s in sizes:
		s = (s or "").strip()
		if s and s not in seen:
			seen.add(s)
			clean.append(s)
	# a size can't be removed while an ACTIVE Order Bag of this type is using it
	# (Cancelled / Sold bags are history — they don't block).
	doc = frappe.get_doc("Design Type", design_type)
	removed = [row.size for row in doc.sizes if row.size not in seen]
	for s in removed:
		used = frappe.db.sql(
			"""SELECT COUNT(*) FROM `tabOrder Bag` ob
			   JOIN `tabDesign` d ON d.name = ob.design
			   WHERE d.design_type = %s AND ob.size = %s
			     AND IFNULL(ob.stock_status, '') NOT IN ('Cancelled', 'Sold')""",
			(design_type, s),
		)[0][0]
		if used:
			frappe.throw(
				frappe._("Cannot remove size '{0}' — {1} active Order Bag(s) of type {2} are using it.").format(s, used, design_type)
			)
	doc.reload()  # fresh copy — guards against stale-cache row loss on rebuild
	old_default = next((r.size for r in doc.sizes if r.is_default), None)
	doc.set("sizes", [])
	for s in clean:
		doc.append("sizes", {"size": s, "is_default": 1 if s == old_default else 0})
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"design_type": design_type, "sizes": clean}


@frappe.whitelist()
def create_design(design_name, design_type, design_style=None, image=None, materials=None):
	"""Quick-create a Design from the Place Order dialog. The Design controller
	provisions the sellable Item + BOM and derives the stone counts. Returns the
	new design + its derived stone profile so the caller can fill a line."""
	if isinstance(materials, str):
		materials = json.loads(materials or "[]")
	materials = materials or []
	if frappe.db.exists("Design", design_name):
		frappe.throw(frappe._("A design named {0} already exists.").format(design_name))

	rows = [
		{
			"item": m.get("item"),
			"qty": m.get("qty") or 0,
			"weight": m.get("weight") or 0,
		}
		for m in materials
		if m.get("item")
	]
	if not rows:
		frappe.throw(frappe._("Add at least one material to the design's BOM."))

	doc = frappe.get_doc(
		{
			"doctype": "Design",
			"design_name": design_name,
			"design_type": design_type,
			"design_style": design_style or None,
			"image": image or None,
			"materials": rows,
		}
	)
	doc.insert(ignore_permissions=True)
	return {
		"name": doc.name,
		"dmd_no": doc.dmd_no, "ps_no": doc.ps_no, "cs_no": doc.cs_no, "purity": doc.purity,
	}


def _profile_from_materials(mats):
	"""Line profile from a materials list (rows with item/qty/purity/weight): stones
	(by item.stone_type) -> dmd/ps/cs counts + carat weights; metal -> nett grams +
	gram-weighted purity. gross = metal grams; nett = gross - stone grams (1 ct = 0.2 g)."""
	out = {
		"dmd_no": 0, "ps_no": 0, "cs_no": 0, "cz_no": 0, "cvd_no": 0, "pdmd_no": 0, "poth_no": 0,
		"dmd_weight": 0.0, "ps_weight": 0.0, "cs_weight": 0.0, "cz_weight": 0.0, "cvd_weight": 0.0, "pdmd_weight": 0.0, "poth_weight": 0.0,
		"gross_weight": 0.0, "nett_weight": 0.0, "purity": 0.0,
	}
	rows = mats or []
	codes = list({m.get("item") for m in rows if m.get("item")})
	stype, purity_map = {}, {}
	if codes:
		for it in frappe.get_all("Item", filters={"name": ["in", codes]}, fields=["name", "stone_type", "purity_percentage"]):
			stype[it.name] = it.stone_type
			purity_map[it.name] = flt(it.purity_percentage)
	NO_BUCKET = {"Diamond": "dmd_no", "Precious Stone": "ps_no", "Color Stone": "cs_no", "Cubic Zirconia": "cz_no", "CVD": "cvd_no", "Party Diamond": "pdmd_no", "Party Other": "poth_no"}
	WT_BUCKET = {"Diamond": "dmd_weight", "Precious Stone": "ps_weight", "Color Stone": "cs_weight", "Cubic Zirconia": "cz_weight", "CVD": "cvd_weight", "Party Diamond": "pdmd_weight", "Party Other": "poth_weight"}
	metal_g = 0.0
	purity_num = 0.0
	metal_purities = []
	for m in rows:
		st = stype.get(m.get("item"))
		if st in NO_BUCKET:
			out[NO_BUCKET[st]] += int(m.get("qty") or 0)
			out[WT_BUCKET[st]] += flt(m.get("weight"))
		else:
			pu = purity_map.get(m.get("item")) or flt(m.get("purity"))
			metal_g += flt(m.get("weight"))
			purity_num += flt(m.get("weight")) * pu
			if pu:
				metal_purities.append(pu)
	stone_g = (out["dmd_weight"] + out["ps_weight"] + out["cs_weight"] + out["cz_weight"] + out["cvd_weight"] + out["pdmd_weight"] + out["poth_weight"]) * 0.2
	out["gross_weight"] = round(metal_g, 3)
	out["nett_weight"] = round(max(metal_g - stone_g, 0.0), 3)
	if metal_g:
		out["purity"] = round(purity_num / metal_g, 3)
	elif metal_purities:
		out["purity"] = round(sum(metal_purities) / len(metal_purities), 3)
	for k in ("dmd_weight", "ps_weight", "cs_weight", "cz_weight", "cvd_weight", "pdmd_weight", "poth_weight"):
		out[k] = round(out[k], 3)
	return out


def _plan_values(bag_bom, qty):
	"""The PLAN weight fields (gross/nett/purity/stones) = profile(bag BOM) x qty.
	Shared by the Order Bag controller (validate) and recalc."""
	mats = [{"item": r.item, "qty": r.qty, "weight": r.weight} for r in (bag_bom or [])]
	p = _profile_from_materials(mats)
	q = max(int(qty or 1), 1)
	# match the ACTUAL convention: gross = gold + stones, nett = gold (metal)
	metal = flt(p["gross_weight"])  # _profile_from_materials' "gross" is metal grams
	buckets = ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth")
	stone_g = sum(flt(p[f"{b}_weight"]) for b in buckets) * 0.2
	vals = {
		"gross_weight": round((metal + stone_g) * q, 3),
		"nett_weight": round(metal * q, 3),
		"purity": p["purity"],
	}
	for b in buckets:
		vals[f"{b}_no"] = int(p[f"{b}_no"]) * q
		vals[f"{b}_weight"] = round(p[f"{b}_weight"] * q, 3)
	return vals


def _actual_profile(order_bag):
	"""The ACTUAL weight profile from what the bag really holds (the ledger): gold
	grams, stone carats by stone_type, gram-weighted metal purity."""
	rows = [it for it in get_bag_contents(order_bag)["items"] if flt(it["qty"])]
	codes = list({it["item"] for it in rows})
	meta = {}
	if codes:
		for i in frappe.get_all("Item", filters={"name": ["in", codes]}, fields=["name", "stone_type", "purity_percentage"]):
			meta[i.name] = (i.stone_type, flt(i.purity_percentage))
	BUCKET = {"Diamond": "dmd", "Precious Stone": "ps", "Color Stone": "cs", "Cubic Zirconia": "cz", "CVD": "cvd", "Party Diamond": "pdmd", "Party Other": "poth"}
	w = {b: 0.0 for b in BUCKET.values()}
	n = {b: 0 for b in BUCKET.values()}
	gold = pnum = 0.0
	mp = []
	for it in rows:
		q = flt(it["qty"])
		pcs = int(it.get("pcs") or 0)
		st, pu = meta.get(it["item"], (None, 0.0))
		b = BUCKET.get(st)
		if b:
			w[b] += q
			n[b] += pcs
		else:
			gold += q
			pnum += q * pu
			if pu:
				mp.append(pu)
	purity = (pnum / gold) if gold else (sum(mp) / len(mp) if mp else 0.0)
	out = {"gross": round(gold + sum(w.values()) * 0.2, 3), "nett": round(gold, 3), "purity": round(purity, 3)}
	for b in BUCKET.values():
		out[f"{b}_weight"] = round(w[b], 3)
		out[f"{b}_no"] = n[b]
	return out


@frappe.whitelist()
def refresh_actual_weights(order_bag):
	"""Recompute + store the ACTUAL weight fields (act_*) from current contents,
	including real stone counts (pcs) from the ledger. Once the piece is a finished
	product its materials are consumed, so the actuals are FROZEN (just return them)."""
	act_fields = ["act_gross_weight", "act_nett_weight", "act_purity", "act_pure_weight"] + [
		f"act_{b}_{k}" for b in ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth") for k in ("weight", "no")]
	if frappe.db.get_value("Order Bag", order_bag, "is_finished"):
		return frappe.db.get_value("Order Bag", order_bag, act_fields, as_dict=True)
	p = _actual_profile(order_bag)
	vals = {
		"act_gross_weight": p["gross"], "act_nett_weight": p["nett"], "act_purity": p["purity"],
		"act_pure_weight": round(p["nett"] * p["purity"] / 100.0, 3),
	}
	for b in ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth"):
		vals[f"act_{b}_weight"] = p[f"{b}_weight"]
		vals[f"act_{b}_no"] = p[f"{b}_no"]
	frappe.db.set_value("Order Bag", order_bag, vals)
	return vals


@frappe.whitelist()
def recalc_bag_weights_from_bom(order_bag):
	"""Recompute a bag's PLAN weights from its OWN BOM x qty + refresh the actuals."""
	bag = frappe.get_doc("Order Bag", order_bag)
	vals = _plan_values(bag.bag_bom, bag.qty)
	bag.db_set(vals)
	refresh_actual_weights(order_bag)
	frappe.db.commit()
	return {"order_bag": order_bag, **vals}


@frappe.whitelist()
def get_design_profile(design):
	"""Full line profile derived from the design's BOM — used to auto-fill a Place
	Order line when a design is picked:
	  - stone counts (dmd/ps/cs no) and carat weights (dmd/ps/cs weight) by stone_type
	  - nett_weight (g) = total metal grams
	  - gross_weight (g) = metal grams + stone grams (1 ct = 0.2 g)
	  - purity (%) = gram-weighted average purity of the metal rows
	"""
	out = {
		"dmd_no": 0, "ps_no": 0, "cs_no": 0,
		"dmd_weight": 0.0, "ps_weight": 0.0, "cs_weight": 0.0,
		"gross_weight": 0.0, "nett_weight": 0.0, "purity": 0.0,
	}
	if not design or not frappe.db.exists("Design", design):
		return out

	mats = frappe.get_all(
		"Design BOM Item",
		filters={"parent": design, "parenttype": "Design"},
		fields=["item", "qty", "purity", "weight"],
	)
	return _profile_from_materials(mats)


@frappe.whitelist()
def get_design_materials(design):
	"""Raw-materials (BOM) table for a Design — for the Place Order 'Design' info dialog.
	Item attributes (name / stone type / uom / purity) are read from the Item master, since
	BOM-row copies are unreliable on programmatically seeded designs."""
	out = {"design": design, "design_type": None, "materials": []}
	if not design or not frappe.db.exists("Design", design):
		return out
	d = frappe.get_doc("Design", design)
	out["design_type"] = d.design_type
	out["design_style"] = d.design_style
	out["image"] = d.image
	for m in d.materials:
		im = frappe.db.get_value(
			"Item", m.item,
			["item_name", "stone_type", "weight_unit", "purity_percentage"], as_dict=True,
		) or {}
		out["materials"].append({
			"item": m.item,
			"item_name": im.get("item_name") or m.item,
			"stone_type": im.get("stone_type") or "",
			"uom": im.get("weight_unit") or getattr(m, "uom", "") or "",
			"purity": im.get("purity_percentage") or getattr(m, "purity", 0) or 0,
			"qty": m.qty or 0,
			"weight": m.weight or 0,
		})
	return out


def classify_item(doc):
	"""Stamp the 4-level classification (main_type / material_type / material_group) from
	the item group's ancestors in the Item Group tree: MAIN TYPE -> TYPE -> GROUP -> leaf.
	A GROUP that is itself the leaf (ALLOY, CVD, …) classifies as its own group."""
	doc.main_type = doc.material_type = doc.material_group = None
	if not doc.get("item_group") or not frappe.db.exists("Item Group", doc.item_group):
		return
	# ancestors root -> leaf (excluding "All Item Groups"), then the leaf itself
	chain = []
	node = doc.item_group
	seen = set()
	while node and node not in seen:
		seen.add(node)
		chain.append(node)
		node = frappe.db.get_value("Item Group", node, "parent_item_group")
	chain = [n for n in reversed(chain) if n != "All Item Groups"]
	# the root node may pre-exist in any case ("Raw Material" ships with ERPNext)
	if not chain or chain[0].upper() not in ("RAW MATERIAL", "PRODUCT"):
		return  # not classified under the tree (legacy flat group)
	doc.main_type = chain[0].upper()
	doc.material_type = chain[1] if len(chain) > 1 else None
	doc.material_group = chain[2] if len(chain) > 2 else (chain[1] if len(chain) > 1 else None)


def set_item_weight_uom(doc, method=None):
	"""Keep the Weight UOM unambiguous: a stone (has stone_type) is weighed in
	carats, everything else in grams. Hooked on Item validate — also stamps the
	classification fields from the Item Group tree."""
	if doc.get("stone_type"):
		doc.weight_unit = "Carat"
	elif not doc.get("weight_unit"):
		doc.weight_unit = "Gram"
	classify_item(doc)


@frappe.whitelist()
def get_raw_material_tree():
	"""The RAW MATERIAL branch of the Item Group tree with every item under its
	group — powers the read-only Raw Materials structure page. Items sort in
	registry order (O-sizes first, then 1→22.5) with anything else natsorted after."""
	import re

	groups = frappe.get_all("Item Group", fields=["name", "parent_item_group", "is_group"])
	root = next((g.name for g in groups if g.name.upper() == "RAW MATERIAL"), None)
	if not root:
		return {"tree": None, "total_groups": 0, "total_items": 0}
	by_parent = {}
	for g in groups:
		by_parent.setdefault(g.parent_item_group, []).append(g)

	# every group under the root (the root ships with ERPNext in title case)
	subtree, queue = [], [root]
	while queue:
		n = queue.pop()
		subtree.append(n)
		queue += [g.name for g in by_parent.get(n, [])]

	rows = frappe.get_all(
		"Item",
		filters={"item_group": ["in", subtree]},
		fields=["name", "item_group", "stock_uom", "stone_type", "purity_percentage", "metal_purity", "disabled", "stone_party"],
	)

	def natkey(s):
		return [(0, int(t)) if t.isdigit() else (1, t.upper()) for t in re.split(r"(\d+)", s or "") if t]

	from jewelima.jewelima.raw_materials import RAW_MATERIALS
	from jewelima.setup import GOLD_COLORS, KARAT_GOLDS

	registry_order = {code: i for i, (code, *_rest) in enumerate(RAW_MATERIALS)}
	in_registry = len(registry_order)
	# karat + standard golds ship through their own seeders — part of the base set too
	shipped = set(registry_order)
	shipped.update(f"{k}{c}" for k in KARAT_GOLDS for c in GOLD_COLORS)
	shipped.update(f"Standard Gold {n}" for n in range(990, 1000))

	items_by_group = {}
	for it in rows:
		items_by_group.setdefault(it.item_group, []).append({
			"name": it.name,
			"uom": it.stock_uom or "",
			"stone_type": it.stone_type or "",
			"purity": it.purity_percentage or 0,
			"metal_purity": it.metal_purity or "",
			"disabled": int(it.disabled or 0),
			# party stones/metals are created on demand — expected, not drift
			"in_registry": it.name in shipped or bool(it.stone_party),
		})

	def build(name):
		kids = sorted(by_parent.get(name, []), key=lambda g: natkey(g.name))
		node_items = sorted(
			items_by_group.get(name, []),
			key=lambda i: (registry_order.get(i["name"], in_registry), natkey(i["name"])),
		)
		children = [build(k.name) for k in kids]
		return {
			"name": name,
			"children": children,
			"items": node_items,
			"count": len(node_items) + sum(c["count"] for c in children),
		}

	tree = build(root)
	return {"tree": tree, "total_groups": len(subtree), "total_items": tree["count"]}


@frappe.whitelist()
def get_stone_buckets():
	"""The six stone buckets (DMD/PS/CS/CVD/PDMD/POTH) with every item that feeds
	each — powers the read-only Stone Buckets page. An item's bucket IS its
	stone_type; the order matches the Order Bag stone columns."""
	import re

	buckets = [
		("DMD", "Diamond"), ("PS", "Precious Stone"), ("CS", "Color Stone"),
		("CZ", "Cubic Zirconia"), ("CVD", "CVD"), ("PDMD", "Party Diamond"), ("POTH", "Party Other"),
	]
	rows = frappe.get_all(
		"Item",
		filters={"stone_type": ["in", [b[1] for b in buckets]]},
		fields=["name", "item_group", "stock_uom", "stone_type", "disabled"],
	)

	def natkey(s):
		return [(0, int(t)) if t.isdigit() else (1, t.upper()) for t in re.split(r"(\d+)", s or "") if t]

	from jewelima.jewelima.raw_materials import RAW_MATERIALS

	registry_order = {code: i for i, (code, *_rest) in enumerate(RAW_MATERIALS)}
	in_registry = len(registry_order)

	by_type = {}
	for it in rows:
		by_type.setdefault(it.stone_type, []).append({
			"name": it.name,
			"group": it.item_group or "",
			"uom": it.stock_uom or "",
			"disabled": int(it.disabled or 0),
			# party stones are created on demand — expected, not drift
			"in_registry": it.name in registry_order or it.stone_type in PARTY_BRACKETS,
		})

	out = []
	for code, st in buckets:
		items = sorted(by_type.get(st, []), key=lambda i: (registry_order.get(i["name"], in_registry), natkey(i["name"])))
		out.append({"code": code, "stone_type": st, "items": items, "count": len(items)})
	return {"buckets": out, "total_items": sum(b["count"] for b in out)}


# --- Party stock (customer-given stones) -----------------------------------------
# A party's 3-letter code prefixes every stone item it owns (EDI -> EDI-VS1).
# Items are created ON DEMAND (no sieve runs) and only under an existing party.
PARTY_BRACKETS = {"Party Diamond": "PARTY DIAMOND", "Party Other": "PARTY OTHER"}  # stone_type -> item group leaf


def _party_code_candidates(party_name):
	"""3-letter code candidates from the party name — first three letters first."""
	import re

	letters = re.sub(r"[^A-Z]", "", (party_name or "").upper())
	seen, out = set(), []

	def add(c):
		if len(c) == 3 and c not in seen:
			seen.add(c)
			out.append(c)

	add(letters[:3])
	for i in range(3, len(letters)):
		add(letters[:2] + letters[i])
	for i in range(1, len(letters) - 1):
		add(letters[0] + letters[i : i + 2])
	return out[:30]


@frappe.whitelist()
def suggest_party_code(party_name):
	"""First free 3-letter code for the name (EDIMINIKAL -> EDI; EDI taken -> EDM …)."""
	for c in _party_code_candidates(party_name):
		if not frappe.db.exists("Stone Party", c):
			return c
	return ""


@frappe.whitelist()
def create_stone_party(party_name, code):
	"""Create a Stone Party — the doctype's validate enforces the 3-letter rule."""
	doc = frappe.get_doc({"doctype": "Stone Party", "party_name": party_name, "code": code})
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def get_stone_parties():
	"""All parties with how many stone items each owns."""
	parties = frappe.get_all("Stone Party", fields=["name", "party_name"], order_by="party_name")
	counts = dict(frappe.db.sql(
		"select stone_party, count(*) from `tabItem` where ifnull(stone_party, '') != '' group by stone_party"
	))
	return [{"code": p.name, "party_name": p.party_name, "items": counts.get(p.name, 0)} for p in parties]


@frappe.whitelist()
def get_party_stones(party):
	"""The stone items a party owns, newest first (metals live on their own page)."""
	rows = frappe.get_all(
		"Item", filters={"stone_party": party, "stone_type": ["in", list(PARTY_BRACKETS)]},
		fields=["name", "stone_type", "disabled"], order_by="creation desc",
	)
	return [{"name": r.name, "bracket": "PDMD" if r.stone_type == "Party Diamond" else "POTH",
	         "stone_type": r.stone_type, "disabled": int(r.disabled or 0)} for r in rows]


# --- Order number reservations ------------------------------------------------------
# The Place Order page CLAIMS its E#### the moment it opens, so the number is
# known before placing. Rules: a user re-opening the page gets their own unused
# claim back (no burn); claims idle past the recycle window return to a pool and
# new sessions drain the pool OLDEST FIRST (gaps fill later); otherwise a fresh
# number is minted off the same E series. Placing consumes the claim atomically —
# if it was recycled away in the meantime, placement self-heals with a fresh one.
ORDER_NO_RECYCLE_HOURS = 2


def _mint_order_no():
	from frappe.model.naming import make_autoname

	return make_autoname("E.####")  # the Job Order series — one counter for both paths


@frappe.whitelist()
def reserve_order_no():
	"""Claim an order number for this session (called when Place Order opens)."""
	me = frappe.session.user
	now = frappe.utils.now_datetime()
	cutoff = frappe.utils.add_to_date(now, hours=-ORDER_NO_RECYCLE_HOURS)

	# 1. my own unused claim — same number every time I come back
	mine = frappe.get_all("Order No Reservation", filters={"reserved_by": me},
	                      order_by="reserved_on desc", pluck="name", limit=1)
	if mine:
		frappe.db.set_value("Order No Reservation", mine[0], "reserved_on", now, update_modified=False)
		frappe.db.commit()
		return mine[0]

	# 2. the pool: abandoned claims, oldest first (this is what fills the gaps)
	for no in frappe.get_all("Order No Reservation", filters={"reserved_on": ["<", cutoff]},
	                         order_by="reserved_on asc", pluck="name", limit=5):
		frappe.db.sql(
			"""update `tabOrder No Reservation` set reserved_by = %s, reserved_on = %s
			   where name = %s and reserved_on < %s""",
			(me, now, no, cutoff),
		)
		if frappe.db._cursor.rowcount:  # atomic — only one session wins a recycled number
			frappe.db.commit()
			return no

	# 3. mint fresh off the series
	no = _mint_order_no()
	frappe.get_doc({
		"doctype": "Order No Reservation", "order_no": no, "reserved_by": me, "reserved_on": now,
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return no


@frappe.whitelist()
def create_job_order(payload):
	"""Create the Job Order under the session's reserved number (autoname would
	override an explicit name client-side). If the claim was recycled away — or
	the number somehow exists — self-heal with a freshly minted one."""
	p = frappe.parse_json(payload)
	no = (p.get("order_no") or "").strip()
	claimed = False
	if no:
		frappe.db.sql("delete from `tabOrder No Reservation` where name = %s", no)
		claimed = bool(frappe.db._cursor.rowcount)
	if not claimed or not no or frappe.db.exists("Job Order", no):
		no = _mint_order_no()
	doc = frappe.get_doc({
		"doctype": "Job Order",
		"order_date": p.get("order_date") or frappe.utils.today(),
		"due_date": p.get("due_date") or None,
		"customer_date": p.get("customer_date") or None,
		"customer": p.get("customer") or None,
		"salesman": p.get("salesman") or None,
		"order_type": p.get("order_type") or None,
	})
	doc.insert(ignore_permissions=True, set_name=no)
	frappe.db.commit()
	return doc.name


# --- Order Requests + repeat orders ------------------------------------------------
# A request = a saved Place Order page (header + basic lines) that hasn't gone
# through. Any base-role user files one; the Order User pulls it up on the page
# (Requests -> Use) and places it — which stamps the Job Order back on the request.


@frappe.whitelist()
def save_order_request(payload):
	"""File an Order Request from the Order Requests page. payload = {customer,
	salesman, order_type, days, cust_days, notes, lines}; a line is either
	{design, qty, size, remark, edited, materials} or {cad: {...}, qty, size, remark}
	— edited BOMs and CAD wishes travel whole so Use restores them 1:1."""
	import json as _json

	p = frappe.parse_json(payload)
	lines = [l for l in (p.get("lines") or []) if l.get("design") or l.get("cad")]
	if not lines:
		frappe.throw(_("Add at least one line with a Design (or a CAD line)."))
	items = []
	for l in lines:
		items.append({
			"design": l.get("design"),
			"qty": cint(l.get("qty")) or 1,
			"size": l.get("size"),
			"remark": l.get("remark"),
			"edited": 1 if l.get("edited") else 0,
			"bom_json": _json.dumps(l["materials"]) if l.get("edited") and l.get("materials") else None,
			"cad_json": _json.dumps(l["cad"]) if l.get("cad") else None,
		})
	doc = frappe.get_doc({
		"doctype": "Order Request",
		"customer": p.get("customer"),
		"salesman": p.get("salesman"),
		"order_type": p.get("order_type"),
		"days": cint(p.get("days")),
		"cust_days": cint(p.get("cust_days")),
		"notes": p.get("notes"),
		"items": items,
	})
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def get_order_requests():
	"""Open requests, newest first, with a line summary for the picker dialog."""
	reqs = frappe.get_all(
		"Order Request", filters={"status": "Open"},
		fields=["name", "request_date", "requested_by", "customer", "notes"],
		order_by="creation desc", limit=50,
	)
	out = []
	for r in reqs:
		items = frappe.get_all("Order Request Item", filters={"parent": r.name},
		                       fields=["design", "qty"], order_by="idx")
		out.append({
			"name": r.name,
			"request_date": r.request_date,
			"requested_by": frappe.utils.get_fullname(r.requested_by),
			"customer": r.customer or "",
			"notes": r.notes or "",
			"lines": len(items),
			"qty": sum(cint(i.qty) or 0 for i in items),
			"designs": ", ".join((i.design or "CAD") for i in items[:4]) + ("…" if len(items) > 4 else ""),
		})
	return out


@frappe.whitelist()
def get_order_request(name):
	"""Full payload of one request — what the page needs to fill itself
	(edited BOMs and CAD lines come back parsed, ready to restore)."""
	import json as _json

	def loads(s):
		try:
			return _json.loads(s) if s else None
		except Exception:
			return None

	doc = frappe.get_doc("Order Request", name)
	return {
		"name": doc.name,
		"status": doc.status,
		"customer": doc.customer, "salesman": doc.salesman, "order_type": doc.order_type,
		"days": doc.days, "cust_days": doc.cust_days, "notes": doc.notes,
		"lines": [{
			"design": i.design, "qty": i.qty, "size": i.size, "remark": i.remark,
			"edited": cint(i.edited), "materials": loads(i.bom_json) or [],
			"cad": loads(i.cad_json),
		} for i in doc.items],
	}


@frappe.whitelist()
def get_all_order_requests(status=None, customer=None, order_type=None, salesman=None, design=None):
	"""Every request, filterable — the review board (Setup > Order Setup > All
	Requests). The design filter matches requests CONTAINING that design."""
	filters = {}
	if status and status != "All":
		filters["status"] = status
	for k, v in (("customer", customer), ("order_type", order_type), ("salesman", salesman)):
		if v:
			filters[k] = v
	if design:
		parents = frappe.get_all("Order Request Item", filters={"design": design}, pluck="parent", distinct=True)
		if not parents:
			return []
		filters["name"] = ["in", parents]
	reqs = frappe.get_all(
		"Order Request", filters=filters,
		fields=["name", "request_date", "requested_by", "customer", "order_type", "salesman", "status", "job_order", "notes"],
		order_by="creation desc", limit=200,
	)
	out = []
	for r in reqs:
		items = frappe.get_all("Order Request Item", filters={"parent": r.name},
		                       fields=["design", "qty"], order_by="idx")
		out.append({
			"name": r.name, "request_date": r.request_date,
			"requested_by": frappe.utils.get_fullname(r.requested_by),
			"customer": r.customer or "", "order_type": r.order_type or "", "salesman": r.salesman or "",
			"status": r.status, "job_order": r.job_order or "", "notes": r.notes or "",
			"lines": len(items),
			"qty": sum(cint(i.qty) or 0 for i in items),
			"designs": ", ".join((i.design or "CAD") for i in items[:4]) + ("…" if len(items) > 4 else ""),
		})
	return out


@frappe.whitelist()
def get_my_order_requests():
	"""The session user's recent requests (any status) — the Order Requests page's
	'My Requests' list, so requesters can see when theirs got placed."""
	reqs = frappe.get_all(
		"Order Request", filters={"requested_by": frappe.session.user},
		fields=["name", "request_date", "customer", "status", "job_order"],
		order_by="creation desc", limit=20,
	)
	for r in reqs:
		r["lines"] = frappe.db.count("Order Request Item", {"parent": r.name})
	return reqs


@frappe.whitelist()
def delete_order_request(name):
	"""Delete a request from the My Requests list. The controller's on_trash
	enforces the rules: never once Placed, and only by the requester (or SM)."""
	doc = frappe.get_doc("Order Request", name)
	doc.delete(ignore_permissions=True)
	frappe.db.commit()


@frappe.whitelist()
def mark_order_request_placed(name, job_order):
	"""Stamp a request once its order went through (called by the Place Order page)."""
	if not frappe.db.exists("Order Request", name):
		return
	frappe.db.set_value("Order Request", name, {"status": "Placed", "job_order": job_order})
	frappe.db.commit()


@frappe.whitelist()
def get_job_order_fill(job_order):
	"""Repeat an order: header + lines of an existing Job Order, shaped for the page.
	CAD bags are skipped (no design to copy) and reported."""
	if not frappe.db.exists("Job Order", job_order):
		frappe.throw(_("Job Order {0} not found.").format(job_order))
	jo = frappe.db.get_value("Job Order", job_order,
	                         ["customer", "salesman", "order_type"], as_dict=True)
	bags = frappe.get_all(
		"Order Bag", filters={"job_order": job_order},
		fields=["design", "qty", "size", "narration", "is_cad"], order_by="creation",
	)
	lines, skipped_cad = [], 0
	for b in bags:
		if not b.design:
			skipped_cad += 1 if b.is_cad else 0
			continue
		lines.append({"design": b.design, "qty": cint(b.qty) or 1, "size": b.size, "remark": b.narration})
	return {"customer": jo.customer, "salesman": jo.salesman, "order_type": jo.order_type,
	        "lines": lines, "skipped_cad": skipped_cad}


# --- Party metal (customer-given gold) --------------------------------------------
# STRICT naming: the suffix must be one of OUR gold codes — the 9 karat golds
# (JOS-22KYG) or a standard gold (JOS-Standard999). Grouped under METAL -> PARTY
# METAL, outside the GOLD branch, so karat/melt pickers never mix party gold in.
PARTY_METAL_GROUP = "PARTY METAL"


def _party_metal_options():
	"""suffix -> {label, purity, metal_purity}; karats first, then standards."""
	from jewelima.setup import GOLD_COLORS, KARAT_GOLDS

	out = {}
	for karat, purity in KARAT_GOLDS.items():
		for color in GOLD_COLORS:
			out[f"{karat}{color}"] = {"kind": "Karat", "purity": purity, "metal_purity": karat}
	for n in range(990, 1000):
		out[f"Standard{n}"] = {"kind": "Standard", "purity": round(n / 10.0, 2), "metal_purity": ""}
	return out


@frappe.whitelist()
def get_party_metal_options():
	"""The allowed metal suffixes for the Party Metal Add page."""
	return [{"suffix": s, **m} for s, m in _party_metal_options().items()]


@frappe.whitelist()
def get_party_metals(party):
	"""The metal items a party owns, newest first."""
	rows = frappe.get_all(
		"Item", filters={"stone_party": party, "item_group": PARTY_METAL_GROUP},
		fields=["name", "metal_purity", "purity_percentage", "disabled"], order_by="creation desc",
	)
	return [{"name": r.name, "metal_purity": r.metal_purity or "", "purity": r.purity_percentage or 0,
	         "disabled": int(r.disabled or 0)} for r in rows]


@frappe.whitelist()
def check_party_metal(party, metal):
	"""Preview + availability of the item code a party/metal pair would create."""
	if not frappe.db.exists("Stone Party", party):
		frappe.throw(_("Create the party first — metal can only come in under a party."))
	if metal not in _party_metal_options():
		frappe.throw(_("Metal must be one of the standard gold codes."))
	code = f"{party}-{metal}"
	return {"item_code": code, "exists": bool(frappe.db.exists("Item", code))}


@frappe.whitelist()
def create_party_metal(party, metal):
	"""Create ONE party gold item on demand: <CODE>-<METAL> where METAL follows OUR
	standard exactly (22KYG… / Standard990…999). Purity derived, Gram, no stone
	bucket — it's metal."""
	chk = check_party_metal(party, metal)
	if chk["exists"]:
		frappe.throw(_("{0} already exists.").format(chk["item_code"]))
	if not frappe.db.exists("Item Group", PARTY_METAL_GROUP):
		frappe.throw(_("Item group {0} is missing — run a migrate.").format(PARTY_METAL_GROUP))
	m = _party_metal_options()[metal]
	frappe.get_doc({
		"doctype": "Item",
		"item_code": chk["item_code"],
		"item_name": chk["item_code"],
		"item_group": PARTY_METAL_GROUP,
		"stock_uom": "Gram",
		"is_stock_item": 1,
		"is_purchase_item": 0,  # customer-given, never purchased
		"is_sales_item": 0,
		"include_item_in_manufacturing": 1,
		"weight_unit": "Gram",
		"metal_purity": m["metal_purity"],
		"purity_percentage": m["purity"],
		"stone_party": party,
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return chk["item_code"]


@frappe.whitelist()
def check_party_stone(party, stone):
	"""Preview + availability of the item code a party/stone pair would create."""
	code = _party_stone_code(party, stone)
	return {"item_code": code, "exists": bool(frappe.db.exists("Item", code))}


def _party_stone_code(party, stone):
	import re

	if not frappe.db.exists("Stone Party", party):
		frappe.throw(_("Create the party first — stones can only be added under a party."))
	s = re.sub(r"\s+", " ", (stone or "").strip().upper())
	if not re.fullmatch(r"[A-Z0-9][A-Z0-9 .\-/]*", s):
		frappe.throw(_("Stone name: letters/numbers (and . - /) only, e.g. VS1 or RUBY."))
	return f"{party}-{s}"


@frappe.whitelist()
def create_party_stone(party, bracket, stone):
	"""Create ONE party stone item on demand: <CODE>-<STONE> under the party-bracket
	group, typed so it tallies into the PDMD/POTH bucket. Everything derived —
	the form only supplies party, bracket and the stone name."""
	if bracket not in PARTY_BRACKETS:
		frappe.throw(_("Bracket must be Party Diamond or Party Other."))
	code = _party_stone_code(party, stone)
	if frappe.db.exists("Item", code):
		frappe.throw(_("{0} already exists.").format(code))
	group = PARTY_BRACKETS[bracket]
	if not frappe.db.exists("Item Group", group):
		frappe.throw(_("Item group {0} is missing — run a migrate.").format(group))
	frappe.get_doc({
		"doctype": "Item",
		"item_code": code,
		"item_name": code,
		"item_group": group,
		"stock_uom": "Carat",
		"is_stock_item": 1,
		"is_purchase_item": 0,  # customer-given, never purchased
		"is_sales_item": 0,
		"include_item_in_manufacturing": 1,
		"stone_type": bracket,
		"weight_unit": "Carat",
		"stone_party": party,
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return code


@frappe.whitelist()
def get_item_stone_profile(item):
	"""Given a finished item, read its (default) BOM and tally the stone counts by
	stone type — Diamond -> dmd_no, Precious Stone -> ps_no, Color Stone -> cs_no.
	Used by the Place Order grid to auto-fill a line when an item is picked.

	Stone counts come from the BOM components whose Item has a `stone_type` set
	(uses the kept Item stone custom-fields). Weights aren't derivable from a plain
	BOM line, so they stay 0 here — that's the case for a richer Design master."""
	out = {
		"item_name": None, "bom": None,
		"dmd_no": 0, "dmd_weight": 0, "ps_no": 0, "ps_weight": 0, "cs_no": 0, "cs_weight": 0,
	}
	if not item:
		return out

	out["item_name"] = frappe.db.get_value("Item", item, "item_name")

	bom = frappe.db.get_value("BOM", {"item": item, "is_default": 1, "is_active": 1}, "name") or \
		frappe.db.get_value("BOM", {"item": item, "is_active": 1}, "name")
	out["bom"] = bom
	if not bom:
		return out

	rows = frappe.get_all("BOM Item", filters={"parent": bom}, fields=["item_code", "qty"])
	if not rows:
		return out

	bucket = {"Diamond": "dmd_no", "Precious Stone": "ps_no", "Color Stone": "cs_no"}
	codes = list({r.item_code for r in rows})
	stone_type = {
		i.name: i.stone_type
		for i in frappe.get_all("Item", filters={"name": ["in", codes]}, fields=["name", "stone_type"])
	}
	for r in rows:
		st = stone_type.get(r.item_code)
		if st in bucket:
			out[bucket[st]] += (r.qty or 0)
	return out


@frappe.whitelist()
def get_order_bag_images(order_bag):
	"""All image files for an Order Bag: native File attachments + the Attachments
	child-table images, deduped, image extensions only."""
	if not order_bag or not frappe.db.exists("Order Bag", order_bag):
		return []
	IMG = (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp")
	files = frappe.get_all(
		"File",
		filters={"attached_to_doctype": "Order Bag", "attached_to_name": order_bag},
		fields=["file_url", "file_name"],
		order_by="creation desc",
	)
	# owner map: only File-backed attachments on THIS bag can be deleted, by the
	# user who added them (File.owner) or a System Manager
	me = frappe.session.user
	is_admin = "System Manager" in set(frappe.get_roles())
	owner_by_url = {f.file_url: f.owner for f in frappe.get_all("File",
		filters={"attached_to_doctype": "Order Bag", "attached_to_name": order_bag},
		fields=["file_url", "owner"])}
	bag = frappe.get_doc("Order Bag", order_bag)
	if bag.image:
		files.insert(0, {"file_url": bag.image, "file_name": bag.design or "design"})  # the bag's held design photo, first
	for a in bag.attachments or []:
		if a.image:
			files.append({"file_url": a.image, "file_name": a.title or a.image.split("/")[-1]})
	out, seen = [], set()
	for f in files:
		url = f.get("file_url")
		if not url or url in seen:
			continue
		if not url.lower().split("?")[0].endswith(IMG):
			continue
		seen.add(url)
		owner = owner_by_url.get(url)
		out.append({"file_url": url, "file_name": f.get("file_name") or url.split("/")[-1],
			"owner": owner, "can_delete": 1 if (owner and (is_admin or owner == me)) else 0})
	return out


@frappe.whitelist()
def delete_order_bag_image(order_bag, file_url):
	"""Delete a card photo — ONLY the File-backed attachment the current user added
	(or any, for a System Manager). The design image and files added by others are
	refused."""
	me = frappe.session.user
	is_admin = "System Manager" in set(frappe.get_roles())
	rec = frappe.get_all("File", filters={"attached_to_doctype": "Order Bag",
		"attached_to_name": order_bag, "file_url": file_url}, fields=["name", "owner"], limit=1)
	if not rec:
		frappe.throw(frappe._("That image isn't a removable attachment on this card."))
	if not (is_admin or rec[0].owner == me):
		frappe.throw(frappe._("You can only delete images you added."))
	bag = frappe.get_doc("Order Bag", order_bag)
	kept = [a for a in bag.attachments if a.image != file_url]
	if len(kept) != len(bag.attachments):
		bag.set("attachments", kept)
		bag.save(ignore_permissions=True)
	frappe.delete_doc("File", rec[0].name, force=True, ignore_permissions=True)
	frappe.db.commit()
	return {"deleted": file_url}


@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def bench_employee_query(doctype, txt, searchfield, start, page_len, filters):
	"""Link query for the bench employee pickers — only employees allotted to the bench
	(filters['bench']). Falls back to all active employees if that bench has no roster yet;
	pass filters['strict'] to return NOBODY instead (roster-only, e.g. Make Tree)."""
	bench = (filters or {}).get("bench")
	strict = frappe.utils.cint((filters or {}).get("strict"))
	roster = []
	if bench and frappe.db.exists("Bench", bench):
		roster = frappe.get_all("Bench Employee", filters={"parent": bench}, pluck="employee")
	like = f"%{txt or ''}%"
	if roster:
		ph = ", ".join(["%s"] * len(roster))
		return frappe.db.sql(
			f"""SELECT name, employee_name FROM `tabEmployee`
			    WHERE status = 'Active' AND (name LIKE %s OR employee_name LIKE %s) AND name IN ({ph})
			    ORDER BY employee_name LIMIT %s, %s""",
			[like, like, *roster, start, page_len],
		)
	if strict:
		return []  # roster-only: an empty Bench roster means no one to pick
	return frappe.db.sql(
		"""SELECT name, employee_name FROM `tabEmployee`
		   WHERE status = 'Active' AND (name LIKE %s OR employee_name LIKE %s)
		   ORDER BY employee_name LIMIT %s, %s""",
		[like, like, start, page_len],
	)


@frappe.whitelist()
def get_order_bag_cards(names):
	"""Print-card data for a list of Order Bags: the bag's own fields plus its
	design's type/image and BOM materials. Used by the Print Order Bags page."""
	if isinstance(names, str):
		names = json.loads(names or "[]")
	cards = []
	for nm in names or []:
		if not frappe.db.exists("Order Bag", nm):
			continue
		b = frappe.get_doc("Order Bag", nm)
		dtype = dstyle = dimg = ""
		materials = []
		if b.design and frappe.db.exists("Design", b.design):
			d = frappe.get_doc("Design", b.design)
			dtype, dstyle, dimg = d.design_type, d.design_style, d.image
			for m in d.materials:
				materials.append({"item": m.item, "purity": m.purity, "qty": m.qty, "weight": m.weight, "uom": m.uom})
		cards.append({
			"name": b.name, "job_order": b.job_order, "design": b.design,
			"design_type": dtype, "design_style": dstyle, "image": dimg or b.image,
			"size": b.size, "qty": b.qty, "location": b.location,
			"customer": b.customer, "salesman": b.salesman, "order_type": b.order_type,
			"order_date": frappe.utils.formatdate(b.order_date, "dd-mm-yyyy") if b.order_date else "",
			"due_date": frappe.utils.formatdate(b.due_date, "dd-mm-yyyy") if b.due_date else "",
			"gross_weight": b.gross_weight, "nett_weight": b.nett_weight, "purity": b.purity,
			"dmd_no": b.dmd_no, "dmd_weight": b.dmd_weight, "ps_no": b.ps_no, "ps_weight": b.ps_weight,
			"cs_no": b.cs_no, "cs_weight": b.cs_weight, "narration": b.narration,
			"materials": materials,
			"is_cad": int(b.is_cad or 0), "cad_design_type": b.cad_design_type, "cad_karat": b.cad_karat,
			"cad_gold_weight": b.cad_gold_weight, "cad_diamond_weight": b.cad_diamond_weight,
			"cad_stone_no": b.cad_stone_no, "cad_reference": b.cad_reference, "cad_remarks": b.cad_remarks,
		})
	return cards


def _all_locations():
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	return list(BENCH_DOCTYPE.keys())


def _transfer_allowed(roles, from_location, to_location):
	"""Role-based from->to permission, PER-ROLE dormancy: a rule only binds the
	roles that appear in the rules table (the Transfer Matrix). A user whose roles
	have no rules at all is unrestricted — so painting the matrix for the transfer
	roles never locks out Stock Managers etc. If any of the user's roles IS in the
	matrix, the union of those roles' rules decides. System Manager always allowed;
	a rule with a blank from/to = wildcard."""
	if "System Manager" in roles:
		return True
	rules = frappe.get_all("Transfer Rule", fields=["role", "from_location", "to_location"])
	ruled = {r.role for r in rules}
	mine = roles & ruled
	if not mine:
		return True   # none of my roles is governed by the matrix
	for r in rules:
		if r.role in mine and (not r.from_location or r.from_location == from_location) and (not r.to_location or r.to_location == to_location):
			return True
	return False


@frappe.whitelist()
def allowed_to_locations(from_location):
	"""Destinations the current user may transfer to from `from_location` (for the
	page's dropdown). Same per-role dormancy as _transfer_allowed."""
	roles = set(frappe.get_roles())
	all_locs = _all_locations()
	if "System Manager" in roles:
		return all_locs
	rules = frappe.get_all("Transfer Rule", fields=["role", "from_location", "to_location"])
	mine = roles & {r.role for r in rules}
	if not mine:
		return all_locs
	allowed = set()
	for r in rules:
		if r.role in mine and (not r.from_location or r.from_location == from_location):
			allowed.update(all_locs if not r.to_location else [r.to_location])
	return [loc for loc in all_locs if loc in allowed]


# --- Transfer Matrix (Setup) — paint which from->to moves each transfer role may make
TRANSFER_MATRIX_ROLES = ("Jewelima Transfer", "Jewelima Transfer Plus")


@frappe.whitelist()
def get_transfer_matrix():
	"""The whole board: locations, the governed roles, and each role's allowed
	from->to pairs (straight from Transfer Rule)."""
	frappe.only_for(("System Manager",))
	matrix = {role: [] for role in TRANSFER_MATRIX_ROLES}
	for r in frappe.get_all("Transfer Rule", fields=["role", "from_location", "to_location"]):
		if r.role in matrix:
			matrix[r.role].append([r.from_location or "", r.to_location or ""])
	return {"locations": _all_locations(), "roles": list(TRANSFER_MATRIX_ROLES), "matrix": matrix}


@frappe.whitelist()
def save_transfer_matrix(role, pairs):
	"""Replace ONE role's rules with the checked cells. `pairs` = [[from,to],...].
	Empty = the role has no rules = it goes dormant (unrestricted) again."""
	frappe.only_for(("System Manager",))
	if role not in TRANSFER_MATRIX_ROLES:
		frappe.throw(frappe._("Only the transfer roles live on this matrix."))
	if isinstance(pairs, str):
		pairs = json.loads(pairs or "[]")
	locs = set(_all_locations())
	frappe.db.delete("Transfer Rule", {"role": role})
	n = 0
	for p in pairs or []:
		f, t = (p[0] or "").strip(), (p[1] or "").strip()
		if f not in locs or t not in locs or f == t:
			continue
		frappe.get_doc({"doctype": "Transfer Rule", "role": role,
			"from_location": f, "to_location": t}).insert(ignore_permissions=True)
		n += 1
	frappe.db.commit()
	return {"role": role, "rules": n}


@frappe.whitelist()
def get_bag_stage_history(order_bag):
	"""Every bench this bag passed through, chronologically: who worked it, status,
	times, weight in/out and loss. Aggregated across the per-bench doctypes."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	rows = []
	for dt in dict.fromkeys(BENCH_DOCTYPE.values()):
		if not frappe.db.exists("DocType", dt):
			continue
		for r in frappe.get_all(
			dt, filters={"order_bag": order_bag},
			fields=["name", "status", "work_type", "collection_state", "employee", "time_in", "time_out", "transferred_at", "issued_at", "receipted_at", "weight_out", "weight_in", "loss", "creation"],
		):
			r["bench"] = dt
			rows.append(r)
	rows.sort(key=lambda x: (x.get("time_in") or x.get("creation")))
	emps = list({r["employee"] for r in rows if r.get("employee")})
	names = {}
	if emps:
		names = {e.name: e.employee_name for e in frappe.get_all("Employee", filters={"name": ["in", emps]}, fields=["name", "employee_name"])}
	for r in rows:
		r["employee_name"] = names.get(r.get("employee")) or r.get("employee") or ""
	return rows


@frappe.whitelist()
def get_bag_transfer_info(order_bag):
	"""Scan lookup for the Transfer page: location/design/qty/due plus the gross &
	nett weight the card ACTUALLY holds right now (from the ledger; 0 if empty)."""
	bag = frappe.db.get_value("Order Bag", order_bag, ["location", "design", "qty", "due_date"], as_dict=True)
	if not bag:
		return {}
	p = _actual_profile(order_bag)  # everything from what the bag ACTUALLY holds
	out = {
		"location": bag.location, "design": bag.design, "qty": bag.qty, "due_date": bag.due_date,
		"gross": p["gross"], "nett": p["nett"],
	}
	for bk in ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth"):
		out[bk + "_weight"] = p[bk + "_weight"]
		out[bk + "_no"] = p[bk + "_no"]
	return out


@frappe.whitelist()
def transfer_order_bag(order_bag, to_location, remarks=None):
	"""The ONLY way an Order Bag changes location: records an Order Bag Transfer
	(from -> to, time, who) and updates the bag's read-only location."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Order Bag {0} not found.").format(order_bag))
	bag = frappe.get_doc("Order Bag", order_bag)
	from_location = bag.location or ""
	if from_location == to_location:
		frappe.throw(frappe._("{0} is already at {1}.").format(order_bag, to_location))
	# CAD gate: an is_cad bag's ONLY legal move is INTO CAD; once there it's pinned until a
	# real design is assigned (finalize clears the flag).
	if bag.is_cad:
		if from_location == "CAD":
			frappe.throw(frappe._("{0} is a CAD job — finalize its design at CAD before moving it on.").format(order_bag))
		if to_location != "CAD":
			frappe.throw(frappe._("{0} is a CAD job — it must go to CAD first.").format(order_bag))
	if not _transfer_allowed(set(frappe.get_roles()), from_location, to_location):
		frappe.throw(frappe._("You don't have permission to move {0} from {1} to {2}.").format(order_bag, from_location or "—", to_location))
	t = frappe.get_doc({
		"doctype": "Order Bag Transfer",
		"order_bag": order_bag,
		"from_location": from_location or None,
		"to_location": to_location,
		"transfer_time": frappe.utils.now_datetime(),
		"transferred_by": frappe.session.user,
		"remarks": remarks,
	}).insert(ignore_permissions=True)
	bag.db_set("location", to_location)
	# the card is leaving its current bench — close out that bench's record.
	# Work that was properly finished stays Completed and gets transferred_at
	# stamped (Completed -> transfer gap = how long the finished card waited);
	# anything still open (In Queue / Issued / ...) becomes EXPIRED — the card
	# left without the work completing, and reports can filter that out.
	try:
		from jewelima.jewelima.benches import bench_doctype, on_bag_arrival

		fdt = bench_doctype(from_location)
		if fdt and frappe.db.exists("DocType", fdt):
			now = frappe.utils.now_datetime()
			open_rec = frappe.get_all(fdt, filters={"order_bag": order_bag,
				"status": ["not in", ["Completed", "Expired"]]},
				order_by="creation desc", limit=1, pluck="name")
			if open_rec:
				frappe.db.set_value(fdt, open_rec[0], {"status": "Expired", "transferred_at": now})
			else:
				done = frappe.get_all(fdt, filters={"order_bag": order_bag, "status": "Completed",
					"transferred_at": ["is", "not set"]}, order_by="creation desc", limit=1, pluck="name")
				if done:
					frappe.db.set_value(fdt, done[0], "transferred_at", now)
		# create the destination bench's record (only if that bench doctype exists yet)
		on_bag_arrival(order_bag, to_location)
	except Exception:
		frappe.log_error(frappe.get_traceback(), "bench record transition failed")
	frappe.db.commit()
	return {"transfer": t.name, "from_location": from_location, "to_location": to_location}


@frappe.whitelist()
def get_bag_contents(order_bag):
	"""Net materials currently held by an Order Bag (summed from the Bag Material
	Ledger) plus the gross weight = gold grams + stone carats * 0.2."""
	out = {"items": [], "gold_grams": 0.0, "stone_carats": 0.0, "gross_weight": 0.0}
	if not order_bag:
		return out
	rows = frappe.get_all(
		"Bag Material Ledger",
		filters={"order_bag": order_bag},
		fields=["item", "uom", "direction", "qty", "pcs"],
	)
	net = {}
	for r in rows:
		sign = 1 if (r.direction or "In") == "In" else -1
		e = net.setdefault(r.item, {"uom": r.uom or "", "qty": 0.0, "pcs": 0})
		e["qty"] += sign * flt(r.qty)
		e["pcs"] += sign * int(r.pcs or 0)
	for item, e in net.items():
		qty = round(e["qty"], 3)
		if abs(qty) < 0.0005:
			continue
		out["items"].append({"item": item, "uom": e["uom"], "qty": qty, "pcs": int(e["pcs"])})
		if e["uom"] == "Carat":
			out["stone_carats"] += qty
		else:
			out["gold_grams"] += qty
	out["gold_grams"] = round(out["gold_grams"], 3)
	out["stone_carats"] = round(out["stone_carats"], 3)
	out["gross_weight"] = round(out["gold_grams"] + out["stone_carats"] * 0.2, 3)
	return out


def _bag_ledger(order_bag, item, direction, qty, entry_type, bench=None, employee=None, remarks=None, reference=None, pcs=0):
	"""Write one Bag Material Ledger row (the per-bag material truth). `pcs` = stone
	count for the line (0 for metal)."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Order Bag {0} not found.").format(order_bag))
	if not item or not frappe.db.exists("Item", item):
		frappe.throw(frappe._("Item {0} not found.").format(item))
	qty = flt(qty)
	if qty <= 0 and not (entry_type == "Adjustment" and cint(pcs) > 0):
		# Adjustment rows may carry ONLY a piece-count correction (Stone Audit)
		frappe.throw(frappe._("Weight / qty must be greater than zero."))
	doc = frappe.get_doc({
		"doctype": "Bag Material Ledger",
		"order_bag": order_bag, "item": item, "direction": direction, "qty": qty, "pcs": int(pcs or 0),
		"entry_type": entry_type, "bench": bench or None, "employee": employee or None,
		"datetime": frappe.utils.now_datetime(), "reference": reference, "remarks": remarks,
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def add_weight(order_bag, item, qty, bench=None, remarks=None):
	"""Give gold (grams) to a bag — the Casting 'add weight' action. Records the
	per-bag ledger row AND moves the gold as real stock: Store -> In Bags pool."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE, RAW_MATERIALS_STORE

	mi = _material_issue_record("Metal", order_bag, _wh(RAW_MATERIALS_STORE),
		items=[{"item": item, "pcs": 0, "qty": flt(qty), "uom": "Gram"}])
	name = _bag_ledger(order_bag, item, "In", qty, "Gold Issue", bench=bench, remarks=remarks, reference=mi.name)
	_stock_move(item, qty, _wh(RAW_MATERIALS_STORE), _wh(IN_PRODUCTION_WAREHOUSE))
	return {"ledger": name, **get_bag_contents(order_bag)}


@frappe.whitelist()
def issue_stones(order_bag, item, qty, pcs=0, bench=None, remarks=None, from_warehouse=None):
	"""Issue stones into a bag (qty = carats, pcs = number of stones) — before work.
	Moves real stock too: source warehouse (default Stone Issue) -> In Bags pool."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE, STONE_ISSUE_WAREHOUSE

	src = from_warehouse if (from_warehouse and frappe.db.exists("Warehouse", from_warehouse)) else _wh(STONE_ISSUE_WAREHOUSE)
	name = _bag_ledger(order_bag, item, "In", qty, "Stone Issue", bench=bench, remarks=remarks, pcs=pcs)
	_stock_move(item, flt(qty), src, _wh(IN_PRODUCTION_WAREHOUSE))
	return {"ledger": name, **get_bag_contents(order_bag)}


@frappe.whitelist()
def get_stone_issue_card(barcode):
	"""Stone Issue station: resolve a scanned card into its BOM's STONE lines with
	plan / already-issued / available-at-Stone-Issue numbers. Metals never show here."""
	from jewelima.setup import STONE_ISSUE_WAREHOUSE

	nm = (barcode or "").strip()
	if not frappe.db.exists("Order Bag", nm):
		return {"error": "not_found", "card": nm or "?",
			"message": frappe._("Card {0} does not exist — check the number and scan again.").format(nm or "?")}
	bag = frappe.get_doc("Order Bag", nm)
	if bag.is_finished or bag.stock_status != "In Production":
		# spell out WHY the card can't take stones (sold / product / cancelled / …)
		st = bag.stock_status or "?"
		if st == "Sold":
			why, code = frappe._("{0} is SOLD — no more stones go into it.").format(nm), "sold"
		elif st == "Cancelled":
			why, code = frappe._("{0} is CANCELLED — nothing gets issued to a cancelled card.").format(nm), "cancelled"
		elif bag.is_finished:
			why, code = frappe._("{0} is already a finished product ({1}) — stones only go into cards on the floor.").format(nm, st), "product"
		else:
			why, code = frappe._("{0} is {1} — stones only go into cards still In Production.").format(nm, st), "status"
		return {"error": code, "card": nm, "status": st, "message": why}
	wh = _wh(STONE_ISSUE_WAREHOUSE)

	# already issued through THIS station (entry_type Stone Issue), per item
	issued = {}
	for r in frappe.get_all("Bag Material Ledger",
			filters={"order_bag": nm, "entry_type": "Stone Issue"},
			fields=["item", "qty", "pcs", "direction"]):
		sign = 1 if (r.direction or "In") == "In" else -1
		e = issued.setdefault(r.item, {"ct": 0.0, "pcs": 0})
		e["ct"] += sign * flt(r.qty)
		e["pcs"] += sign * int(r.pcs or 0)

	lines = []
	for r in bag.bag_bom:
		stone_type = frappe.db.get_value("Item", r.item, "stone_type")
		if not stone_type:
			continue  # metals are issued at Casting, never here
		got = issued.get(r.item, {"ct": 0.0, "pcs": 0})
		lines.append({
			"item": r.item, "stone_type": stone_type,
			"bucket": (_BUCKET_OF_STONE_TYPE.get(stone_type) or "poth").upper(),
			"plan_pcs": int(flt(r.qty) * (bag.qty or 1)), "plan_ct": round(flt(r.weight) * (bag.qty or 1), 3),
			"issued_pcs": got["pcs"], "issued_ct": round(got["ct"], 3),
			"available_ct": flt(frappe.db.get_value("Bin", {"item_code": r.item, "warehouse": wh}, "actual_qty")),
		})
	if not lines:
		return {"error": "no_stones", "card": nm,
			"message": frappe._("{0} has no stones on its BOM — nothing to issue here.").format(nm)}
	return {
		"order_bag": bag.name, "design": bag.design or "", "qty": bag.qty or 1,
		"location": bag.location or "", "warehouse": wh,
		"design_type": (frappe.db.get_value("Design", bag.design, "design_type") if bag.design else "") or "",
		"lines": lines,
	}


STONE_BUCKET_CODES = ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth")
STONE_ISSUE_ROLE = "Jewelima Stone Issue"


def _stone_issue_admin():
	"""Only an admin (System Manager) may hand-pick who is issuing; everyone else
	is locked to their own linked Employee."""
	return "System Manager" in frappe.get_roles()


def _employee_from_user(user=None):
	"""The Employee linked to a desk user (user_id), if any."""
	return frappe.db.get_value("Employee", {"user_id": user or frappe.session.user}, "name")


def _employee_allowed_buckets(employee):
	"""The uppercase bucket codes this employee may issue. No access record on
	file means everything is allowed (the page defaults to all-on)."""
	all_codes = {c.upper() for c in STONE_BUCKET_CODES}
	if not employee:
		return all_codes
	acc = frappe.db.get_value("Stone Issue Access", {"employee": employee},
		["allow_" + c for c in STONE_BUCKET_CODES], as_dict=True)
	if not acc:
		return all_codes
	return {c.upper() for c in STONE_BUCKET_CODES if acc.get("allow_" + c)}


@frappe.whitelist()
def get_stone_issue_context():
	"""Who the Stone Issue station will book against, and what they may issue.
	Admins choose the issuer; a plain Stone Issue user is locked to themselves."""
	admin = _stone_issue_admin()
	self_emp = _employee_from_user()
	effective = None if admin else self_emp
	return {
		"can_choose_issuer": bool(admin),
		"self_employee": self_emp,
		"self_employee_name": frappe.db.get_value("Employee", self_emp, "employee_name") if self_emp else None,
		"allowed_buckets": sorted(_employee_allowed_buckets(effective)) if not admin else sorted({c.upper() for c in STONE_BUCKET_CODES}),
	}


@frappe.whitelist()
def get_employee_buckets(employee):
	"""The buckets a given employee may issue — used when an admin picks an issuer."""
	return {"allowed_buckets": sorted(_employee_allowed_buckets(employee))}


def _require_stone_issue_admin():
	if not _stone_issue_admin():
		frappe.throw(frappe._("Only an administrator can manage issue access."), frappe.PermissionError)


@frappe.whitelist()
def get_issue_access():
	"""Setup > Issue > Issue Access: every relevant issuer and the buckets they may
	issue. Relevant = has the Stone Issue role, or already has an access record.
	No record on file = all buckets on (that's the default the grid shows)."""
	_require_stone_issue_admin()
	users = frappe.get_all("Has Role", filters={"role": STONE_ISSUE_ROLE, "parenttype": "User"}, pluck="parent")
	emps = set(frappe.get_all("Employee", filters={"user_id": ["in", users]}, pluck="name")) if users else set()
	emps |= set(frappe.get_all("Stone Issue Access", pluck="employee"))
	codes = [c.upper() for c in STONE_BUCKET_CODES]
	rows = []
	for e in sorted(emps):
		allowed = _employee_allowed_buckets(e)
		rows.append({"employee": e, "employee_name": frappe.db.get_value("Employee", e, "employee_name"),
			"buckets": {c: (1 if c in allowed else 0) for c in codes}})
	return {"buckets": codes, "rows": rows}


@frappe.whitelist()
def save_issue_access(rows):
	"""Upsert the per-employee bucket locks from the setup page. Each row carries the
	employee and a {BUCKET: 0/1} map; a fully-on row leaves the employee unrestricted."""
	_require_stone_issue_admin()
	rows = frappe.parse_json(rows) if isinstance(rows, str) else (rows or [])
	saved = 0
	for r in rows:
		emp = r.get("employee")
		if not emp or not frappe.db.exists("Employee", emp):
			continue
		buckets = r.get("buckets") or {}
		name = frappe.db.get_value("Stone Issue Access", {"employee": emp}, "name")
		doc = frappe.get_doc("Stone Issue Access", name) if name else frappe.new_doc("Stone Issue Access")
		doc.employee = emp
		for c in STONE_BUCKET_CODES:
			doc.set("allow_" + c, 1 if buckets.get(c.upper(), 1) else 0)
		doc.save(ignore_permissions=True)
		saved += 1
	frappe.db.commit()
	return {"saved": saved}


@frappe.whitelist()
def stone_issue_apply(order_bag, lines, issued_by=None):
	"""Issue several stone lines into one card (pcs + carats each). Per line: a Bag
	Material Ledger 'Stone Issue' row + real stock Stone Issue -> In Bags, plus one
	Material Issue record (the who/what/when paper trail). Only items on the card's
	BOM that ARE stones; availability checked up front."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE, STONE_ISSUE_WAREHOUSE

	if isinstance(lines, str):
		lines = json.loads(lines or "[]")
	lines = [l for l in (lines or []) if flt(l.get("ct")) > 0 or cint(l.get("pcs"))]
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Order Bag {0} not found.").format(order_bag))
	if not lines:
		frappe.throw(frappe._("Enter a Qty + Carat weight on at least one stone line."))

	# non-admins can only ever issue as themselves — the client's issued_by is ignored
	if not _stone_issue_admin():
		issued_by = _employee_from_user()
		if not issued_by:
			frappe.throw(frappe._("Your login isn't linked to an Employee, so you can't issue stones."))
	if not issued_by or not frappe.db.exists("Employee", issued_by):
		frappe.throw(frappe._("Pick who is issuing these stones."))

	# the issuer may be locked to only certain stone buckets
	allowed = _employee_allowed_buckets(issued_by)

	bag = frappe.get_doc("Order Bag", order_bag)
	if bag.is_finished or bag.stock_status != "In Production":
		frappe.throw(frappe._("{0} is not on the floor anymore.").format(order_bag))
	bom_items = {r.item for r in bag.bag_bom}
	wh = _wh(STONE_ISSUE_WAREHOUSE)

	for l in lines:
		item, ct, pcs = l.get("item"), flt(l.get("ct")), cint(l.get("pcs"))
		if item not in bom_items:
			frappe.throw(frappe._("{0} is not on this card's BOM.").format(item))
		stone_type = frappe.db.get_value("Item", item, "stone_type")
		if not stone_type:
			frappe.throw(frappe._("{0} is not a stone — only stones are issued here.").format(item))
		bucket = (_BUCKET_OF_STONE_TYPE.get(stone_type) or "poth").upper()
		if bucket not in allowed:
			frappe.throw(frappe._("The issuer isn't allowed to issue {0} stones ({1}).").format(bucket, item))
		if ct <= 0 or pcs <= 0:
			frappe.throw(frappe._("{0}: enter both a Qty (pcs) and a Carat weight.").format(item))
		avail = flt(frappe.db.get_value("Bin", {"item_code": item, "warehouse": wh}, "actual_qty"))
		if ct > avail + 0.0005:
			frappe.throw(frappe._("Only {0} ct of {1} at {2} — can't issue {3} ct.").format(avail, item, wh, ct))

	mi = _material_issue_record("Stone", order_bag, wh, issued_by=issued_by, items=[
		{"item": l.get("item"), "pcs": cint(l.get("pcs")), "qty": flt(l.get("ct")), "uom": "Carat"} for l in lines
	])
	bag_qty = bag.qty or 1
	plan_dirty = False
	for l in lines:
		item, ct, pcs = l.get("item"), flt(l.get("ct")), cint(l.get("pcs"))
		_bag_ledger(order_bag, item, "In", ct, "Stone Issue", pcs=pcs, employee=issued_by,
			remarks="Stone Issue station", reference=mi.name)
		_stock_move(item, ct, wh, _wh(IN_PRODUCTION_WAREHOUSE))
		# a freshly added line may carry a BLANK plan — the actual fills it in
		row = next((r for r in bag.bag_bom if r.item == item), None)
		if row and flt(row.weight) <= 0:
			row.weight = ct / bag_qty
			if cint(row.qty) <= 0:
				row.qty = pcs / bag_qty
			plan_dirty = True
	if plan_dirty:
		bag.save(ignore_permissions=True)
	frappe.db.commit()
	out = get_stone_issue_card(order_bag)
	out["material_issue"] = mi.name
	return out


def _material_issue_record(issue_type, order_bag, warehouse, issued_by=None, items=None):
	"""One Material Issue record — the who/what/when paper trail, fully built."""
	return frappe.get_doc({
		"doctype": "Material Issue", "issue_type": issue_type, "order_bag": order_bag,
		"issued_by": issued_by or None, "warehouse": warehouse,
		"posting": frappe.utils.now_datetime(), "recorded_by": frappe.session.user,
		"items": items or [],
	}).insert(ignore_permissions=True)


@frappe.whitelist()
def get_stone_issuer_today(employee):
	"""What the picked issuer has handed out TODAY (Stone Issue station side panel):
	totals + the line-by-line history (item, pcs, ct, card, time), newest first."""
	if not employee:
		return {"pcs": 0, "ct": 0.0, "cards": 0, "lines": []}
	rows = frappe.db.sql("""
		SELECT i.item, i.pcs, i.qty, m.order_bag, m.posting
		FROM `tabMaterial Issue Item` i
		JOIN `tabMaterial Issue` m ON m.name = i.parent
		WHERE m.issue_type = 'Stone' AND m.issued_by = %s AND DATE(m.posting) = CURDATE()
		ORDER BY m.posting DESC
	""", employee, as_dict=True)
	return {
		"pcs": sum(cint(r.pcs) for r in rows),
		"ct": round(sum(flt(r.qty) for r in rows), 3),
		"cards": len({r.order_bag for r in rows}),
		"lines": [{"item": r.item, "pcs": cint(r.pcs), "ct": flt(r.qty),
			"order_bag": r.order_bag, "time": str(r.posting)} for r in rows],
	}


@frappe.whitelist()
def get_day_sheet_html(date):
	"""Day Sheet page: the rendered 4-page sheet for a date (single source of
	truth = the 'Day Sheet' print format), or exists=False when no record yet."""
	d = str(date or "").strip()
	if not d or not frappe.db.exists("Day Record", d):
		return {"exists": False, "name": d}
	# render ONLY the sheet markup (get_print wraps it in the whole print-view page)
	pf = frappe.get_doc("Print Format", "Day Sheet")
	html = frappe.render_template(pf.html, {"doc": frappe.get_doc("Day Record", d)})
	return {"exists": True, "name": d, "html": html}


@frappe.whitelist()
def get_usage_report():
	"""Reports > Usage (SM only): DB size + top tables, core document counts,
	bags-per-day trend, users/sessions, and server disk — the capacity dashboard."""
	frappe.only_for(("System Manager",))
	import shutil

	db_rows = frappe.db.sql("""
		SELECT table_name, table_rows, ROUND((data_length+index_length)/1024/1024, 2) mb
		FROM information_schema.tables WHERE table_schema = DATABASE()
		ORDER BY (data_length+index_length) DESC LIMIT 10
	""", as_dict=True)
	total_mb = flt(frappe.db.sql("""
		SELECT ROUND(SUM(data_length+index_length)/1024/1024, 1)
		FROM information_schema.tables WHERE table_schema = DATABASE()
	""")[0][0])

	def cnt(dt, filters=None):
		return frappe.db.count(dt, filters or {})

	docs = {
		"Order Bags (total)": cnt("Order Bag"),
		"Order Bags on the floor": cnt("Order Bag", {"stock_status": "In Production", "is_finished": 0}),
		"Job Orders": cnt("Job Order"),
		"Bag Material Ledger rows": cnt("Bag Material Ledger"),
		"Stock Ledger Entries": cnt("Stock Ledger Entry"),
		"Material Issues": cnt("Material Issue"),
		"Product Sales": cnt("Product Sale"),
		"Designs": cnt("Design"),
	}

	trend = frappe.db.sql("""
		SELECT DATE(creation) d, COUNT(*) n FROM `tabOrder Bag`
		WHERE creation > DATE_SUB(CURDATE(), INTERVAL 14 DAY)
		GROUP BY DATE(creation) ORDER BY d
	""", as_dict=True)

	sessions = frappe.db.sql("""
		SELECT user, MAX(lastupdate) last_seen FROM tabSessions
		WHERE user NOT IN ('Guest') GROUP BY user ORDER BY last_seen DESC LIMIT 25
	""", as_dict=True)
	enabled_users = frappe.db.count("User", {"enabled": 1, "user_type": "System User"})

	du = shutil.disk_usage("/")
	return {
		"db": {"total_mb": total_mb, "tables": db_rows},
		"docs": docs,
		"trend": [{"date": str(r.d), "n": r.n} for r in trend],
		"users": {"enabled": enabled_users,
			"sessions": [{"user": r.user, "last_seen": str(r.last_seen)} for r in sessions]},
		"disk": {"total_gb": round(du.total / 1e9, 1), "free_gb": round(du.free / 1e9, 1)},
	}


# --- Data retention (prune BY CHOICE — the Day Record keeps the day's story) -------
# Only informative/monitoring rows of CLOSED bags (Sold / Cancelled) ever qualify;
# Stock Ledger Entries are valuation truth and are never touched.

def _prune_candidates(months):
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	cutoff = frappe.utils.add_months(frappe.utils.today(), -cint(months or 3))
	closed = "SELECT name FROM `tabOrder Bag` WHERE stock_status IN ('Sold','Cancelled')"
	kinds = {}
	for dt in dict.fromkeys(BENCH_DOCTYPE.values()):
		if frappe.db.exists("DocType", dt):
			n = frappe.db.sql("""SELECT COUNT(*) FROM `tab{0}`
				WHERE creation < %s AND order_bag IN ({1})""".format(dt, closed), cutoff)[0][0]
			if n:
				kinds["bench:" + dt] = n
	kinds["ledger"] = frappe.db.sql("""SELECT COUNT(*) FROM `tabBag Material Ledger`
		WHERE datetime < %s AND order_bag IN ({0})""".format(closed), cutoff)[0][0]
	kinds["material_issue"] = frappe.db.sql("""SELECT COUNT(*) FROM `tabMaterial Issue`
		WHERE posting < %s AND order_bag IN ({0})""".format(closed), cutoff)[0][0]
	unsealed = frappe.db.sql("""
		SELECT COUNT(DISTINCT DATE(l.datetime)) FROM `tabBag Material Ledger` l
		WHERE l.datetime < %s AND l.order_bag IN ({0})
		AND DATE(l.datetime) NOT IN (SELECT date FROM `tabDay Record`)""".format(closed), cutoff)[0][0]
	return cutoff, kinds, unsealed


@frappe.whitelist()
def get_prune_preview(months=3):
	"""DRY RUN: what a prune would delete. Nothing is touched."""
	frappe.only_for(("System Manager",))
	cutoff, kinds, unsealed = _prune_candidates(months)
	return {"cutoff": str(cutoff), "kinds": kinds, "total": sum(kinds.values()),
		"unsealed_days": unsealed}


@frappe.whitelist()
def prune_execute(months=3, confirm_text=None):
	"""Actually delete the preview's rows. Demands the literal confirmation text —
	this is the one destructive button in the app, used only when space runs out."""
	frappe.only_for(("System Manager",))
	if (confirm_text or "").strip().upper() != "PRUNE":
		frappe.throw(frappe._('Type PRUNE to confirm — this permanently deletes old monitoring rows.'))
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	cutoff, kinds, _un = _prune_candidates(months)
	closed = "SELECT name FROM `tabOrder Bag` WHERE stock_status IN ('Sold','Cancelled')"
	deleted = {}
	for dt in dict.fromkeys(BENCH_DOCTYPE.values()):
		if ("bench:" + dt) in kinds:
			frappe.db.sql("""DELETE FROM `tab{0}` WHERE creation < %s
				AND order_bag IN ({1})""".format(dt, closed), cutoff)
			deleted[dt] = kinds["bench:" + dt]
	frappe.db.sql("""DELETE FROM `tabBag Material Ledger` WHERE datetime < %s
		AND order_bag IN ({0})""".format(closed), cutoff)
	deleted["Bag Material Ledger"] = kinds.get("ledger", 0)
	mi = frappe.get_all("Material Issue", filters={"posting": ["<", cutoff]}, pluck="name")
	mi = [m for m in mi if frappe.db.get_value("Material Issue", m, "order_bag") in
		set(frappe.get_all("Order Bag", filters={"stock_status": ["in", ["Sold", "Cancelled"]]}, pluck="name"))]
	for m in mi:
		frappe.delete_doc("Material Issue", m, force=True, ignore_permissions=True)
	deleted["Material Issue"] = len(mi)
	frappe.db.commit()
	frappe.get_doc({"doctype": "Comment", "comment_type": "Info", "reference_doctype": "User",
		"reference_name": frappe.session.user,
		"content": "Prune executed: cutoff {0}, deleted {1}".format(cutoff, deleted)}).insert(ignore_permissions=True)
	return {"cutoff": str(cutoff), "deleted": deleted}


_STONE_AUDIT_TOL = 0.005  # carats below this are rounding dust, treated as zero


@frappe.whitelist()
def get_stone_audit():
	"""Per-card stone lines whose net PIECES and net CARATS tell different stories
	(cards still on the floor). Weight is the stock truth — the audit exists to
	make the piece story match it."""
	rows = frappe.db.sql("""
		SELECT l.order_bag, l.item,
			SUM(IF(l.direction='Out', -l.qty, l.qty)) net_ct,
			SUM(IF(l.direction='Out', -l.pcs, l.pcs)) net_pcs,
			b.design, b.location
		FROM `tabBag Material Ledger` l
		JOIN `tabOrder Bag` b ON b.name = l.order_bag
		JOIN `tabItem` i ON i.name = l.item
		WHERE b.stock_status = 'In Production' AND b.is_finished = 0
			AND IFNULL(i.stone_type, '') != ''
		GROUP BY l.order_bag, l.item
	""", as_dict=True)
	out = []
	for r in rows:
		ct, pcs = flt(r.net_ct), cint(r.net_pcs)
		if abs(ct) < _STONE_AUDIT_TOL and pcs == 0:
			continue  # clean (or pure rounding dust on both axes)
		problem = None
		if ct < -_STONE_AUDIT_TOL or pcs < 0:
			problem = "negative"  # books went below zero — data error
		elif pcs > 0 and abs(ct) < _STONE_AUDIT_TOL:
			problem = "count_without_weight"  # pcs left, carats gone
		elif pcs == 0 and ct >= _STONE_AUDIT_TOL:
			problem = "weight_without_count"  # residual carats, no stones
		if problem:
			out.append({"order_bag": r.order_bag, "design": r.design or "", "location": r.location or "",
				"item": r.item, "net_pcs": pcs, "net_ct": round(ct, 3), "problem": problem})
	loss_benches = [w.replace(" -LOSS", "") for w in frappe.get_all("Warehouse",
		filters={"warehouse_name": ["like", "% -LOSS"]}, pluck="warehouse_name")]
	return {"rows": out, "loss_benches": sorted(loss_benches)}


@frappe.whitelist()
def stone_audit_fix(order_bag, item, action, bench=None):
	"""Resolve one audit line. 'zero_pcs': corrective Adjustment row so the count
	matches the (zero) weight. 'sweep': residual carats go to a stage's -LOSS
	bucket (Option B — residue is collected, never vanishes)."""
	frappe.only_for(("System Manager", "Stock Manager"))
	net = frappe.db.sql("""
		SELECT SUM(IF(direction='Out', -qty, qty)) ct, SUM(IF(direction='Out', -pcs, pcs)) pcs
		FROM `tabBag Material Ledger` WHERE order_bag = %s AND item = %s
	""", (order_bag, item), as_dict=True)[0]
	ct, pcs = flt(net.ct), cint(net.pcs)

	if action == "zero_pcs":
		if abs(ct) >= _STONE_AUDIT_TOL:
			frappe.throw(frappe._("{0} still nets {1} ct — only counts orphaned from weight can be zeroed.").format(item, round(ct, 3)))
		if pcs == 0:
			frappe.throw(frappe._("Nothing to correct — the count is already zero."))
		_bag_ledger(order_bag, item, "Out" if pcs > 0 else "In", 0, "Adjustment",
			pcs=abs(pcs), remarks="Stone audit: count corrected to match weight")
	elif action == "sweep":
		if ct < _STONE_AUDIT_TOL:
			frappe.throw(frappe._("No residual carats to sweep on {0}.").format(item))
		if pcs != 0:
			frappe.throw(frappe._("{0} still counts {1} pcs — zero the count first if the stones are truly gone.").format(item, pcs))
		# ledger benches are the UPPERCASE locations; the -LOSS warehouse uses the Title
		if not bench or not _wh("{0} -LOSS".format(bench)):
			frappe.throw(frappe._("Pick which stage's -LOSS bucket takes the residue."))
		book_loss(order_bag, item, ct, bench=bench.upper(), remarks="Stone audit: residual carats swept")
	else:
		frappe.throw(frappe._("Unknown action."))
	frappe.db.commit()
	return get_stone_audit()


# --- Selection: pick photos from the catalog, keep the record ----------------------

@frappe.whitelist()
def get_selection_photos(search=None, design_type=None, provider=None, tag=None,
		in_stock=None, gold_min=None, gold_max=None, cts_min=None, cts_max=None, limit=500):
	"""The photo catalog for the Selection page + everything there is to filter by
	(design types, providers, tags). ONLY reviewed photos show — the unreviewed
	imports live on the Review page until a human confirms them."""
	design_types = frappe.db.sql_list(
		"SELECT DISTINCT design_type FROM `tabSelection Photo` WHERE IFNULL(design_type,'') != '' AND reviewed=1 ORDER BY design_type")
	providers = frappe.db.sql_list(
		"SELECT DISTINCT provider FROM `tabSelection Photo` WHERE IFNULL(provider,'') != '' AND reviewed=1 ORDER BY provider")
	# EVERY tag ever created (the master), not just the used ones — the filter bar
	# shows the whole vocabulary, with its colours
	tags_all = frappe.get_all("Selection Tag", fields=["name as tag", "color"], order_by="tag_name")

	filters = {"active": 1, "reviewed": 1}
	if design_type:
		filters["design_type"] = design_type
	if provider:
		filters["provider"] = provider
	if search:
		filters["code"] = ["like", "%{0}%".format(search)]
	if cint(in_stock):
		filters["stock_pcs"] = [">", 0]
	rows = frappe.get_all("Selection Photo", filters=filters,
		fields=["name", "code", "image", "gold_gms", "cts", "design_type", "provider", "stock_pcs"],
		order_by="code", limit_page_length=cint(limit) or 500)
	# weight-range filters (a photo with no value entered never matches a range)
	if gold_min not in (None, ""):
		rows = [r for r in rows if flt(r.gold_gms) >= flt(gold_min)]
	if gold_max not in (None, ""):
		rows = [r for r in rows if flt(r.gold_gms) and flt(r.gold_gms) <= flt(gold_max)]
	if cts_min not in (None, ""):
		rows = [r for r in rows if flt(r.cts) >= flt(cts_min)]
	if cts_max not in (None, ""):
		rows = [r for r in rows if flt(r.cts) and flt(r.cts) <= flt(cts_max)]

	# tags per photo, painted on the cards + used for the tag filter
	photo_tags = {}
	for r in frappe.db.sql("""SELECT parent, tag FROM `tabSelection Photo Tag`
		WHERE parenttype='Selection Photo' ORDER BY idx""", as_dict=True):
		photo_tags.setdefault(r.parent, []).append(r.tag)
	for r in rows:
		r["tags"] = photo_tags.get(r.name, [])
	if tag:
		rows = [r for r in rows if tag in r["tags"]]
	return {"design_types": design_types, "providers": providers,
		"tags": tags_all, "photos": rows, "total": len(rows)}


@frappe.whitelist()
def update_selection_photo(name, design_type=None, provider=None, stock_pcs=None, tags=None):
	"""Edit one catalog photo from the Selection page's viewer: assign the design
	type, who makes it, the pieces in stock, and its tags. Unknown tags are created
	on the fly (Design Tag is a plain master, shared with the Design Bank)."""
	if not frappe.db.exists("Selection Photo", name):
		frappe.throw(frappe._("Photo {0} not found.").format(name))
	if isinstance(tags, str):
		tags = json.loads(tags or "[]")
	doc = frappe.get_doc("Selection Photo", name)
	if design_type is not None:
		if design_type and not frappe.db.exists("Design Type", design_type):
			frappe.throw(frappe._("Design Type {0} does not exist.").format(design_type))
		doc.design_type = design_type or None
	if provider is not None:
		if provider and not frappe.db.exists("Supplier", provider):
			frappe.throw(frappe._("Provider (Supplier) {0} does not exist.").format(provider))
		doc.provider = provider or None
	if stock_pcs is not None:
		doc.stock_pcs = max(cint(stock_pcs), 0)
	if tags is not None:
		clean = []
		for t in tags:
			t = (t or "").strip().upper()
			if not t or t in clean:
				continue
			if not frappe.db.exists("Selection Tag", t):
				frappe.get_doc({"doctype": "Selection Tag", "tag_name": t}).insert(ignore_permissions=True)
			clean.append(t)
		doc.set("tags", [{"tag": t} for t in clean])
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "design_type": doc.design_type, "provider": doc.provider,
		"stock_pcs": doc.stock_pcs, "tags": [t.tag for t in doc.tags]}


# --- Selection Review (page: selection-review) --------------------------------
# Imported photos carry OCR'd weights; a human confirms each one (code, photo,
# gold, dia) and ticks Reviewed. A photo's CODE is its identity — one design,
# one code — so a code change that collides brings up both photos side by side
# and one of them has to go.
@frappe.whitelist()
def get_selection_review(status="pending", search=None, limit=100):
	"""The review queue. status: pending | done | all."""
	frappe.only_for(("System Manager", "Stock Manager"))
	filters = {"active": 1}
	if status == "pending":
		filters["reviewed"] = 0
	elif status == "done":
		filters["reviewed"] = 1
	if search:
		filters["code"] = ["like", "%{0}%".format(search)]
	rows = frappe.get_all("Selection Photo", filters=filters,
		fields=["name", "code", "image", "gold_gms", "cts", "stock_pcs",
			"design_type", "provider", "reviewed"],
		order_by="reviewed asc, code asc", limit_page_length=cint(limit) or 100)
	return {"rows": rows,
		"total": frappe.db.count("Selection Photo", {"active": 1}),
		"reviewed": frappe.db.count("Selection Photo", {"active": 1, "reviewed": 1})}


@frappe.whitelist()
def review_save(name, gold_gms=None, cts=None, stock_pcs=None, reviewed=None,
		design_type=None, provider=None):
	"""One review row: save values and/or the Reviewed tick."""
	frappe.only_for(("System Manager", "Stock Manager"))
	if not frappe.db.exists("Selection Photo", name):
		frappe.throw(frappe._("Photo {0} not found.").format(name))
	doc = frappe.get_doc("Selection Photo", name)
	if gold_gms is not None:
		doc.gold_gms = flt(gold_gms)
	if cts is not None:
		doc.cts = flt(cts)
	if stock_pcs is not None:
		doc.stock_pcs = max(cint(stock_pcs), 0)
	if design_type is not None:
		doc.design_type = design_type or None
	if provider is not None:
		doc.provider = provider or None
	if reviewed is not None:
		doc.reviewed = cint(reviewed)
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "reviewed": doc.reviewed}


@frappe.whitelist()
def review_rename_code(name, new_code):
	"""Change a photo's code. If the target code exists, DON'T rename — return
	both photos so the page can show them side by side; the reviewer then keeps
	one via review_delete_photo and retries."""
	frappe.only_for(("System Manager", "Stock Manager"))
	new_code = (new_code or "").strip().upper()
	if not new_code:
		frappe.throw(frappe._("Code is required."))
	if new_code == name:
		return {"renamed": 0, "name": name}
	if frappe.db.exists("Selection Photo", new_code):
		fields = ["name", "code", "image", "gold_gms", "cts", "stock_pcs",
			"design_type", "provider", "reviewed"]
		mine = frappe.db.get_value("Selection Photo", name, fields, as_dict=True)
		other = frappe.db.get_value("Selection Photo", new_code, fields, as_dict=True)
		for d in (mine, other):
			d["selections"] = frappe.db.count("Selection Item", {"photo": d.name})
		return {"conflict": 1, "mine": mine, "existing": other}
	frappe.rename_doc("Selection Photo", name, new_code, force=True)
	frappe.db.set_value("Selection Photo", new_code, "code", new_code, update_modified=False)
	frappe.db.commit()
	return {"renamed": 1, "name": new_code}


@frappe.whitelist()
def review_delete_photo(name):
	"""Remove a duplicate photo. Any Selection lines pointing at it are dropped
	and their Selections re-totalled first, so nothing dangles."""
	frappe.only_for(("System Manager", "Stock Manager"))
	if not frappe.db.exists("Selection Photo", name):
		frappe.throw(frappe._("Photo {0} not found.").format(name))
	parents = frappe.get_all("Selection Item", filters={"photo": name}, pluck="parent")
	for sel in set(parents):
		doc = frappe.get_doc("Selection", sel)
		doc.set("items", [i for i in doc.items if i.photo != name])
		if doc.items:
			doc.save(ignore_permissions=True)
		else:
			frappe.delete_doc("Selection", sel, force=True, ignore_permissions=True)
	frappe.delete_doc("Selection Photo", name, force=True, ignore_permissions=True)
	frappe.db.commit()
	return {"deleted": name, "selections_touched": len(set(parents))}


# --- Diamond Sieve chart (Stones > Sieve Chart) --------------------------------
@frappe.whitelist()
def get_sieve_chart():
	"""The whole chart, in chart order, for the excel-style page."""
	return frappe.get_all("Diamond Sieve", fields=["name", "sieve_size", "mm_size", "avg_cts", "idx_order"],
		order_by="idx_order asc, name asc", limit_page_length=0)


@frappe.whitelist()
def save_sieve_chart(rows):
	"""Save edited cells from the Sieve Chart page. rows = [{name, mm_size, avg_cts}]."""
	frappe.only_for(("System Manager", "Stock Manager"))
	if isinstance(rows, str):
		rows = json.loads(rows or "[]")
	n = 0
	for r in rows or []:
		if not frappe.db.exists("Diamond Sieve", r.get("name")):
			continue
		frappe.db.set_value("Diamond Sieve", r["name"], {
			"mm_size": flt(r.get("mm_size")), "avg_cts": flt(r.get("avg_cts"))})
		n += 1
	frappe.db.commit()
	return {"saved": n}


@frappe.whitelist()
def get_sieve_map():
	"""size label -> avg cts/stone, for the pages that auto-fill qty<->carat
	(purchase entry, BOM entry). The size label matches the tail of the diamond
	item codes: 'SI-IJ 1-1.5' -> '1-1.5'."""
	return {r.sieve_size: flt(r.avg_cts) for r in frappe.get_all(
		"Diamond Sieve", fields=["sieve_size", "avg_cts"], limit_page_length=0) if flt(r.avg_cts) > 0}


# --- Stone Stock (CAD) — read-only: is that stone FREE to use? -----------------
@frappe.whitelist()
def get_stone_stock(search=None, family=None):
	"""Stone Issue warehouse only. FREE = what's physically there MINUS what open
	production cards still plan to draw (their plan lines less what's already
	been issued to them). CAD sees only stones that are genuinely available."""
	from jewelima.setup import STONE_ISSUE_WAREHOUSE

	wh = _wh(STONE_ISSUE_WAREHOUSE)
	bins = frappe.db.sql("""
		SELECT b.item_code, i.item_group, i.stone_type, b.actual_qty
		FROM `tabBin` b JOIN `tabItem` i ON i.name = b.item_code
		WHERE b.warehouse = %s AND IFNULL(i.stone_type, '') != '' AND b.actual_qty > 0.0005
		ORDER BY b.item_code""", wh, as_dict=True)

	# planned demand: stone plan lines of OPEN cards, minus what those cards
	# already received (Bag Material Ledger In) — clamped at zero per card+item
	plan = frappe.db.sql("""
		SELECT bi.item, ob.name AS bag, SUM(bi.weight) AS w
		FROM `tabOrder Bag BOM Item` bi
		JOIN `tabOrder Bag` ob ON ob.name = bi.parent
		JOIN `tabItem` i ON i.name = bi.item
		WHERE ob.is_finished = 0 AND ob.stock_status = 'In Production'
			AND IFNULL(i.stone_type, '') != ''
		GROUP BY bi.item, ob.name""", as_dict=True)
	got = {}
	if plan:
		for r in frappe.db.sql("""
			SELECT item, order_bag, SUM(CASE WHEN direction='In' THEN qty ELSE -qty END) AS q
			FROM `tabBag Material Ledger`
			WHERE order_bag IN %(bags)s
			GROUP BY item, order_bag""", {"bags": tuple({r.bag for r in plan})}, as_dict=True):
			got[(r.item, r.order_bag)] = flt(r.q)
	pending = {}
	for r in plan:
		short = flt(r.w) - max(got.get((r.item, r.bag), 0.0), 0.0)
		if short > 0.0005:
			pending[r.item] = pending.get(r.item, 0.0) + short

	sieve = {r.sieve_size: flt(r.avg_cts) for r in frappe.get_all(
		"Diamond Sieve", fields=["sieve_size", "avg_cts"], limit_page_length=0) if flt(r.avg_cts) > 0}
	out, fams = [], set()
	for b in bins:
		fam = "DIAMOND" if (b.item_group or "").startswith("DIAMOND") else (b.item_group or "")
		fams.add(fam)
		if family and fam != family:
			continue
		if search and search.upper() not in b.item_code.upper():
			continue
		stock = flt(b.actual_qty)
		planned = round(pending.get(b.item_code, 0.0), 3)
		free = round(stock - planned, 3)
		if free <= 0.0005:
			continue   # fully spoken for — not available to CAD
		avg = sieve.get(b.item_code.split(" ", 1)[1]) if (b.stone_type == "Diamond" and " " in b.item_code) else None
		out.append({"item": b.item_code, "group": b.item_group, "family": fam,
			"stock": round(stock, 3), "planned": planned, "free": free,
			"est_pcs": int(free / avg) if avg else None})
	return {"rows": out, "families": sorted(fams), "warehouse": wh}


# --- Repack Stock (Stones) — split bulk stone stock into sieves, with approval
# The requester proposes the split; someone with the bigger role (System Manager /
# Stock Manager) approves, and only THEN does a Repack Stock Entry move the stock.
# Locked to the Stone Issue warehouse. Family rule: a stone can only repack into
# its own parent group — CZ -> CZ sieves; DIAMOND -> any diamond quality + sieve.
def _stone_family(item):
	group = frappe.db.get_value("Item", item, "item_group") or ""
	if group.startswith("DIAMOND"):
		return "DIAMOND"
	return group   # CUBIC ZIRCONIA, CVD, SWAROVSKI, ... are their own family


@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def stone_item_search(doctype, txt, searchfield, start, page_len, filters):
	"""Link-field search that puts the EXACT / shortest match first — Frappe's
	default relevance drowns the bare 'CZ' under its 49 sieve children."""
	filters = filters or {}
	cond, vals = ["i.disabled = 0", "i.is_stock_item = 1"], {"txt": f"%{txt}%", "exact": txt}
	if filters.get("item_group"):
		ig = filters["item_group"]
		if isinstance(ig, (list, tuple)) and len(ig) == 2 and ig[0] == "in":
			cond.append("i.item_group IN %(groups)s")
			vals["groups"] = tuple(ig[1]) or ("",)
		else:
			cond.append("i.item_group = %(group)s")
			vals["group"] = ig
	if filters.get("stone_only"):
		cond.append("IFNULL(i.stone_type, '') != ''")
	if filters.get("exclude"):
		cond.append("i.name != %(excl)s")
		vals["excl"] = filters["exclude"]
	vals.update({"start": start, "page_len": page_len})
	return frappe.db.sql("""
		SELECT i.name, i.item_group
		FROM `tabItem` i
		WHERE {cond} AND i.name LIKE %(txt)s
		ORDER BY (i.name = %(exact)s) DESC, CHAR_LENGTH(i.name), i.name
		LIMIT %(start)s, %(page_len)s""".format(cond=" AND ".join(cond)), vals)


@frappe.whitelist()
def get_repack_context(source_item=None):
	"""What the Repack page needs: the locked warehouse, and (given a source)
	its available stock there + the items it may legally split into."""
	wh = _wh("Stone Issue")
	out = {"warehouse": wh, "can_approve": bool({"System Manager", "Stock Manager"} & set(frappe.get_roles()))}
	if source_item:
		fam = _stone_family(source_item)
		out["family"] = fam
		out["available"] = flt(frappe.db.get_value("Bin", {"item_code": source_item, "warehouse": wh}, "actual_qty"))
		if fam == "DIAMOND":
			groups = frappe.get_all("Item Group", filters={"name": ["like", "DIAMOND %"], "is_group": 0}, pluck="name")
		else:
			groups = [fam]
		out["target_groups"] = groups
	return out


@frappe.whitelist()
def create_repack_request(source_item, qty, targets, remarks=None):
	"""Place a repack request (Pending). Validates the family rule and that the
	target quantities add up EXACTLY to the source qty — repacking never creates
	or loses carats."""
	if isinstance(targets, str):
		targets = json.loads(targets or "[]")
	qty = flt(qty)
	if not frappe.db.exists("Item", source_item):
		frappe.throw(frappe._("Item {0} not found.").format(source_item))
	if not frappe.db.get_value("Item", source_item, "stone_type"):
		frappe.throw(frappe._("{0} is not a stone.").format(source_item))
	if qty <= 0:
		frappe.throw(frappe._("Qty must be positive."))
	fam = _stone_family(source_item)
	rows = []
	total = 0.0
	for t in targets or []:
		it, q = t.get("item"), flt(t.get("qty"))
		if not it or q <= 0:
			continue
		if it == source_item:
			frappe.throw(frappe._("{0} can't be repacked into itself.").format(it))
		if _stone_family(it) != fam:
			frappe.throw(frappe._("{0} is outside the {1} family — a stone only repacks within its own group.").format(it, fam))
		rows.append({"item": it, "qty": q})
		total += q
	if not rows:
		frappe.throw(frappe._("Add at least one target line."))
	if abs(total - qty) > 0.0005:
		frappe.throw(frappe._("Targets add up to {0} ct but the source is {1} ct — a repack must balance exactly.").format(round(total, 3), qty))
	wh = _wh("Stone Issue")
	doc = frappe.get_doc({
		"doctype": "Repack Request", "source_item": source_item, "qty": qty,
		"warehouse": wh, "status": "Pending", "remarks": remarks,
		"requested_by": frappe.session.user, "requested_on": frappe.utils.now_datetime(),
		"targets": rows,
	})
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name}


@frappe.whitelist()
def list_repack_requests(status=None, limit=50):
	filters = {}
	if status and status != "all":
		filters["status"] = status.title()
	rows = frappe.get_all("Repack Request", filters=filters,
		fields=["name", "source_item", "qty", "status", "requested_by", "requested_on",
			"approved_by", "stock_entry", "reject_reason"],
		order_by="creation desc", limit_page_length=cint(limit) or 50)
	tmap = {}
	for t in frappe.get_all("Repack Request Item", filters={"parent": ["in", [r.name for r in rows]]},
			fields=["parent", "item", "qty"], order_by="idx"):
		tmap.setdefault(t.parent, []).append({"item": t.item, "qty": flt(t.qty)})
	for r in rows:
		r["targets"] = tmap.get(r.name, [])
		r["requested_on"] = str(r.requested_on or "")
	return rows


@frappe.whitelist()
def approve_repack(name):
	"""The bigger role signs off: checks live stock, writes ONE Repack Stock Entry
	(source out, sieves in — same warehouse), stamps the request Approved."""
	frappe.only_for(("System Manager", "Stock Manager"))
	doc = frappe.get_doc("Repack Request", name)
	if doc.status != "Pending":
		frappe.throw(frappe._("{0} is already {1}.").format(name, doc.status))
	have = flt(frappe.db.get_value("Bin", {"item_code": doc.source_item, "warehouse": doc.warehouse}, "actual_qty"))
	if have + 0.0005 < flt(doc.qty):
		frappe.throw(frappe._("Only {0} ct of {1} at {2} — the request needs {3} ct.").format(
			round(have, 3), doc.source_item, doc.warehouse, doc.qty))
	items = [{
		"item_code": doc.source_item, "qty": flt(doc.qty),
		"uom": frappe.db.get_value("Item", doc.source_item, "stock_uom") or "Carat",
		"s_warehouse": doc.warehouse, "allow_zero_valuation_rate": 1,
	}]
	# multiple finished goods in a Repack need explicit rates — the sieves inherit
	# the source parcel's per-carat valuation, so value is conserved like weight
	src_rate = flt(frappe.db.get_value("Bin", {"item_code": doc.source_item, "warehouse": doc.warehouse}, "valuation_rate"))
	for t in doc.targets:
		items.append({
			"item_code": t.item, "qty": flt(t.qty),
			"uom": frappe.db.get_value("Item", t.item, "stock_uom") or "Carat",
			"t_warehouse": doc.warehouse, "allow_zero_valuation_rate": 1,
			"set_basic_rate_manually": 1, "basic_rate": src_rate,
			"is_finished_item": 1,
		})
	se = frappe.get_doc({"doctype": "Stock Entry", "stock_entry_type": "Repack",
		"company": _company(), "items": items})
	se.insert(ignore_permissions=True)
	se.submit()
	doc.db_set("status", "Approved")
	doc.db_set("approved_by", frappe.session.user)
	doc.db_set("approved_on", frappe.utils.now_datetime())
	doc.db_set("stock_entry", se.name)
	frappe.db.commit()
	return {"name": name, "stock_entry": se.name}


@frappe.whitelist()
def reject_repack(name, reason=None):
	frappe.only_for(("System Manager", "Stock Manager"))
	doc = frappe.get_doc("Repack Request", name)
	if doc.status != "Pending":
		frappe.throw(frappe._("{0} is already {1}.").format(name, doc.status))
	doc.db_set("status", "Rejected")
	doc.db_set("approved_by", frappe.session.user)
	doc.db_set("approved_on", frappe.utils.now_datetime())
	if reason:
		doc.db_set("reject_reason", reason)
	frappe.db.commit()
	return {"name": name}


# --- Selection Tags (their own master — different purpose from the bank's Design Tags)
@frappe.whitelist()
def get_selection_tags(with_counts=1):
	"""All Selection Tags (name + colour) + how many catalog photos carry each."""
	tags = frappe.get_all("Selection Tag", fields=["name as tag", "color"], order_by="tag_name asc")
	if cint(with_counts):
		counts = dict(frappe.db.sql(
			"SELECT tag, COUNT(*) FROM `tabSelection Photo Tag` GROUP BY tag"))
		for t in tags:
			t["count"] = int(counts.get(t["tag"], 0))
	return tags


@frappe.whitelist()
def create_selection_tag(tag_name, color=None):
	tag_name = (tag_name or "").strip()
	if not tag_name:
		frappe.throw(frappe._("Tag name is required"))
	if frappe.db.exists("Selection Tag", tag_name):
		frappe.throw(frappe._("Tag '{0}' already exists").format(tag_name))
	doc = frappe.get_doc({"doctype": "Selection Tag", "tag_name": tag_name,
		"color": color or "#6b7280"}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"tag": doc.name, "color": doc.color, "count": 0}


@frappe.whitelist()
def rename_selection_tag(old, new):
	new = (new or "").strip()
	if not new:
		frappe.throw(frappe._("New name is required"))
	if old == new:
		return {"tag": new}
	frappe.rename_doc("Selection Tag", old, new, force=True)  # cascades into Selection Photo Tag
	frappe.db.commit()
	return {"tag": new}


@frappe.whitelist()
def delete_selection_tag(tag_name):
	frappe.db.delete("Selection Photo Tag", {"tag": tag_name})
	frappe.delete_doc("Selection Tag", tag_name, force=True, ignore_permissions=True)
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def set_selection_tag_color(tag_name, color):
	frappe.db.set_value("Selection Tag", tag_name, "color", color)
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def bulk_add_selection_tags(photos, tags):
	"""Stick tag(s) on many catalog photos at once (the Selection page's Add Tag
	button). Unknown tags are created; a photo that already carries the tag is
	left alone."""
	if isinstance(photos, str):
		photos = json.loads(photos or "[]")
	if isinstance(tags, str):
		tags = json.loads(tags or "[]")
	clean = []
	for t in tags or []:
		t = (t or "").strip().upper()
		if t and t not in clean:
			if not frappe.db.exists("Selection Tag", t):
				frappe.get_doc({"doctype": "Selection Tag", "tag_name": t}).insert(ignore_permissions=True)
			clean.append(t)
	if not clean or not photos:
		frappe.throw(frappe._("Pick photos and give at least one tag."))
	touched = 0
	for name in photos:
		if not frappe.db.exists("Selection Photo", name):
			continue
		doc = frappe.get_doc("Selection Photo", name)
		have = {t.tag for t in doc.tags}
		add = [t for t in clean if t not in have]
		if add:
			for t in add:
				doc.append("tags", {"tag": t})
			doc.save(ignore_permissions=True)
			touched += 1
	frappe.db.commit()
	return {"tags": clean, "photos_tagged": touched, "photos_given": len(photos)}


# --- Selection Providers (who makes the pieces — plain Suppliers underneath) ---
@frappe.whitelist()
def get_selection_providers():
	"""Every Supplier + how many catalog photos each one provides."""
	rows = frappe.get_all("Supplier", fields=["name", "supplier_name"], order_by="supplier_name")
	counts = dict(frappe.db.sql(
		"SELECT provider, COUNT(*) FROM `tabSelection Photo` WHERE IFNULL(provider,'') != '' GROUP BY provider"))
	for r in rows:
		r["count"] = int(counts.get(r.name, 0))
	return rows


@frappe.whitelist()
def create_selection_provider(provider_name):
	provider_name = (provider_name or "").strip()
	if not provider_name:
		frappe.throw(frappe._("Provider name is required"))
	if frappe.db.exists("Supplier", provider_name):
		frappe.throw(frappe._("Provider '{0}' already exists").format(provider_name))
	doc = frappe.get_doc({
		"doctype": "Supplier", "supplier_name": provider_name,
		"supplier_group": frappe.db.get_value("Supplier Group", {}, "name"),
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "supplier_name": doc.supplier_name, "count": 0}


@frappe.whitelist()
def rename_selection_provider(old, new):
	new = (new or "").strip()
	if not new:
		frappe.throw(frappe._("New name is required"))
	if old == new:
		return {"name": new}
	frappe.rename_doc("Supplier", old, new, force=True)  # cascades into Selection Photo.provider
	frappe.db.commit()
	return {"name": new}


@frappe.whitelist()
def delete_selection_provider(name):
	used = frappe.db.count("Selection Photo", {"provider": name})
	if used:
		frappe.throw(frappe._("{0} provides {1} photo(s) — reassign them first.").format(name, used))
	frappe.delete_doc("Supplier", name, ignore_permissions=True)
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def create_selection(payload):
	"""Record what a party picked: one Selection with a line per photo."""
	p = frappe.parse_json(payload)
	party = p.get("party")
	codes = p.get("photos") or []
	if not party or not frappe.db.exists("Customer", party):
		frappe.throw(frappe._("Pick the Party."))
	if not codes:
		frappe.throw(frappe._("Select at least one photo."))
	doc = frappe.get_doc({
		"doctype": "Selection", "party": party,
		"selection_date": p.get("selection_date") or frappe.utils.today(),
		"batch": p.get("batch") or None, "remarks": p.get("remarks"),
		"items": [{"photo": c} for c in codes],
	})
	doc.insert(ignore_permissions=True)   # controller fills code/image/weights + totals
	frappe.db.commit()
	return {"name": doc.name, "total_photos": doc.total_photos,
		"total_gold": doc.total_gold, "total_cts": doc.total_cts}


@frappe.whitelist()
def get_selected_pieces(party=None, batch=None, from_date=None, to_date=None, selection=None):
	"""Selected Pieces board: the selection records that match + every picked photo
	(with which record/party it came from), plus headline totals."""
	cond, vals = ["1=1"], {}
	if party:
		cond.append("s.party = %(party)s")
		vals["party"] = party
	if batch:
		cond.append("s.batch = %(batch)s")
		vals["batch"] = batch
	if from_date:
		cond.append("s.selection_date >= %(fd)s")
		vals["fd"] = from_date
	if to_date:
		cond.append("s.selection_date <= %(td)s")
		vals["td"] = to_date
	if selection:
		cond.append("s.name = %(sel)s")
		vals["sel"] = selection
	where = " AND ".join(cond)

	sels = frappe.db.sql("""
		SELECT s.name, s.party, s.selection_date, s.batch, s.total_photos, s.total_gold, s.total_cts, s.remarks
		FROM `tabSelection` s WHERE {0} ORDER BY s.selection_date DESC, s.creation DESC
	""".format(where), vals, as_dict=True)

	items = frappe.db.sql("""
		SELECT i.photo, i.code, i.image, i.gold_gms, i.cts,
			s.name AS selection, s.party, s.selection_date, s.batch
		FROM `tabSelection Item` i JOIN `tabSelection` s ON s.name = i.parent
		WHERE {0} ORDER BY s.selection_date DESC, s.creation DESC, i.idx
	""".format(where), vals, as_dict=True)

	parties = [r.party for r in frappe.db.sql("SELECT DISTINCT party FROM `tabSelection` ORDER BY party", as_dict=True) if r.party]
	batches = [r.batch for r in frappe.db.sql("SELECT DISTINCT batch FROM `tabSelection` WHERE IFNULL(batch,'') != '' ORDER BY batch DESC", as_dict=True)]
	return {
		"selections": [{**r, "selection_date": str(r.selection_date or "")} for r in sels],
		"items": [{**r, "selection_date": str(r.selection_date or "")} for r in items],
		"parties": parties, "batches": batches,
		"total_selections": len(sels), "total_photos": len(items),
		"total_gold": round(sum(flt(r.gold_gms) for r in items), 3),
		"total_cts": round(sum(flt(r.cts) for r in items), 3),
		"unique_photos": len({r.photo for r in items}),
	}


@frappe.whitelist()
def update_selection(name, photos):
	"""Second pass: keep only these photos on an existing Selection. Totals
	recompute from the remaining lines (controller). Empty is refused — delete the
	record instead if the party dropped everything."""
	if isinstance(photos, str):
		photos = json.loads(photos or "[]")
	photos = [p for p in photos if p]
	if not frappe.db.exists("Selection", name):
		frappe.throw(frappe._("Selection {0} not found.").format(name))
	if not photos:
		frappe.throw(frappe._("Nothing left — remove the whole Selection instead."))
	doc = frappe.get_doc("Selection", name)
	doc.set("items", [{"photo": p} for p in photos])
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "total_photos": doc.total_photos,
		"total_gold": doc.total_gold, "total_cts": doc.total_cts}


@frappe.whitelist()
def get_recent_selections(limit=15):
	"""Latest selection records for the page's side list."""
	return frappe.get_all("Selection", fields=["name", "party", "selection_date", "batch", "total_photos", "total_gold", "total_cts"],
		order_by="creation desc", limit_page_length=cint(limit) or 15)


@frappe.whitelist()
def export_table_xlsx(title, data):
	"""Generic table -> xlsx download. `data` = [[header...], [row...], ...] exactly
	as the page shows it (visible columns, current filter + sort). Streams the file."""
	from frappe.utils.xlsxutils import build_xlsx_response

	if isinstance(data, str):
		data = json.loads(data or "[]")
	build_xlsx_response(data, (title or "export"))


def _render_table_png(data, heading=None):
	"""Render [[header],[row]...] to PNG bytes (Pillow, no browser). Header row is
	styled, the LAST row is a totals row, numeric columns right-align."""
	from io import BytesIO
	from PIL import Image, ImageDraw, ImageFont

	if isinstance(data, str):
		data = json.loads(data or "[]")
	if not data:
		frappe.throw(frappe._("Nothing to export."))

	def font(sz, bold=False):
		for name in ((["DejaVuSans-Bold.ttf"] if bold else ["DejaVuSans.ttf"])):
			try:
				return ImageFont.truetype(name, sz)
			except Exception:
				pass
		return ImageFont.load_default()

	f_head = font(15, bold=True)
	f_cell = font(15)
	f_tot = font(15, bold=True)
	f_title = font(20, bold=True)

	ncols = max(len(r) for r in data)
	rows = [list(r) + [""] * (ncols - len(r)) for r in data]
	pad_x, pad_y, line_h = 16, 10, 34
	scratch = ImageDraw.Draw(Image.new("RGB", (1, 1)))

	def text_w(txt, fnt):
		return scratch.textbbox((0, 0), str(txt), font=fnt)[2]

	# column widths = widest cell + padding
	col_w = []
	for c in range(ncols):
		w = 0
		for ri, r in enumerate(rows):
			fnt = f_head if ri == 0 else f_cell
			w = max(w, text_w(r[c], fnt))
		col_w.append(w + pad_x * 2)

	# which columns are numeric (right-align) — sample non-header rows
	def is_num(v):
		try:
			float(str(v).replace(",", ""))
			return str(v).strip() != ""
		except Exception:
			return False
	num_col = []
	for c in range(ncols):
		vals = [rows[ri][c] for ri in range(1, len(rows)) if str(rows[ri][c]).strip() != ""]
		num_col.append(bool(vals) and all(is_num(v) for v in vals))

	title_h = 40 if heading else 0
	W = sum(col_w) + 2
	H = title_h + line_h * len(rows) + 2 + pad_y

	img = Image.new("RGB", (W, H), "#ffffff")
	d = ImageDraw.Draw(img)
	y = 0
	if heading:
		d.text((4, 8), heading, fill="#111111", font=f_title)
		y = title_h

	for ri, r in enumerate(rows):
		is_header = ri == 0
		is_totals = ri == len(rows) - 1 and len(rows) > 2
		bg = "#e9ecef" if is_header else ("#f5f5f5" if is_totals else "#ffffff")
		d.rectangle([1, y, W - 1, y + line_h], fill=bg, outline="#bbbbbb")
		x = 1
		for c in range(ncols):
			cw = col_w[c]
			d.rectangle([x, y, x + cw, y + line_h], outline="#cccccc")
			fnt = f_head if is_header else (f_tot if is_totals else f_cell)
			txt = str(r[c])
			tw = text_w(txt, fnt)
			tx = (x + cw - pad_x - tw) if (num_col[c] and not is_header) else (x + pad_x)
			d.text((tx, y + 8), txt, fill="#111111", font=fnt)
			x += cw
		y += line_h

	buf = BytesIO()
	img.save(buf, "PNG")
	return buf.getvalue()


@frappe.whitelist()
def export_table_image(title, data, heading=None):
	"""Generic table -> PNG download."""
	png = _render_table_png(data, heading)
	frappe.local.response.filename = "{0}.png".format(title or "export")
	frappe.local.response.filecontent = png
	frappe.local.response.type = "download"


@frappe.whitelist()
def attach_table_image_to_card(order_bag, data, title=None, heading=None, remarks=None):
	"""Render the table to PNG and attach it into the card's photos (the Order Bag
	Attachment table) — same store the Card page shows."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Card {0} not found.").format(order_bag))
	png = _render_table_png(data, heading)
	fname = "{0}-{1}.png".format(title or "table", frappe.utils.now().replace(" ", "_").replace(":", ""))
	f = frappe.get_doc({
		"doctype": "File", "file_name": fname, "content": png, "is_private": 0,
		"attached_to_doctype": "Order Bag", "attached_to_name": order_bag,
	}).insert(ignore_permissions=True)
	bag = frappe.get_doc("Order Bag", order_bag)
	bag.append("attachments", {"image": f.file_url, "title": (title or "Weight Check"),
		"remarks": remarks or heading or ""})
	bag.save(ignore_permissions=True)
	frappe.db.commit()
	return {"order_bag": order_bag, "file_url": f.file_url, "count": len(bag.attachments)}


def _cad_rows(payload):
	rows = [r for r in (payload.get("rows") or []) if cint(r.get("qty"))]
	tot_qty = sum(cint(r.get("qty")) for r in rows)
	tot_ct = round(sum(flt(r.get("total")) for r in rows), 4)
	return rows, tot_qty, tot_ct


def _cad_image_any(ref):
	"""A PIL image from either a base64 data-URL or a stored /files/ url. Used by
	the composite so a saved CAD Sheet (stored image) re-renders like a fresh one."""
	if not ref:
		return None
	if ref.startswith("data:") or (len(ref) > 80 and "/" not in ref[:24]):
		return _cad_image_from_b64(ref)
	try:
		from io import BytesIO
		from PIL import Image
		try:
			content = frappe.get_doc("File", {"file_url": ref}).get_content()
		except Exception:
			# design-bank images have no File record — read straight from disk
			from urllib.parse import unquote
			path = frappe.get_site_path("public", unquote(ref).lstrip("/"))
			content = open(path, "rb").read()
		return Image.open(BytesIO(content)).convert("RGB")
	except Exception:
		return None


def _cad_image_from_b64(b64):
	"""Decode the browser-supplied image (data URL or bare base64). NEVER stored —
	it lives only for the duration of this request."""
	if not b64:
		return None
	import base64
	from io import BytesIO
	from PIL import Image
	if "," in b64 and b64.strip().startswith("data:"):
		b64 = b64.split(",", 1)[1]
	try:
		return Image.open(BytesIO(base64.b64decode(b64))).convert("RGB")
	except Exception:
		return None


def _cad_compose(payload):
	"""Composite the CAD sheet -> PIL Image (design image + header + stone table +
	notes + sub-images). Uploaded images are base64 in the payload — never stored."""
	from PIL import Image, ImageDraw, ImageFont

	rows, tot_qty, tot_ct = _cad_rows(payload)
	photo = _cad_image_any(payload.get("image_b64"))
	subs = [im for im in (_cad_image_any(b) for b in (payload.get("sub_images") or [])) if im]
	payload["approver"] = frappe.utils.get_fullname(frappe.session.user)

	def font(sz, bold=False):
		try:
			return ImageFont.truetype("DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf", sz)
		except Exception:
			return ImageFont.load_default()
	f_h, f_b, f_c, f_t = font(15), font(18, True), font(15), font(22, True)

	scratch = ImageDraw.Draw(Image.new("RGB", (1, 1)))
	tw = lambda t, fn: scratch.textbbox((0, 0), str(t), font=fn)[2]

	IMG_W, PAD = 380, 24
	header = [
		("DESIGN NUMBER", payload.get("style_no") or ""),
		("DESIGN TYPE", payload.get("design_type") or ""),
		("KARAT", payload.get("karat") or ""),
		("APPROX. GOLD WT", (payload.get("gold_wt") or "") and f"{payload.get('gold_wt')} Gms"),
		("DIAMOND WT", (payload.get("dia_wt") or "") and f"{payload.get('dia_wt')} cts"),
		("LENGTH", payload.get("length") or ""),
	]
	cols = [("COL", "col"), ("S.Size", "sieve"), ("SIZE/MM", "mm"), ("WT/CT", "wt"), ("Qty", "qty"), ("Total WT", "total")]
	tbl = [[c[0] for c in cols]]
	for r in rows:
		tbl.append([r.get("col") or "", r.get("sieve") or "", r.get("mm") or "", r.get("wt") or "", cint(r.get("qty")), round(flt(r.get("total")), 4)])
	tbl.append(["", "", "", "TOTAL", tot_qty, tot_ct])
	ncol = len(cols)
	cw = []
	for c in range(ncol):
		w = max(tw(tbl[ri][c], f_b if ri == 0 else f_c) for ri in range(len(tbl)))
		cw.append(max(w + 24, 60 if c == 0 else 0))
	tbl_w = sum(cw)
	line_h = 32

	top = 44
	right_x = PAD + IMG_W + 40
	table_h = line_h * len(tbl)
	img_h = int(photo.height * (IMG_W / photo.width)) if photo else 0

	# sub-images: laid out in the LEFT column below the main photo, 2 across
	SUB_W, SUB_GAP = 182, 16
	sub_thumbs = []
	for im in subs:
		th = int(im.height * (SUB_W / im.width))
		sub_thumbs.append((im.resize((SUB_W, th)), th))
	sub_rows = (len(sub_thumbs) + 1) // 2
	sub_block_h = 0
	if sub_thumbs:
		# height = sum of the taller thumb in each pair-row + gaps
		for ri in range(sub_rows):
			pair = sub_thumbs[ri * 2:ri * 2 + 2]
			sub_block_h += max(t[1] for t in pair) + SUB_GAP

	manual = [str(m) for m in (payload.get("manual_lines") or []) if str(m).strip()]
	manual_h = (26 * len(manual) + 30) if manual else 0

	left_h = top + img_h + (20 + sub_block_h if sub_thumbs else 0)
	right_h = top + 30 * len(header) + 16 + table_h + 50 + manual_h
	W = int(max(right_x + tbl_w + PAD, PAD + IMG_W + PAD))
	H = int(max(left_h, right_h, top + 60) + 40)

	img = Image.new("RGB", (W, H), "#ffffff")
	d = ImageDraw.Draw(img)
	d.text((PAD, 10), payload.get("style_no") or "CAD SHEET", fill="#111", font=f_t)

	if photo:
		img.paste(photo.resize((IMG_W, img_h)), (PAD, top))
		d.rectangle([PAD, top, PAD + IMG_W, top + img_h], outline="#999")

	# sub-images grid (left, below main)
	sy = top + img_h + 20
	for ri in range(sub_rows):
		pair = sub_thumbs[ri * 2:ri * 2 + 2]
		sx = PAD
		for thumb, th in pair:
			img.paste(thumb, (sx, sy))
			d.rectangle([sx, sy, sx + SUB_W, sy + th], outline="#bbb")
			sx += SUB_W + SUB_GAP
		sy += max(t[1] for t in pair) + SUB_GAP

	hy = top
	for label, val in header:
		d.text((right_x, hy), f"{label} :", fill="#666", font=f_h)
		d.text((right_x + 190, hy), str(val or ""), fill="#111", font=f_b)
		hy += 30
	hy += 16

	ty = hy
	for ri, row in enumerate(tbl):
		is_head = ri == 0
		is_tot = ri == len(tbl) - 1
		bg = "#e9ecef" if is_head else ("#f5f5f5" if is_tot else "#ffffff")
		d.rectangle([right_x, ty, right_x + tbl_w, ty + line_h], fill=bg, outline="#bbb")
		x = right_x
		for c in range(ncol):
			d.rectangle([x, ty, x + cw[c], ty + line_h], outline="#ccc")
			fn = f_b if is_head else (font(15, True) if is_tot else f_c)
			t = str(row[c])
			if c == 0 and not is_head and t.startswith("#") and len(t) in (4, 7):
				d.rectangle([x + 10, ty + 6, x + cw[c] - 10, ty + line_h - 6], fill=t, outline="#888")
			else:
				align_r = c >= 3 and not is_head
				tx = (x + cw[c] - 12 - tw(t, fn)) if align_r else (x + 12)
				d.text((tx, ty + 7), t, fill="#111", font=fn)
			x += cw[c]
		ty += line_h

	ay = ty
	d.text((right_x, ay + 14), "Created by: " + str(payload.get("approver")), fill="#111", font=f_b)
	ay += 40
	if manual:
		my = ay + 16
		d.text((right_x, my), "OTHER / NOTES", fill="#666", font=f_h)
		my += 24
		for ln in manual:
			d.text((right_x, my), "\u2022 " + ln, fill="#111", font=f_c)
			my += 26
	return img


def _cad_store_image(ref, attach_to):
	"""Persist an image the page holds. A data-URL becomes a File attached to the
	CAD Sheet; an existing /files/ url is kept as-is. Returns the file_url."""
	if not ref:
		return ""
	if not ref.startswith("data:"):
		return ref
	import base64
	head, b64 = ref.split(",", 1)
	ext = "png"
	if "jpeg" in head or "jpg" in head:
		ext = "jpg"
	f = frappe.get_doc({"doctype": "File", "file_name": "cadimg-{0}.{1}".format(frappe.generate_hash(length=8), ext),
		"content": base64.b64decode(b64), "is_private": 0,
		"attached_to_doctype": "CAD Sheet Record", "attached_to_name": attach_to}).insert(ignore_permissions=True)
	return f.file_url


def _cad_sheet_name_for_bag(bag):
	"""The CAD Sheet Record driving this bag — as its primary bag or one of its
	'also linked' bags. Returns the record name, or None."""
	if not bag:
		return None
	name = frappe.db.get_value("CAD Sheet Record", {"order_bag": bag}, "name")
	if name:
		return name
	return frappe.db.get_value(
		"CAD Sheet Order Bag", {"order_bag": bag, "parenttype": "CAD Sheet Record"}, "parent")


@frappe.whitelist()
def get_cad_sheet(order_bag):
	"""The saved CAD Sheet for an order bag (edit-and-pull-again). {} if none yet.
	Matches whether the bag is the sheet's primary bag or one of its linked bags."""
	name = _cad_sheet_name_for_bag(order_bag)
	if not name:
		return {}
	d = frappe.get_doc("CAD Sheet Record", name)
	return {
		"name": d.name, "order_bag": d.order_bag, "design": d.design,
		"linked_bags": [r.order_bag for r in d.order_bags if r.order_bag],
		"style_no": d.design_number, "design_type": d.design_type, "karat": d.karat,
		"gold_wt": d.gold_wt, "length": d.length, "dia_wt": d.diamond_wt,
		"image_url": d.main_image or "", "rendered": d.rendered_sheet or "",
		"stones": [{"col": r.colour, "sieve": r.sieve, "mm": r.mm, "wt": r.wt_ct, "qty": r.qty, "total": r.total_wt} for r in d.stones],
		"manual_lines": [r.line for r in d.notes if r.line],
		"sub_image_urls": [r.image for r in d.sub_images if r.image],
		"created_by": d.owner, "created_on": str(d.creation),
		"edited_by": d.modified_by, "edited_on": str(d.modified),
	}


@frappe.whitelist()
def save_cad_sheet(payload):
	"""Upsert the CAD Sheet for an order bag (one per bag, edited in place — the
	Version history logs who/what/when). Renders the composite PNG into the record
	AND pushes it to Order Bag Photos under a deterministic name, replacing the
	previous push so only the latest sheet image sits on the card."""
	payload = frappe.parse_json(payload) if isinstance(payload, str) else payload
	order_bag = payload.get("order_bag")
	if not order_bag or not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Pick a valid Order Bag."))

	name = _cad_sheet_name_for_bag(order_bag)
	doc = frappe.get_doc("CAD Sheet Record", name) if name else frappe.new_doc("CAD Sheet Record")
	if not doc.get("order_bag"):
		doc.order_bag = order_bag          # primary bag is set once, on creation
	if payload.get("design") and frappe.db.exists("Design", payload["design"]):
		doc.design = payload["design"]
	elif frappe.db.get_value("Order Bag", order_bag, "design"):
		doc.design = frappe.db.get_value("Order Bag", order_bag, "design")
	doc.design_number = payload.get("style_no") or ""
	doc.design_type = payload.get("design_type") or None
	doc.karat = payload.get("karat") or ""
	doc.gold_wt = payload.get("gold_wt") or ""
	doc.length = payload.get("length") or ""
	doc.diamond_wt = payload.get("dia_wt") or ""
	if not doc.name:
		doc.insert(ignore_permissions=True)          # need a name to attach files
	doc.main_image = _cad_store_image(payload.get("image_b64"), doc.name)
	doc.set("stones", [{"colour": r.get("col"), "sieve": r.get("sieve"), "mm": flt(r.get("mm")),
		"wt_ct": flt(r.get("wt")), "qty": cint(r.get("qty")), "total_wt": flt(r.get("total"))}
		for r in (payload.get("rows") or [])])
	doc.set("notes", [{"line": ln} for ln in (payload.get("manual_lines") or []) if ln])
	doc.set("sub_images", [{"image": _cad_store_image(b, doc.name)} for b in (payload.get("sub_images") or []) if b])

	# additional order bags this same sheet drives (same design in production).
	# dedupe, drop the primary, and refuse a bag already owned by a different sheet.
	linked, seen = [], {doc.order_bag}
	candidates = list(payload.get("linked_bags") or [])
	if order_bag != doc.order_bag:
		candidates.insert(0, order_bag)   # keep the bag being worked on linked
	for lb in candidates:
		if not lb or lb in seen or not frappe.db.exists("Order Bag", lb):
			continue
		owner = _cad_sheet_name_for_bag(lb)
		if owner and owner != doc.name:
			frappe.throw(frappe._("Order Bag {0} is already on CAD Sheet {1}.").format(lb, owner))
		seen.add(lb)
		linked.append({"order_bag": lb})
	doc.set("order_bags", linked)

	# render the composite from what we just stored (stored urls re-render via _cad_image_any)
	from io import BytesIO
	render_payload = dict(payload)
	render_payload["image_b64"] = doc.main_image
	render_payload["sub_images"] = [r.image for r in doc.sub_images]
	buf = BytesIO()
	_cad_compose(render_payload).save(buf, "PNG")
	rf = frappe.get_doc({"doctype": "File", "file_name": "CADSHEET-{0}.png".format(doc.name),
		"content": buf.getvalue(), "is_private": 0,
		"attached_to_doctype": "CAD Sheet Record", "attached_to_name": doc.name}).insert(ignore_permissions=True)
	doc.rendered_sheet = rf.file_url
	doc.save(ignore_permissions=True)

	# push to Order Bag Photos — replace the prior CADSHEET-<name> image on EVERY
	# bag this sheet drives (primary + linked); each gets the same latest image.
	tag = "CADSHEET-{0}".format(doc.name)
	all_bags = [doc.order_bag] + [r.order_bag for r in doc.order_bags if r.order_bag]
	pushed_url = ""
	for b in all_bags:
		for old in frappe.get_all("File", filters={"attached_to_doctype": "Order Bag",
				"attached_to_name": b, "file_name": ["like", tag + "%"]}, pluck="name"):
			frappe.delete_doc("File", old, force=True, ignore_permissions=True)
		bag = frappe.get_doc("Order Bag", b)
		bag.set("attachments", [a for a in bag.attachments if not (a.title or "").startswith(tag)])
		pushed = frappe.get_doc({"doctype": "File", "file_name": tag + ".png",
			"content": buf.getvalue(), "is_private": 0,
			"attached_to_doctype": "Order Bag", "attached_to_name": b}).insert(ignore_permissions=True)
		bag.append("attachments", {"image": pushed.file_url, "title": tag, "remarks": "CAD sheet"})
		bag.save(ignore_permissions=True)
		if b == order_bag:
			pushed_url = pushed.file_url
	frappe.db.commit()
	return {"name": doc.name, "rendered": doc.rendered_sheet,
		"pushed_to_card": pushed_url, "bags": all_bags}


@frappe.whitelist()
def export_cad_sheet_image(payload):
	"""CAD sheet -> PNG (composite; images base64, never stored)."""
	from io import BytesIO
	payload = frappe.parse_json(payload) if isinstance(payload, str) else payload
	buf = BytesIO()
	_cad_compose(payload).save(buf, "PNG")
	frappe.local.response.filename = "{0}.png".format((payload.get("style_no") or "cad-sheet").replace(" ", "-"))
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "download"


@frappe.whitelist()
def export_cad_sheet_pdf(payload):
	"""CAD sheet -> single-page PDF of the same composite."""
	from io import BytesIO
	payload = frappe.parse_json(payload) if isinstance(payload, str) else payload
	buf = BytesIO()
	_cad_compose(payload).convert("RGB").save(buf, "PDF", resolution=150)
	frappe.local.response.filename = "{0}.pdf".format((payload.get("style_no") or "cad-sheet").replace(" ", "-"))
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "download"


@frappe.whitelist()
def attach_cad_sheet_to_card(order_bag, payload):
	"""Render the CAD sheet composite and attach it into the card's photos (Order
	Bag Attachment) — the same store Order Bag Photos shows. Image is base64 in
	the payload; the rendered PNG is stored as the card's attachment."""
	from io import BytesIO
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Card {0} not found.").format(order_bag))
	payload = frappe.parse_json(payload) if isinstance(payload, str) else payload
	buf = BytesIO()
	_cad_compose(payload).save(buf, "PNG")
	stamp = frappe.utils.now().replace(" ", "_").replace(":", "")
	fname = "cadsheet-{0}-{1}.png".format((payload.get("style_no") or "sheet").replace(" ", "-"), stamp)
	f = frappe.get_doc({"doctype": "File", "file_name": fname, "content": buf.getvalue(),
		"is_private": 0, "attached_to_doctype": "Order Bag", "attached_to_name": order_bag}).insert(ignore_permissions=True)
	bag = frappe.get_doc("Order Bag", order_bag)
	bag.append("attachments", {"image": f.file_url, "title": (payload.get("style_no") or "CAD Sheet"), "remarks": "CAD sheet"})
	bag.save(ignore_permissions=True)
	frappe.db.commit()
	return {"order_bag": order_bag, "file_url": f.file_url, "count": len(bag.attachments)}


@frappe.whitelist()
def export_cad_sheet_xlsx(payload):
	"""The CAD sheet as an .xlsx — header cells, stone table, embedded design image.
	The image is base64 in the payload (never stored server-side)."""
	from io import BytesIO
	from openpyxl import Workbook
	from openpyxl.drawing.image import Image as XLImage
	from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

	payload = frappe.parse_json(payload) if isinstance(payload, str) else payload
	rows, tot_qty, tot_ct = _cad_rows(payload)
	photo = _cad_image_from_b64(payload.get("image_b64"))
	payload["approver"] = frappe.utils.get_fullname(frappe.session.user)

	wb = Workbook()
	ws = wb.active
	ws.title = "CAD Sheet"
	bold = Font(bold=True)
	hdr_fill = PatternFill("solid", fgColor="E9ECEF")
	tot_fill = PatternFill("solid", fgColor="F5F5F5")
	thin = Border(*[Side(style="thin", color="BBBBBB")] * 4)

	# header block (col A)
	ws["A1"] = payload.get("style_no") or "CAD SHEET"; ws["A1"].font = Font(bold=True, size=14)
	meta = [("DESIGN NUMBER :", payload.get("style_no")), ("APPROX. GOLD WT :", (payload.get("gold_wt") and f"{payload.get('gold_wt')} Gms")),
		("DIAMOND WT :", (payload.get("dia_wt") and f"{payload.get('dia_wt')} cts")), ("LENGTH :", payload.get("length"))]
	for i, (k, v) in enumerate(meta, start=3):
		ws.cell(row=i, column=1, value=k).font = bold
		ws.cell(row=i, column=2, value=v or "")

	# stone table (starts col H, row 1)
	cols = ["COL.", "S.Size", "SIZE/MM", "WT/CT", "Qty", "Total WT"]
	base_c = 8
	for j, c in enumerate(cols):
		cell = ws.cell(row=1, column=base_c + j, value=c)
		cell.font = bold; cell.fill = hdr_fill; cell.border = thin
	r = 2
	for row in rows:
		colhex = str(row.get("col") or "")
		vals = ["", row.get("sieve") or "", row.get("mm") or "", row.get("wt") or "", cint(row.get("qty")), round(flt(row.get("total")), 4)]
		for j, v in enumerate(vals):
			cc = ws.cell(row=r, column=base_c + j, value=v); cc.border = thin
			if j == 0 and colhex.startswith("#") and len(colhex) in (4, 7):
				cc.fill = PatternFill("solid", fgColor=colhex.lstrip("#").upper().rjust(6, colhex.lstrip("#")[0]))
		r += 1
	for j, v in enumerate(["", "", "", "TOTAL", tot_qty, tot_ct]):
		cc = ws.cell(row=r, column=base_c + j, value=v); cc.font = bold; cc.fill = tot_fill; cc.border = thin
	mr = r + 2
	if payload.get("approver"):
		ws.cell(row=mr, column=base_c + 3, value="Created by: " + str(payload.get("approver"))).font = bold
		mr += 2
	manual = [str(m) for m in (payload.get("manual_lines") or []) if str(m).strip()]
	if manual:
		ws.cell(row=mr, column=base_c, value="OTHER / NOTES").font = bold
		mr += 1
		for ln in manual:
			ws.cell(row=mr, column=base_c, value=ln)
			mr += 1

	for j in range(len(cols)):
		ws.column_dimensions[chr(ord("A") + base_c - 1 + j)].width = 12

	# embed the design image (top-left), scaled to ~360px wide
	anchor_row = 10
	def _embed(pil, cell, w):
		nonlocal anchor_row
		pil = pil.resize((w, int(pil.height * (w / pil.width))))
		bio = BytesIO(); pil.save(bio, "PNG"); bio.seek(0)
		ws.add_image(XLImage(bio), cell)
	if photo:
		_embed(photo, "A" + str(anchor_row), 360); anchor_row += 20
	for sub in [im for im in (_cad_image_from_b64(b) for b in (payload.get("sub_images") or [])) if im]:
		_embed(sub, "A" + str(anchor_row), 240); anchor_row += 14

	out = BytesIO(); wb.save(out)
	frappe.local.response.filename = "{0}.xlsx".format((payload.get("style_no") or "cad-sheet").replace(" ", "-"))
	frappe.local.response.filecontent = out.getvalue()
	frappe.local.response.type = "binary"


# --- CAD Workstation (Workstation) — CAD users see their assigned cards; a lead
# (System Manager) assigns queue cards. Read-only dashboard + assign. Approval
# workflow: deferred.
def _cad_karat(bag):
	"""'22K' / '18K' etc. from the card's cad_karat, else its design's gold row."""
	import re as _re
	k = bag.get("cad_karat") or ""
	if not k and bag.get("design"):
		gold = frappe.db.sql_list("""SELECT bi.item FROM `tabDesign BOM Item` bi
			JOIN `tabItem` i ON i.name = bi.item
			WHERE bi.parent = %s AND i.material_group = 'GOLD' LIMIT 1""", bag.get("design"))
		k = gold[0] if gold else ""
	m = _re.match(r"^(\d+)K", k or "")
	return f"{m.group(1)}K" if m else ""


def _my_employee():
	return frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")


def _cad_roster():
	b = frappe.db.get_value("Bench", "CAD", "name") or frappe.db.get_value("Bench", {"bench_name": "CAD"}, "name")
	return frappe.get_all("Bench Employee", filters={"parent": b}, pluck="employee") if b else []


@frappe.whitelist()
def get_cad_workstation():
	"""The Workstation: CAD queue count, the logged-in user's assigned cards, the
	per-user summary, and (for a lead) the unassigned queue to hand out."""
	me = _my_employee()
	roster = _cad_roster()
	is_lead = "System Manager" in set(frappe.get_roles())
	bags = frappe.get_all("Order Bag", filters={"is_cad": 1},
		fields=["name", "customer", "order_date", "due_date", "job_order", "cad_design_type", "qty",
			"cad_karat", "design", "cad_gold_weight", "cad_diamond_weight"],
		order_by="due_date asc, creation asc", limit_page_length=0)
	jd = _jd_stock_customer()
	names = [b.name for b in bags] or [""]
	assign = {}
	for r in frappe.get_all("CAD", filters={"order_bag": ["in", names], "status": ["not in", ["Completed", "Expired"]]},
			fields=["order_bag", "employee"], order_by="creation desc"):
		assign.setdefault(r.order_bag, r.employee)
	empname = {e.name: e.employee_name for e in frappe.get_all("Employee",
		filters={"name": ["in", (roster or []) + [me] if me else roster or [""]]}, fields=["name", "employee_name"])}
	counts = {e: 0 for e in roster}
	my_cards, unassigned = [], []
	for b in bags:
		emp = assign.get(b.name)
		row = dict(b)
		row["employee"] = emp
		row["employee_name"] = empname.get(emp, "")
		row["order_date"] = str(b.order_date or "")
		row["due_date"] = str(b.due_date or "")
		row["karat"] = _cad_karat(b)
		row["is_bulk"] = 1 if (b.customer and b.customer == jd) else 0
		if emp:
			counts[emp] = counts.get(emp, 0) + 1
		if me and emp == me:
			my_cards.append(row)
		elif not emp:
			unassigned.append(row)
	cad_users = sorted(
		[{"employee": e, "name": empname.get(e, e), "count": counts.get(e, 0), "is_me": e == me} for e in roster],
		key=lambda x: -x["count"])
	return {"in_queue": len(bags), "assigned_total": sum(counts.values()),
		"my_cards": my_cards, "unassigned": unassigned if is_lead else [],
		"cad_users": cad_users, "is_lead": is_lead, "my_employee": me,
		"my_name": empname.get(me, "") if me else ""}


@frappe.whitelist()
def assign_cad_card(order_bag, employee=None):
	"""Lead assigns (or reassigns / clears) a CAD queue card to a CAD user."""
	frappe.only_for(("System Manager",))
	if employee and employee not in _cad_roster():
		frappe.throw(frappe._("{0} isn't on the CAD bench.").format(employee))
	rec = frappe.get_all("CAD", filters={"order_bag": order_bag, "status": ["not in", ["Completed", "Expired"]]},
		order_by="creation desc", limit=1, pluck="name")
	if rec:
		frappe.db.set_value("CAD", rec[0], "employee", employee or None)
	else:
		frappe.get_doc({"doctype": "CAD", "order_bag": order_bag, "bench": "CAD",
			"status": "In Queue", "employee": employee or None,
			"time_in": frappe.utils.now_datetime()}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"order_bag": order_bag, "employee": employee or ""}


@frappe.whitelist()
def get_cad_card_detail(order_bag):
	"""One card's materials + CAD brief + photos, for the Workstation dialogs."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Card {0} not found.").format(order_bag))
	doc = frappe.get_doc("Order Bag", order_bag)
	# materials come from the DESIGN's BOM — an existing/made design often comes to
	# CAD only for a size change, so the design recipe is what CAD works against.
	# Fall back to the bag's own BOM if it isn't tied to a design.
	if doc.design and frappe.db.exists("Design", doc.design):
		dm = frappe.get_doc("Design", doc.design).materials
		materials = [{"item": m.item, "qty": flt(m.qty), "weight": flt(m.weight)} for m in dm if m.item]
		mat_source = "design"
	else:
		materials = [{"item": m.item, "qty": flt(m.qty), "weight": flt(m.weight)} for m in doc.bag_bom]
		mat_source = "bag"
	brief = {k: doc.get(k) for k in ("cad_design_type", "cad_karat", "cad_gold_weight",
		"cad_diamond_weight", "cad_stone_no", "cad_reference", "cad_remarks")}
	photos = []
	if doc.image:
		photos.append({"image": doc.image, "title": "Design"})
	for a in doc.attachments:
		if a.image:
			photos.append({"image": a.image, "title": a.title or ""})
	return {"name": order_bag, "design": doc.design, "customer": doc.customer,
		"materials": materials, "mat_source": mat_source, "brief": brief, "photos": photos,
		"design_type": doc.cad_design_type or "",
		"karat": _cad_karat({"cad_karat": doc.cad_karat, "design": doc.design}),
		"gold_wt": doc.cad_gold_weight or "", "dia_wt": doc.cad_diamond_weight or ""}


@frappe.whitelist()
def get_bench_board(bench):
	"""One bench's info board (no actions). Returns every card sitting there with
	its OWN stock (gold g, pure g, stone buckets), salesman, party, type, status —
	the page filters + rolls up client-side, so any filter is instant."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	bench = (bench or "").upper()
	if bench not in BENCH_DOCTYPE:
		frappe.throw(frappe._("Unknown bench: {0}").format(bench or "?"))

	bags = frappe.db.sql("""
		SELECT b.name, b.design, b.qty, b.due_date, jo.customer, jo.salesman, jo.order_type,
			IFNULL(d.design_type, IFNULL(b.cad_design_type, '')) design_type
		FROM `tabOrder Bag` b
		LEFT JOIN `tabJob Order` jo ON jo.name = b.job_order
		LEFT JOIN `tabDesign` d ON d.name = b.design
		WHERE b.location = %s AND b.stock_status = 'In Production' AND b.is_finished = 0
	""", bench, as_dict=True)
	names = [b.name for b in bags]

	# each card's LATEST work record decides its status; no record yet = In Queue
	dt = BENCH_DOCTYPE[bench]
	got = {}
	if names and frappe.db.exists("DocType", dt):
		for r in frappe.db.sql("""
			SELECT t.order_bag, t.status FROM `tab{0}` t
			JOIN (SELECT order_bag, MAX(creation) mc FROM `tab{0}`
				WHERE order_bag IN %(bags)s GROUP BY order_bag) x
			ON x.order_bag = t.order_bag AND x.mc = t.creation
		""".format(dt), {"bags": tuple(names)}, as_dict=True):
			got[r.order_bag] = r.status or "In Queue"

	# per-card stock from the ledgers
	stock = {n: {"gold_g": 0.0, "pure_g": 0.0, "buckets": {}} for n in names}
	if names:
		for r in frappe.db.sql("""
			SELECT l.order_bag bag, IFNULL(i.stone_type,'') st,
				SUM(IF(l.direction='Out',-l.qty,l.qty)) q,
				SUM(IF(l.direction='Out',-l.pcs,l.pcs)) pcs,
				SUM(IF(IFNULL(i.stone_type,'')='', IF(l.direction='Out',-l.qty,l.qty) * IFNULL(i.purity_percentage,0)/100, 0)) pure
			FROM `tabBag Material Ledger` l JOIN `tabItem` i ON i.name = l.item
			WHERE l.order_bag IN %(bags)s GROUP BY l.order_bag, i.stone_type
		""", {"bags": tuple(names)}, as_dict=True):
			sc = stock[r.bag]
			if r.st:
				bk = (_BUCKET_OF_STONE_TYPE.get(r.st) or "poth").upper()
				e = sc["buckets"].setdefault(bk, {"pcs": 0, "ct": 0.0})
				e["pcs"] += cint(r.pcs)
				e["ct"] += flt(r.q)
			else:
				sc["gold_g"] += flt(r.q)
				sc["pure_g"] += flt(r.pure)

	rows = []
	for b in bags:
		sc = stock[b.name]
		rows.append({
			"name": b.name, "design": b.design or "", "design_type": b.design_type or "",
			"qty": cint(b.qty) or 1, "party": b.customer or "", "salesman": b.salesman or "",
			"order_type": b.order_type or "", "due": str(b.due_date or ""),
			"status": got.get(b.name, "In Queue"),
			"gold_g": round(sc["gold_g"], 3), "pure_g": round(sc["pure_g"], 3),
			"buckets": {k: {"pcs": v["pcs"], "ct": round(v["ct"], 3)}
				for k, v in sc["buckets"].items() if v["pcs"] or abs(v["ct"]) > 0.0005},
		})
	return {"bench": bench, "rows": rows}


@frappe.whitelist()
def get_stone_issues_day(date, stone_type=None, item_group=None):
	"""Stone Issues report: everything issued on one day, line by line (item, card,
	who, when) + totals, filterable by stone type and item group. Also returns the
	distinct types/groups seen that day so the page can build its filter pills."""
	d = str(date or frappe.utils.today())
	rows = frappe.db.sql("""
		SELECT l.item, i.stone_type, i.item_group, l.order_bag, l.pcs, l.qty ct,
			IFNULL(e.employee_name, IFNULL(l.employee, '')) who, l.datetime
		FROM `tabBag Material Ledger` l
		JOIN `tabItem` i ON i.name = l.item
		LEFT JOIN `tabEmployee` e ON e.name = l.employee
		WHERE l.entry_type = 'Stone Issue' AND l.direction = 'In' AND DATE(l.datetime) = %s
		ORDER BY l.datetime""", d, as_dict=True)
	types = sorted({r.stone_type for r in rows if r.stone_type})
	groups = sorted({r.item_group for r in rows if r.item_group})
	if stone_type:
		rows = [r for r in rows if r.stone_type == stone_type]
	if item_group:
		rows = [r for r in rows if r.item_group == item_group]
	agg = {}
	for r in rows:
		e = agg.setdefault(r.item, {"item": r.item, "stone_type": r.stone_type,
			"item_group": r.item_group, "pcs": 0, "ct": 0.0, "lines": 0})
		e["pcs"] += cint(r.pcs)
		e["ct"] += flt(r.ct)
		e["lines"] += 1
	for e in agg.values():
		e["ct"] = round(e["ct"], 3)
	return {
		"date": d, "types": types, "groups": groups,
		"rows": [{"item": r.item, "stone_type": r.stone_type, "item_group": r.item_group,
			"order_bag": r.order_bag, "pcs": cint(r.pcs), "ct": flt(r.ct), "who": r.who,
			"datetime": str(r.datetime)} for r in rows],
		"items": sorted(agg.values(), key=lambda x: -x["ct"]),
		"total_pcs": sum(cint(r.pcs) for r in rows),
		"total_ct": round(sum(flt(r.ct) for r in rows), 3),
		"cards": len({r.order_bag for r in rows}),
	}


@frappe.whitelist()
def get_stone_issue_stock():
	"""Everything sitting in the Stone Issue warehouse right now (side panel)."""
	from jewelima.setup import STONE_ISSUE_WAREHOUSE

	wh = _wh(STONE_ISSUE_WAREHOUSE)
	rows = frappe.get_all("Bin", filters={"warehouse": wh, "actual_qty": [">", 0]},
		fields=["item_code", "actual_qty"], order_by="item_code")
	return {"warehouse": wh, "items": [
		{"item": r.item_code, "ct": flt(r.actual_qty)} for r in rows
	], "total_ct": round(sum(flt(r.actual_qty) for r in rows), 3)}


@frappe.whitelist()
def book_loss(order_bag, item, qty, bench=None, employee=None, remarks=None):
	"""Record metal loss out of a bag (the out-minus-in difference at a bench).
	Per-bag ledger row AND real stock: In Bags pool -> '<bench> -LOSS' warehouse.
	`bench` may be the UPPERCASE location or the Title bench name."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE

	name = _bag_ledger(order_bag, item, "Out", qty, "Loss", bench=bench, employee=employee, remarks=remarks)
	bench_title = BENCH_DOCTYPE.get(bench, bench) if bench else None
	loss_wh = _wh("{0} -LOSS".format(bench_title)) if bench_title else None
	_stock_move(item, qty, _wh(IN_PRODUCTION_WAREHOUSE), loss_wh)
	return {"ledger": name, **get_bag_contents(order_bag)}


# ---------------------------------------------------------------------------
# Weight Add / Weight Reduce screens (single card, warehouse-aware).
# ---------------------------------------------------------------------------
def _recompute_bag_from_contents(order_bag):
	"""Sync the bag's ACTUAL weight fields to its current contents (the ledger).
	The PLAN fields are left alone (they come from the BOM)."""
	refresh_actual_weights(order_bag)


@frappe.whitelist()
def get_card_for_weight(order_bag):
	"""Header + materials (the bag's BOM merged with what it actually holds) for the
	Weight Add / Reduce screens."""
	bag = frappe.db.get_value(
		"Order Bag", order_bag,
		["name", "design", "qty", "size", "purity", "gross_weight", "nett_weight",
		 "job_order", "customer", "salesman", "order_date", "is_finished"],
		as_dict=True,
	)
	if not bag:
		return {}
	item = item_name = None
	if bag.design:
		item = frappe.db.get_value("Design", bag.design, "item")
		if item:
			item_name = frappe.db.get_value("Item", item, "item_name")
	cur = {c["item"]: c for c in get_bag_contents(order_bag)["items"]}
	mats = []
	for r in frappe.get_all(
		"Order Bag BOM Item", filters={"parent": order_bag, "parenttype": "Order Bag"},
		fields=["item", "item_name", "purity", "uom", "stone_type", "qty", "weight"], order_by="idx asc",
	):
		mats.append({
			"item": r.item, "item_name": r.item_name, "purity": r.purity, "uom": r.uom, "stone_type": r.stone_type,
			"bom_qty": r.qty, "bom_weight": r.weight,
			"cur_qty": 0, "cur_weight": flt((cur.get(r.item) or {}).get("qty")),
			"cur_pcs": cint((cur.get(r.item) or {}).get("pcs")),
		})
	return {"bag": bag, "item": item, "item_name": item_name, "materials": mats}


@frappe.whitelist()
def weight_add(order_bag, lines, from_warehouse=None):
	"""Add real weight to a card: ledger In + stock move from_warehouse -> In Bags,
	per line {item, weight}. Source defaults to the Raw Materials Store."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE, RAW_MATERIALS_STORE

	if isinstance(lines, str):
		lines = json.loads(lines or "[]")
	if frappe.db.get_value("Order Bag", order_bag, "is_finished"):
		frappe.throw(frappe._("{0} is finished — its materials are locked.").format(order_bag))
	# from_warehouse from the UI is already the full "<name> - <abbr>"; only fall back to
	# the Raw Materials Store (which needs the abbr appended) when nothing was passed.
	src = from_warehouse if (from_warehouse and frappe.db.exists("Warehouse", from_warehouse)) else _wh(RAW_MATERIALS_STORE)
	tgt = _wh(IN_PRODUCTION_WAREHOUSE)
	added = 0.0
	for ln in lines or []:
		item, wt = ln.get("item"), flt(ln.get("weight"))
		if not item or wt <= 0:
			continue
		avail = flt(frappe.db.get_value("Bin", {"item_code": item, "warehouse": src}, "actual_qty"))
		if wt > avail + 0.0005:
			frappe.throw(frappe._("Only {0} of {1} available in {2} — can't issue {3}.").format(round(avail, 3), item, src, wt))
		_bag_ledger(order_bag, item, "In", wt, "Weight Add", remarks=ln.get("remarks"))
		_stock_move(item, wt, src, tgt)
		added += wt
	_recompute_bag_from_contents(order_bag)
	frappe.db.commit()
	return {"added": round(added, 3), **get_bag_contents(order_bag)}


@frappe.whitelist()
def weight_reduce(order_bag, lines, to_warehouse=None):
	"""Remove weight from a card: ledger Out + stock move In Bags -> to_warehouse,
	per line {item, weight}. Destination defaults to the Raw Materials Store."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE, RAW_MATERIALS_STORE

	if isinstance(lines, str):
		lines = json.loads(lines or "[]")
	if frappe.db.get_value("Order Bag", order_bag, "is_finished"):
		frappe.throw(frappe._("{0} is finished — its materials are locked.").format(order_bag))
	src = _wh(IN_PRODUCTION_WAREHOUSE)
	tgt = to_warehouse if (to_warehouse and frappe.db.exists("Warehouse", to_warehouse)) else _wh(RAW_MATERIALS_STORE)
	removed = 0.0
	for ln in lines or []:
		item, wt = ln.get("item"), flt(ln.get("weight"))
		if not item or wt <= 0:
			continue
		# stones leave with their COUNT too — pcs keep the per-card story honest
		_bag_ledger(order_bag, item, "Out", wt, "Weight Reduce", pcs=cint(ln.get("pcs")), remarks=ln.get("remarks"))
		_stock_move(item, wt, src, tgt)
		removed += wt
	_recompute_bag_from_contents(order_bag)
	frappe.db.commit()
	return {"removed": round(removed, 3), **get_bag_contents(order_bag)}


@frappe.whitelist()
def get_order_for_edit(job_order):
	"""Everything the Edit Order page needs: the header + every bag with its plan
	BOM, the design's ORIGINAL materials (to show diversions), and whether the
	bag's plan is still editable (locked once the ornament is made)."""
	jo = frappe.db.get_value("Job Order", job_order,
		["name", "customer", "salesman", "order_type", "order_date", "due_date", "customer_date"], as_dict=True)
	if not jo:
		frappe.throw(frappe._("Job Order {0} not found.").format(job_order))
	bags = []
	for b in frappe.get_all("Order Bag", filters={"job_order": job_order},
			fields=["name", "design", "qty", "size", "location", "stock_status", "is_finished",
				"gross_weight", "nett_weight", "purity", "narration"], order_by="name"):
		doc = frappe.get_doc("Order Bag", b.name)
		b["bom"] = [{"item": r.item, "qty": flt(r.qty), "weight": flt(r.weight)} for r in doc.bag_bom]
		design_mats = []
		if b.design and frappe.db.exists("Design", b.design):
			d = frappe.get_doc("Design", b.design)
			design_mats = [{"item": m.item, "qty": flt(m.qty), "weight": flt(m.weight)} for m in d.materials if m.item]
		b["design_bom"] = design_mats
		b["diverged"] = sorted((r["item"], r["qty"], r["weight"]) for r in b["bom"]) != \
			sorted((r["item"], r["qty"], r["weight"]) for r in design_mats) if design_mats else bool(b["bom"])
		bags.append(b)
	return {"order": jo, "bags": bags}


@frappe.whitelist()
def get_card_for_edit(order_bag):
	"""Edit Order, card-scan flow: ONE card with its plan BOM, the design's
	original materials (to show the diversion), and whether it's locked."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Card {0} not found.").format(order_bag))
	doc = frappe.get_doc("Order Bag", order_bag)
	design_mats = []
	if doc.design and frappe.db.exists("Design", doc.design):
		d = frappe.get_doc("Design", doc.design)
		design_mats = [{"item": m.item, "qty": flt(m.qty), "weight": flt(m.weight)} for m in d.materials if m.item]
	bom = [{"item": r.item, "qty": flt(r.qty), "weight": flt(r.weight)} for r in doc.bag_bom]
	return {
		"name": doc.name, "design": doc.design, "image": doc.image, "qty": doc.qty, "size": doc.size,
		"location": doc.location, "stock_status": doc.stock_status, "is_finished": doc.is_finished,
		"customer": doc.customer, "salesman": doc.salesman, "due_date": str(doc.due_date or ""),
		"job_order": doc.job_order,
		"gross_weight": flt(doc.gross_weight), "nett_weight": flt(doc.nett_weight), "purity": flt(doc.purity),
		"bom": bom, "design_bom": design_mats,
		"diverged": sorted((r["item"], r["qty"], r["weight"]) for r in bom) != \
			sorted((r["item"], r["qty"], r["weight"]) for r in design_mats) if design_mats else bool(bom),
	}


@frappe.whitelist()
def update_order_dates(job_order, due_date=None, customer_date=None):
	"""Edit Order: move the order's dates — on the Job Order and every one of its bags."""
	if not frappe.db.exists("Job Order", job_order):
		frappe.throw(frappe._("Job Order {0} not found.").format(job_order))
	vals = {}
	if due_date:
		vals["due_date"] = due_date
	if customer_date:
		vals["customer_date"] = customer_date
	if not vals:
		return {"ok": 0}
	frappe.db.set_value("Job Order", job_order, vals)
	for bag in frappe.get_all("Order Bag", filters={"job_order": job_order}, pluck="name"):
		frappe.db.set_value("Order Bag", bag, vals, update_modified=False)
	frappe.db.commit()
	return {"ok": 1, **vals}


@frappe.whitelist()
def save_bag_bom(order_bag, rows):
	"""Replace a card's BOM (editable plan) from the Weight Add screen. Blocked once
	the ornament is made."""
	if isinstance(rows, str):
		rows = json.loads(rows or "[]")
	bag = frappe.get_doc("Order Bag", order_bag)
	if bag.is_finished:
		frappe.throw(frappe._("The BOM is locked — the ornament for {0} has already been made.").format(order_bag))
	bag.set("bag_bom", [])
	for r in rows or []:
		if not r.get("item"):
			continue
		bag.append("bag_bom", {"item": r.get("item"), "qty": r.get("qty") or 0, "weight": r.get("weight") or 0})
	bag.save(ignore_permissions=True)
	frappe.db.commit()
	return {"ok": 1, "rows": len(bag.bag_bom)}


@frappe.whitelist()
def get_in_bags_matrix():
	"""The In Bags pool, exploded: WHERE each material physically sits. Rows =
	items (gold in g, stones in ct), columns = benches holding material (from bag
	ledgers x current bag locations), plus a per-bench status split (the bench
	record's In Queue / Issued / Receipted) so you see what's out with workers."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	locs = list(BENCH_DOCTYPE)
	bags = frappe.get_all(
		"Order Bag",
		filters={"location": ["in", locs], "is_finished": 0, "stock_status": ["not in", ["Cancelled", "Sold"]]},
		fields=["name", "location"],
	)
	if not bags:
		return {"items": [], "locations": [], "totals": {}}
	bag_loc = {b.name: b.location for b in bags}
	names = list(bag_loc)

	# item x location material sums
	cells, meta = {}, {}
	for item, loc, qty, stone_type, group in frappe.db.sql(
		"""select l.item, ob.location,
		          sum(case when l.direction = 'In' then l.qty else -l.qty end),
		          ifnull(i.stone_type, ''), i.item_group
		   from `tabBag Material Ledger` l
		   join `tabOrder Bag` ob on ob.name = l.order_bag
		   join `tabItem` i on i.name = l.item
		   where l.order_bag in %(names)s
		   group by l.item, ob.location""",
		{"names": names},
	):
		if flt(qty) <= 0.0005:
			continue
		cells.setdefault(item, {})[loc] = round(flt(qty), 3)
		meta[item] = {"is_stone": bool(stone_type), "group": group or "", "bucket": stone_type or ""}

	# per-bag gold/stone totals -> the per (location, status) split
	bag_mat = {}
	for bag, gold, stones in frappe.db.sql(
		"""select l.order_bag,
		          sum(case when ifnull(i.stone_type, '') = '' then (case when l.direction = 'In' then l.qty else -l.qty end) else 0 end),
		          sum(case when ifnull(i.stone_type, '') != '' then (case when l.direction = 'In' then l.qty else -l.qty end) else 0 end)
		   from `tabBag Material Ledger` l join `tabItem` i on i.name = l.item
		   where l.order_bag in %(names)s group by l.order_bag""",
		{"names": names},
	):
		bag_mat[bag] = (max(flt(gold), 0), max(flt(stones), 0))

	# latest bench-record status per bag at its CURRENT location
	bag_status, by_loc_names = {}, {}
	for b, loc in bag_loc.items():
		by_loc_names.setdefault(loc, []).append(b)
	for loc, bnames in by_loc_names.items():
		dt = BENCH_DOCTYPE.get(loc)
		if not dt or not frappe.db.exists("DocType", dt):
			continue
		for r in frappe.get_all(dt, filters={"order_bag": ["in", bnames]},
		                        fields=["order_bag", "status"], order_by="creation asc"):
			bag_status[r.order_bag] = r.status  # ascending scan -> the latest record wins

	loc_stats = {}
	for bag, (gold, stones) in bag_mat.items():
		if gold <= 0.0005 and stones <= 0.0005:
			continue
		loc = bag_loc[bag]
		st = bag_status.get(bag) or "In Queue"
		d = loc_stats.setdefault(loc, {"gold": 0.0, "stones": 0.0, "statuses": {}})
		d["gold"] += gold
		d["stones"] += stones
		sd = d["statuses"].setdefault(st, {"gold": 0.0, "stones": 0.0, "cards": 0})
		sd["gold"] += gold
		sd["stones"] += stones
		sd["cards"] += 1

	live_locs = [loc for loc in locs if loc in loc_stats]
	items = [{
		"item": item, "group": meta[item]["group"], "is_stone": meta[item]["is_stone"],
		"bucket": meta[item]["bucket"],
		"total": round(sum(cells[item].values()), 3),
		"cells": cells[item],
	} for item in sorted(cells, key=lambda i: (meta[i]["is_stone"], i))]
	locations = [{
		"location": loc, "label": BENCH_DOCTYPE.get(loc, loc),
		"gold": round(loc_stats[loc]["gold"], 3), "stones": round(loc_stats[loc]["stones"], 3),
		"statuses": {k: {"gold": round(v["gold"], 3), "stones": round(v["stones"], 3), "cards": v["cards"]}
		             for k, v in loc_stats[loc]["statuses"].items()},
	} for loc in live_locs]
	return {
		"items": items,
		"locations": locations,
		"totals": {
			"gold": round(sum(x["gold"] for x in locations), 3),
			"stones": round(sum(x["stones"] for x in locations), 3),
			"benches": len(locations),
			"materials": len(items),
		},
	}


@frappe.whitelist()
def get_finished_stock_matrix(status="In Stock"):
	"""One slice of the finished pool, grouped by DESIGN TYPE: rows = design types,
	columns = the customer holding the pieces (held_by, JD Stock first), cells =
	piece counts (+ gold/stones/barcodes for the tooltip). status "In Stock"
	feeds the Finished Stock page, "At Certification" the At Certification page.
	Piece weights come from the finished bags' Convert rows — raw materials are
	never shown loose here."""
	if status not in ("In Stock", "At Certification"):
		frappe.throw(frappe._("Unknown status: {0}").format(status))
	bags = frappe.get_all(
		"Order Bag",
		filters={"is_finished": 1, "stock_status": status},
		fields=["name", "held_by", "design"],
	)
	if not bags:
		return {"types": [], "locations": [], "totals": {}}
	names = [b.name for b in bags]

	# per-bag gold/stone totals from the Convert (Out) rows
	bag_mat = {}
	for bag, gold, stones in frappe.db.sql(
		"""select l.order_bag,
		          sum(case when ifnull(i.stone_type, '') = '' then l.qty else 0 end),
		          sum(case when ifnull(i.stone_type, '') != '' then l.qty else 0 end)
		   from `tabBag Material Ledger` l join `tabItem` i on i.name = l.item
		   where l.order_bag in %(names)s and l.entry_type = 'Convert' and l.direction = 'Out'
		   group by l.order_bag""",
		{"names": names},
	):
		bag_mat[bag] = (flt(gold), flt(stones))

	designs = list({b.design for b in bags if b.design})
	dtype = {d.name: (d.design_type or "—") for d in frappe.get_all(
		"Design", filters={"name": ["in", designs or [""]]}, fields=["name", "design_type"])}

	# design type x holder cells + per-holder totals
	cells, trow, holder_stats = {}, {}, {}
	for b in bags:
		gold, stones = bag_mat.get(b.name, (0.0, 0.0))
		ty = dtype.get(b.design, "—")
		holder = b.held_by or "—"
		c = cells.setdefault(ty, {}).setdefault(holder, {"pc": 0, "gold": 0.0, "stones": 0.0, "bags": []})
		c["pc"] += 1
		c["gold"] += gold
		c["stones"] += stones
		if len(c["bags"]) < 8:
			c["bags"].append(b.name)
		t = trow.setdefault(ty, {"pc": 0, "gold": 0.0, "stones": 0.0})
		t["pc"] += 1
		t["gold"] += gold
		t["stones"] += stones
		d = holder_stats.setdefault(holder, {"gold": 0.0, "stones": 0.0, "cards": 0})
		d["gold"] += gold
		d["stones"] += stones
		d["cards"] += 1

	# JD Stock (our own shelf) first, then holders by pieces held
	order = sorted(holder_stats, key=lambda h: (h != "JD Stock", -holder_stats[h]["cards"], h))
	types = [{
		"design_type": ty,
		"pieces": trow[ty]["pc"],
		"gold": round(trow[ty]["gold"], 3), "stones": round(trow[ty]["stones"], 3),
		"cells": {h: {"pc": c["pc"], "gold": round(c["gold"], 3), "stones": round(c["stones"], 3), "bags": c["bags"]}
		          for h, c in cells[ty].items()},
	} for ty in sorted(trow, key=lambda t: (-trow[t]["pc"], t))]
	locations = [{
		"location": h, "label": h, "cards": holder_stats[h]["cards"],
		"gold": round(holder_stats[h]["gold"], 3), "stones": round(holder_stats[h]["stones"], 3),
	} for h in order]
	return {
		"types": types,
		"locations": locations,
		"totals": {
			"gold": round(sum(x["gold"] for x in locations), 3),
			"stones": round(sum(x["stones"] for x in locations), 3),
			"pieces": sum(x["cards"] for x in locations),
			"holders": len(locations),
			"types": len(types),
		},
	}


# ---------------------------------------------------------------------------
# Warehouse Stock dashboard — live balances straight from the stock ledger (Bin).
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_item_stock(warehouse=None):
	"""Every stock item (gold + all stones) with its balance in a chosen warehouse
	(blank = totalled across all warehouses). Shows the full item list, 0 where there
	is no stock; skips disabled/non-stock items. For the Item Stock screen."""
	bin_filters = {}
	if warehouse:
		bin_filters["warehouse"] = warehouse
	qmap = {}
	for b in frappe.get_all("Bin", filters=bin_filters, fields=["item_code", "actual_qty"]):
		qmap[b.item_code] = qmap.get(b.item_code, 0.0) + flt(b.actual_qty)
	rows = []
	for it in frappe.get_all(
		"Item", filters={"is_stock_item": 1, "disabled": 0},
		fields=["name", "item_name", "stock_uom", "stone_type"], order_by="item_name asc",
	):
		rows.append({
			"item": it.name, "item_name": it.item_name, "uom": it.stock_uom,
			"type": it.stone_type or "Metal", "qty": round(flt(qmap.get(it.name, 0)), 3),
		})
	return sorted(rows, key=lambda x: (-x["qty"], (x["item_name"] or x["item"]).lower()))


# ---------------------------------------------------------------------------
# Bag Extraction — split one N-piece bag into N individual order bags.
# ---------------------------------------------------------------------------
def _even_split(total, n, prec=3):
	"""Split `total` into `n` parts rounded to `prec` dp, summing back to `total`
	(the rounding remainder is spread over the first parts)."""
	total = flt(total)
	if n <= 0:
		return []
	base = round(total / n, prec)
	parts = [base] * n
	diff = round(total - base * n, prec)
	step = 10 ** (-prec)
	i = 0
	while abs(diff) >= step / 2 and i < n:
		parts[i] = round(parts[i] + (step if diff > 0 else -step), prec)
		diff = round(diff - (step if diff > 0 else -step), prec)
		i += 1
	return parts


def _split_counts(total, n):
	"""Split an integer count into n parts (first parts get the remainder)."""
	total = int(total or 0)
	base, rem = divmod(total, n) if n else (0, 0)
	return [base + (1 if i < rem else 0) for i in range(n)]


@frappe.whitelist()
def get_bag_for_split(order_bag):
	"""Bag Extraction scan: validate the card is AT Bag Extraction and In Queue, then
	return its header, contents and a suggested even split across its qty pieces."""
	from jewelima.jewelima.benches import bench_doctype

	bag = frappe.db.get_value(
		"Order Bag", order_bag,
		["name", "location", "design", "qty", "size", "purity", "gross_weight", "nett_weight",
		 "dmd_no", "dmd_weight", "ps_no", "ps_weight", "cs_no", "cs_weight", "job_order", "customer", "salesman"],
		as_dict=True,
	)
	if not bag:
		return {"error": frappe._("No Order Bag {0}.").format(order_bag)}
	loc = (bag.location or "").upper()
	if loc != "BAG EXTRACTION":
		return {"error": frappe._("{0} is at {1} — not at Bag Extraction.").format(order_bag, bag.location or "—")}
	dt = bench_doctype(loc)
	status = None
	if dt and frappe.db.exists("DocType", dt):
		recs = frappe.get_all(dt, filters={"order_bag": order_bag}, fields=["status"], order_by="creation desc", limit=1)
		status = recs[0].status if recs else None
	if status == "Completed":
		return {"error": frappe._("{0} has already been extracted (Completed).").format(order_bag)}
	if not status:
		return {"error": frappe._("{0} has no Bag Extraction record.").format(order_bag)}
	n = max(int(bag.qty or 1), 1)
	contents = get_bag_contents(order_bag)
	cmap = {it["item"]: flt(it["qty"]) for it in contents["items"]}  # actual available per item
	gold_total = flt(contents.get("gold_grams"))
	# item order: BOM first (gives counts + order), then any held item not in the BOM
	bom_qty = {}
	order = []
	for r in frappe.get_all("Order Bag BOM Item", filters={"parent": order_bag, "parenttype": "Order Bag"}, fields=["item", "qty"], order_by="idx asc"):
		bom_qty[r.item] = flt(r.qty)
		if r.item not in order:
			order.append(r.item)
	for it in contents["items"]:
		if it["item"] not in order:
			order.append(it["item"])
	# stone_type/uom/purity are authoritative on the Item (BOM-row fetched copies are
	# unreliable on programmatic seed). Stones split + rounded to .000; gold left empty.
	items = []
	for code in order:
		meta = frappe.db.get_value("Item", code, ["item_name", "stone_type", "stock_uom", "purity_percentage"], as_dict=True) or {}
		st = meta.get("stone_type")
		total = round(flt(cmap.get(code, 0.0)), 3)
		if st:
			cnt = _split_counts(int(round(flt(bom_qty.get(code, 0)) * n)), n)
			wt = _even_split(total, n)
			per = [{"qty": cnt[i], "weight": wt[i]} for i in range(n)]
		else:
			per = [{"qty": 0, "weight": 0.0} for _ in range(n)]
		items.append({
			"item": code, "item_name": meta.get("item_name"), "purity": meta.get("purity_percentage") or 0,
			"uom": meta.get("stock_uom") or ("Carat" if st else "Gram"),
			"stone_type": st, "is_gold": not st, "total": total, "per_piece": per,
		})
	return {
		"bag": bag, "n": n, "status": status, "gold_total": round(gold_total, 3),
		"stone_ct_total": round(flt(contents.get("stone_carats")), 3),
		"items": items,
	}


@frappe.whitelist()
def start_bag_split(order_bag, employee=None):
	"""Begin extraction: assign the logged-in employee to the Bag Extraction record,
	set it Ongoing and stamp the start time. Card must be In Queue."""
	from jewelima.jewelima.benches import bench_doctype

	dt = bench_doctype("BAG EXTRACTION")
	if not (dt and frappe.db.exists("DocType", dt)):
		frappe.throw(frappe._("Bag Extraction bench is not set up."))
	rec = frappe.get_all(dt, filters={"order_bag": order_bag}, order_by="creation desc", limit=1, fields=["name", "status", "time_in", "issued_at", "employee"])
	if not rec:
		frappe.throw(frappe._("No Bag Extraction record for {0}.").format(order_bag))
	if rec[0].status == "Completed":
		frappe.throw(frappe._("{0} has already been extracted.").format(order_bag))
	if not employee:
		employee = rec[0].employee or frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")
	now = frappe.utils.now_datetime()
	# resuming keeps the original start time + employee
	frappe.db.set_value(dt, rec[0].name, {
		"status": "Ongoing", "employee": employee,
		"issued_at": rec[0].issued_at or now, "time_in": rec[0].time_in or now,
	})
	frappe.db.commit()
	return {"ok": 1, "employee": employee, "resumed": rec[0].status == "Ongoing"}


@frappe.whitelist()
def split_bag(order_bag, pieces, employee=None):
	"""Split a Bag Extraction card into one bag per piece. Parent stays as piece 1;
	pieces 2..N are new bags (named <parent>-2 …). Distributes gold (entered) + stones
	(per piece) via the ledger, sets each bag's weights, marks the Bag Extraction
	record Completed (+ employee)."""
	from jewelima.jewelima.benches import bench_doctype

	if isinstance(pieces, str):
		pieces = json.loads(pieces or "[]")
	bag = frappe.get_doc("Order Bag", order_bag)
	if (bag.location or "").upper() != "BAG EXTRACTION":
		frappe.throw(frappe._("{0} is not at Bag Extraction.").format(order_bag))
	n = len(pieces)
	if n < 1:
		frappe.throw(frappe._("Nothing to split."))
	dt = bench_doctype("BAG EXTRACTION")
	rec = frappe.get_all(dt, filters={"order_bag": order_bag}, order_by="creation desc", limit=1, fields=["name", "status", "employee"]) if dt else []
	if not rec or rec[0].status != "Ongoing":
		frappe.throw(frappe._("Press Start first — {0} isn't in progress.").format(order_bag))
	if not employee:
		employee = rec[0].employee or frappe.db.get_value("Employee", {"user_id": frappe.session.user}, "name")

	# pieces[j].items = [{item, qty, weight}] (gold item weight already computed as
	# gross - stone weight). Build per-item amount/count arrays + a stone_type cache.
	amounts, counts, stype = {}, {}, {}
	for j, p in enumerate(pieces):
		for it in p.get("items", []):
			code = it.get("item")
			if not code:
				continue
			amounts.setdefault(code, [0.0] * n)[j] = flt(it.get("weight"))
			counts.setdefault(code, [0] * n)[j] = int(it.get("qty") or 0)
			if code not in stype:
				stype[code] = frappe.db.get_value("Item", code, "stone_type")

	# never take out more than the bag holds (no negative balances)
	cmap = {it["item"]: flt(it["qty"]) for it in get_bag_contents(order_bag)["items"]}
	for code, arr in amounts.items():
		want = round(sum(flt(x) for x in arr), 3)
		have = round(flt(cmap.get(code, 0.0)), 3)
		if want > have + 0.0005:
			frappe.throw(frappe._("Not enough {0} in the bag — assigning {1} but only {2} available.").format(code, want, have))

	def piece_fields(j):
		dmd_no = ps_no = cs_no = cz_no = 0
		dmd_w = ps_w = cs_w = cz_w = gold = 0.0
		for code, arr in amounts.items():
			w, q, st = arr[j], counts[code][j], stype.get(code)
			if st == "Diamond":
				dmd_no += q; dmd_w += w
			elif st == "Precious Stone":
				ps_no += q; ps_w += w
			elif st == "Color Stone":
				cs_no += q; cs_w += w
			elif st == "Cubic Zirconia":
				cz_no += q; cz_w += w
			else:
				gold += w
		gross = gold + (dmd_w + ps_w + cs_w + cz_w) * 0.2
		# the piece's ACTUAL weights (what it physically holds after the split)
		return {
			"qty": 1, "act_nett_weight": round(gold, 3), "act_gross_weight": round(gross, 3), "act_purity": bag.purity,
			"act_dmd_no": dmd_no, "act_dmd_weight": round(dmd_w, 3),
			"act_ps_no": ps_no, "act_ps_weight": round(ps_w, 3),
			"act_cs_no": cs_no, "act_cs_weight": round(cs_w, 3),
			"act_cz_no": cz_no, "act_cz_weight": round(cz_w, 3),
		}

	created = []
	for j in range(n):
		if j == 0:
			frappe.db.set_value("Order Bag", order_bag, piece_fields(0))
			frappe.db.set_value("Order Bag", order_bag, _plan_values(bag.bag_bom, 1))  # parent now 1 piece -> plan per-unit
			created.append(order_bag)
			continue
		child = frappe.get_doc({
			"doctype": "Order Bag",
			"split_of": order_bag, "piece_no": j,
			"job_order": bag.job_order, "design": bag.design, "size": bag.size,
			"location": "BAG EXTRACTION", "customer": bag.customer, "salesman": bag.salesman,
			"order_type": bag.order_type, "order_date": bag.order_date, "due_date": bag.due_date,
		})
		child.insert(ignore_permissions=True)
		frappe.db.set_value("Order Bag", child.name, piece_fields(j))
		for code, arr in amounts.items():
			if flt(arr[j]) > 0:
				_bag_ledger(child.name, code, "In", arr[j], "Split In", reference=order_bag, pcs=counts[code][j])
		created.append(child.name)

	# parent gives up everything beyond piece 1's share
	for code, arr in amounts.items():
		out = round(sum(flt(x) for x in arr[1:]), 3)
		if out > 0:
			_bag_ledger(order_bag, code, "Out", out, "Split Out", pcs=sum(counts[code][1:]))

	# close the Bag Extraction record
	dt = bench_doctype("BAG EXTRACTION")
	if dt and frappe.db.exists("DocType", dt):
		rec = frappe.get_all(dt, filters={"order_bag": order_bag}, order_by="creation desc", limit=1, pluck="name")
		if rec:
			vals = {"status": "Completed", "time_out": frappe.utils.now_datetime()}
			if employee:
				vals["employee"] = employee
			frappe.db.set_value(dt, rec[0], vals)
	frappe.db.commit()
	return {"created": created, "count": len(created)}


# ---------------------------------------------------------------------------
# Job Work — the bench Issue / Receipt screens (scan-driven, batch).
# ---------------------------------------------------------------------------
def _current_bench_record(dt, order_bag):
	"""The latest bench record for a bag at a bench, or None."""
	if not dt or not frappe.db.exists("DocType", dt):
		return None
	recs = frappe.get_all(dt, filters={"order_bag": order_bag}, order_by="creation desc", limit=1, pluck="name")
	return recs[0] if recs else None


@frappe.whitelist()
def get_bench_card(order_bag):
	"""Job Work scan lookup: the bag's location, its current bench record + status,
	and the gold grams it holds (the 'weight out')."""
	from jewelima.jewelima.benches import bench_doctype

	bag = frappe.db.get_value("Order Bag", order_bag, ["location", "design", "qty", "is_cad", "cad_design_type"], as_dict=True)
	if not bag:
		return {}
	dt = bench_doctype(bag.location)
	rec = None
	if dt and frappe.db.exists("DocType", dt):
		recs = frappe.get_all(
			dt, filters={"order_bag": order_bag},
			fields=["name", "status", "employee", "weight_out", "weight_in", "loss"],
			order_by="creation desc", limit=1,
		)
		rec = recs[0] if recs else None
	return {
		"order_bag": order_bag, "location": bag.location, "design": bag.design, "qty": bag.qty,
		"is_cad": int(bag.is_cad or 0), "cad_design_type": bag.cad_design_type,
		"doctype": dt, "record": rec, "status": (rec or {}).get("status"),
		"gold": flt(get_bag_contents(order_bag)["gold_grams"]),
	}


# ---------------------------------------------------------------------------
# Bench work options — per-bench Work Types (picked at issue/assign) and
# Collection States (picked at collect/receipt). Configured on Setup > Bench
# Setup > Work Types & States; options can always be renamed (the rename
# follows through to every record) but only deleted while nothing uses them.
# ---------------------------------------------------------------------------
def _bench_option_field(kind):
	return "work_type" if kind == "Work Type" else "collection_state"


def _bench_option_usage(bench, kind, value):
	"""How many bench records at this location carry the option right now."""
	from jewelima.jewelima.benches import bench_doctype
	dt = bench_doctype(bench)
	if not dt or not frappe.db.exists("DocType", dt):
		return 0
	return frappe.db.count(dt, {_bench_option_field(kind): value})


@frappe.whitelist()
def get_bench_work_options(location):
	"""The configured Work Types + Collection States for one bench (for the
	issue/assign and collect/receipt pickers). Empty lists = nothing configured,
	the pages then skip the picker."""
	loc = (location or "").upper()
	rows = frappe.get_all("Bench Work Option", filters={"bench": loc},
		fields=["name", "kind", "value"], order_by="creation")
	return {
		"work_types": [r.value for r in rows if r.kind == "Work Type"],
		"collection_states": [r.value for r in rows if r.kind == "Collection State"],
	}


@frappe.whitelist()
def get_bench_work_setup(location):
	"""Setup page: the options WITH their usage counts (rename always allowed;
	delete only at zero usage)."""
	_require_stone_issue_admin()
	loc = (location or "").upper()
	rows = frappe.get_all("Bench Work Option", filters={"bench": loc},
		fields=["name", "kind", "value"], order_by="creation")
	return {"options": [{"name": r.name, "kind": r.kind, "value": r.value,
		"in_use": _bench_option_usage(loc, r.kind, r.value)} for r in rows]}


@frappe.whitelist()
def bench_work_option_add(location, kind, value):
	_require_stone_issue_admin()
	loc, value = (location or "").upper(), (value or "").strip()
	from jewelima.jewelima.benches import BENCH_DOCTYPE
	if loc not in BENCH_DOCTYPE:
		frappe.throw(frappe._("{0} is not a bench.").format(loc or "?"))
	if kind not in ("Work Type", "Collection State"):
		frappe.throw(frappe._("Kind must be Work Type or Collection State."))
	if not value:
		frappe.throw(frappe._("Enter a value."))
	if frappe.db.exists("Bench Work Option", {"bench": loc, "kind": kind, "value": value}):
		frappe.throw(frappe._("{0} already has {1} '{2}'.").format(loc, kind, value))
	d = frappe.get_doc({"doctype": "Bench Work Option", "bench": loc, "kind": kind, "value": value}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name}


@frappe.whitelist()
def bench_work_option_rename(name, new_value):
	"""Rename an option — and carry the rename onto every bench record that used
	the old spelling, so history stays consistent."""
	_require_stone_issue_admin()
	from jewelima.jewelima.benches import bench_doctype
	new_value = (new_value or "").strip()
	if not new_value:
		frappe.throw(frappe._("Enter the new name."))
	d = frappe.get_doc("Bench Work Option", name)
	if frappe.db.exists("Bench Work Option", {"bench": d.bench, "kind": d.kind, "value": new_value, "name": ["!=", d.name]}):
		frappe.throw(frappe._("{0} already has {1} '{2}'.").format(d.bench, d.kind, new_value))
	old = d.value
	d.value = new_value
	d.save(ignore_permissions=True)
	dt = bench_doctype(d.bench)
	updated = 0
	if dt and frappe.db.exists("DocType", dt) and old != new_value:
		field = _bench_option_field(d.kind)
		for rec in frappe.get_all(dt, filters={field: old}, pluck="name"):
			frappe.db.set_value(dt, rec, field, new_value, update_modified=False)
			updated += 1
	frappe.db.commit()
	return {"renamed": old + " -> " + new_value, "records_updated": updated}


@frappe.whitelist()
def bench_work_option_delete(name):
	"""Delete an option — refused while any bench record still carries it."""
	_require_stone_issue_admin()
	d = frappe.get_doc("Bench Work Option", name)
	used = _bench_option_usage(d.bench, d.kind, d.value)
	if used:
		frappe.throw(frappe._("'{0}' is on {1} record(s) at {2} — rename it instead; delete only works when nothing uses it.").format(d.value, used, d.bench))
	frappe.delete_doc("Bench Work Option", name, ignore_permissions=True)
	frappe.db.commit()
	return {"deleted": d.value}


def _valid_bench_option(location, kind, value):
	"""A picked option must actually be configured for THIS bench (or blank)."""
	if not value:
		return None
	if not frappe.db.exists("Bench Work Option", {"bench": (location or "").upper(), "kind": kind, "value": value}):
		frappe.throw(frappe._("'{0}' is not a configured {1} for {2}.").format(value, kind, location))
	return value


@frappe.whitelist()
def issue_bench_cards(names, location, employee=None, work_type=None):
	"""Issue a batch of bags at one bench: status -> Issued, snapshot weight_out
	(gold grams), stamp issued_at (+ employee if given). Skips already-Issued cards;
	with an employee, bumps their held-weight balance."""
	from jewelima.jewelima.benches import ISSUE_RECEIPT_LOCATIONS, bench_doctype

	if isinstance(names, str):
		names = json.loads(names or "[]")
	if (location or "").upper() not in ISSUE_RECEIPT_LOCATIONS:
		frappe.throw(frappe._("Job Work (Issue / Receipt) is only for {0}.").format(", ".join(sorted(ISSUE_RECEIPT_LOCATIONS))))
	work_type = _valid_bench_option(location, "Work Type", work_type)
	dt = bench_doctype(location)
	now = frappe.utils.now_datetime()
	done, errors = [], []
	for nm in names or []:
		try:
			rec = _current_bench_record(dt, nm)
			if not rec:
				errors.append({"name": nm, "error": frappe._("No bench record at {0}").format(location)})
				continue
			doc = frappe.get_doc(dt, rec)
			if doc.status == "Issued":
				errors.append({"name": nm, "error": frappe._("Already issued")})
				continue
			gold = flt(get_bag_contents(nm)["gold_grams"])
			doc.status = "Issued"
			doc.weight_out = gold
			doc.issued_at = now
			if work_type:
				doc.work_type = work_type
			if not doc.time_in:
				doc.time_in = now
			if employee:
				doc.employee = employee
			doc.save(ignore_permissions=True)
			if employee:
				_adjust_employee_balance(employee, gold)
			done.append(nm)
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	frappe.db.commit()
	return {"count": len(done), "done": done, "errors": errors}


@frappe.whitelist()
def assign_bench_cards(names, location, employee=None, work_type=None):
	"""Assign a batch of bags at a transfer bench (CAD / Wax Injecting / Wax Cleaning):
	status -> Issued, stamp issued_at (+ employee / work type if given). Times only —
	no weight/loss."""
	from jewelima.jewelima.benches import ASSIGN_COLLECT_LOCATIONS, bench_doctype

	if isinstance(names, str):
		names = json.loads(names or "[]")
	loc = (location or "").upper()
	if loc not in ASSIGN_COLLECT_LOCATIONS:
		frappe.throw(frappe._("Assign / Collect is only for {0}.").format(", ".join(sorted(ASSIGN_COLLECT_LOCATIONS))))
	# CAD work is always owned by someone — it lands on their Workstation
	if loc == "CAD" and not employee:
		frappe.throw(frappe._("CAD cards must be assigned TO an employee — pick who takes the work."))
	work_type = _valid_bench_option(loc, "Work Type", work_type)
	dt = bench_doctype(loc)
	now = frappe.utils.now_datetime()
	done, errors = [], []
	for nm in names or []:
		try:
			rec = _current_bench_record(dt, nm)
			if not rec:
				errors.append({"name": nm, "error": frappe._("No bench record at {0}").format(location)})
				continue
			doc = frappe.get_doc(dt, rec)
			if doc.status == "Issued":
				errors.append({"name": nm, "error": frappe._("Already assigned")})
				continue
			doc.status = "Issued"
			doc.issued_at = now
			if work_type:
				doc.work_type = work_type
			if not doc.time_in:
				doc.time_in = now
			if employee:
				doc.employee = employee
			doc.save(ignore_permissions=True)
			done.append(nm)
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	frappe.db.commit()
	return {"count": len(done), "done": done, "errors": errors}


@frappe.whitelist()
def collect_bench_cards(names, location, collection_state=None):
	"""Collect a batch of assigned bags at a transfer bench: status -> Completed, stamp
	receipted_at + time_out (+ collection state — complete / failed / QC failed / … —
	if given). Times only — no weight/loss. Only assigned (Issued) cards."""
	from jewelima.jewelima.benches import ASSIGN_COLLECT_LOCATIONS, bench_doctype

	if isinstance(names, str):
		names = json.loads(names or "[]")
	loc = (location or "").upper()
	if loc not in ASSIGN_COLLECT_LOCATIONS:
		frappe.throw(frappe._("Assign / Collect is only for {0}.").format(", ".join(sorted(ASSIGN_COLLECT_LOCATIONS))))
	collection_state = _valid_bench_option(loc, "Collection State", collection_state)
	dt = bench_doctype(loc)
	now = frappe.utils.now_datetime()
	done, errors = [], []
	for nm in names or []:
		try:
			# CAD gate: a bag still awaiting its CAD design cannot be collected —
			# finalize (assign the real design) first; that clears is_cad.
			if frappe.db.get_value("Order Bag", nm, "is_cad"):
				errors.append({"name": nm, "error": frappe._("CAD design not finalized — create the design first")})
				continue
			rec = _current_bench_record(dt, nm)
			if not rec:
				errors.append({"name": nm, "error": frappe._("No bench record at {0}").format(location)})
				continue
			doc = frappe.get_doc(dt, rec)
			if doc.status != "Issued":
				errors.append({"name": nm, "error": frappe._("Not assigned (is {0})").format(doc.status)})
				continue
			doc.status = "Completed"
			doc.receipted_at = now
			doc.time_out = now
			if collection_state:
				doc.collection_state = collection_state
			doc.save(ignore_permissions=True)
			done.append(nm)
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	frappe.db.commit()
	return {"count": len(done), "done": done, "errors": errors}


@frappe.whitelist()
def receipt_bench_cards(lines, location, employee=None, collection_state=None):
	"""Receive a batch of issued bags at one bench (one employee). Per line
	{order_bag, weight_in}: loss = weight_out - weight_in, status -> Receipted,
	loss booked (per-bag ledger + In Bags -> '<bench> -LOSS' stock). The optional
	collection state (complete / failed / QC failed / …) lands on every line."""
	from jewelima.jewelima.benches import ISSUE_RECEIPT_LOCATIONS, bench_doctype

	if isinstance(lines, str):
		lines = json.loads(lines or "[]")
	if (location or "").upper() not in ISSUE_RECEIPT_LOCATIONS:
		frappe.throw(frappe._("Job Work (Issue / Receipt) is only for {0}.").format(", ".join(sorted(ISSUE_RECEIPT_LOCATIONS))))
	collection_state = _valid_bench_option(location, "Collection State", collection_state)
	dt = bench_doctype(location)
	now = frappe.utils.now_datetime()
	done, errors, total_loss = [], [], 0.0
	for ln in lines or []:
		nm = ln.get("order_bag")
		win = flt(ln.get("weight_in"))
		try:
			rec = _current_bench_record(dt, nm)
			if not rec:
				errors.append({"name": nm, "error": frappe._("No bench record")})
				continue
			doc = frappe.get_doc(dt, rec)
			if doc.status != "Issued":
				errors.append({"name": nm, "error": frappe._("Not in Issued state")})
				continue
			wout = flt(doc.weight_out)
			loss = max(wout - win, 0.0)
			issue_emp = doc.employee
			if issue_emp:
				_adjust_employee_balance(issue_emp, -wout)  # held weight returns
			doc.employee = employee or issue_emp
			doc.weight_in = win
			doc.loss = loss
			doc.status = "Receipted"
			doc.receipted_at = now
			doc.time_out = now
			if collection_state:
				doc.collection_state = collection_state
			doc.save(ignore_permissions=True)
			if loss > 0:
				book_loss(nm, _bag_gold_item(nm), loss, bench=location, employee=doc.employee)
			total_loss += loss
			done.append({"name": nm, "loss": round(loss, 3)})
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	frappe.db.commit()
	return {"count": len(done), "done": done, "errors": errors, "total_loss": round(total_loss, 3), "employee": employee}


def _bag_gold_item(order_bag):
	"""The bag's main metal item (the largest non-Carat holding) — loss is booked here."""
	golds = [it for it in get_bag_contents(order_bag)["items"] if it["uom"] != "Carat" and it["qty"] > 0]
	golds.sort(key=lambda x: -x["qty"])
	return golds[0]["item"] if golds else None


def _adjust_employee_balance(employee, delta):
	"""Running total of weight currently out with an employee (informational)."""
	if frappe.db.exists("Employee Metal Balance", employee):
		bal = frappe.get_doc("Employee Metal Balance", employee)
	else:
		bal = frappe.get_doc({"doctype": "Employee Metal Balance", "employee": employee, "current_weight": 0}).insert(ignore_permissions=True)
	bal.db_set("current_weight", round(flt(bal.current_weight) + flt(delta), 3))
	bal.db_set("last_updated", frappe.utils.now_datetime())


@frappe.whitelist()
def issue_to_employee(order_bag, employee, weight_out, bench=None, item=None):
	"""Issue a bag's piece to an employee to work (time tracked). Bumps the
	employee's running weight total. Loss is settled on receipt."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Order Bag {0} not found.").format(order_bag))
	if not employee or not frappe.db.exists("Employee", employee):
		frappe.throw(frappe._("Employee {0} not found.").format(employee))
	weight_out = flt(weight_out)
	if weight_out <= 0:
		frappe.throw(frappe._("Weight out must be greater than zero."))
	ei = frappe.get_doc({
		"doctype": "Employee Issue",
		"order_bag": order_bag, "employee": employee, "bench": bench or None,
		"item": item or _bag_gold_item(order_bag),
		"status": "Issued", "weight_out": weight_out, "time_out": frappe.utils.now_datetime(),
	}).insert(ignore_permissions=True)
	_adjust_employee_balance(employee, weight_out)
	frappe.db.commit()
	return {"issue": ei.name, "employee": employee, "weight_out": weight_out}


@frappe.whitelist()
def receive_from_employee(issue, weight_in, remarks=None):
	"""Receive the piece back. Loss = weight_out - weight_in, booked to the bag's
	metal (via book_loss). Clears the employee's running total."""
	ei = frappe.get_doc("Employee Issue", issue)
	if ei.status != "Issued":
		frappe.throw(frappe._("{0} is already returned.").format(issue))
	weight_in = flt(weight_in)
	loss = round(max(flt(ei.weight_out) - weight_in, 0), 3)
	ei.db_set("weight_in", weight_in)
	ei.db_set("time_in", frappe.utils.now_datetime())
	ei.db_set("loss", loss)
	ei.db_set("status", "Returned")
	if remarks:
		ei.db_set("remarks", remarks)
	_adjust_employee_balance(ei.employee, -flt(ei.weight_out))
	if loss > 0 and ei.item:
		_bag_ledger(ei.order_bag, ei.item, "Out", loss, "Loss", bench=ei.bench, employee=ei.employee, reference=ei.name)
	frappe.db.commit()
	return {"issue": ei.name, "loss": loss, **get_bag_contents(ei.order_bag)}


@frappe.whitelist()
def convert_to_ornament(order_bag, move_stock=True):
	"""The piece is finished: consume the bag's remaining materials (zero it out),
	one Convert (Out) row per held item. Finished-good stock posting is added with
	the coarse-stock wiring. move_stock=False when the caller already landed the
	materials in Finished Goods itself (stock import)."""
	from jewelima.setup import IN_PRODUCTION_WAREHOUSE

	c = get_bag_contents(order_bag)
	held = [it for it in c["items"] if it["qty"] > 0]
	if not held:
		frappe.throw(frappe._("{0} holds no materials to convert.").format(order_bag))
	refresh_actual_weights(order_bag)  # freeze the finished piece's actual weights before consuming
	in_bags, fg = _wh(IN_PRODUCTION_WAREHOUSE), _wh("Finished Goods")
	for it in held:
		_bag_ledger(order_bag, it["item"], "Out", it["qty"], "Convert")
		# EVERYTHING the piece holds drains from the In Bags pool into Finished Goods —
		# the warehouse book stays raw-material-denominated (grams + carats); the
		# piece dimension rides on top via the bag's Convert rows + frozen actuals
		if move_stock:
			_stock_move(it["item"], it["qty"], in_bags, fg)
	frappe.db.set_value("Order Bag", order_bag, "is_finished", 1)  # locks the BOM (plan)
	frappe.db.commit()
	return {"order_bag": order_bag, "consumed": held}


# ---------------------------------------------------------------------------
# Finished products — turn finished (qty-1) bags into stock products.
# ---------------------------------------------------------------------------
def _jd_stock_customer():
	return "JD Stock" if frappe.db.exists("Customer", "JD Stock") else None


@frappe.whitelist()
def get_makeable_bags(location=None):
	"""Qty-1 bags that hold materials and aren't products yet — ready to make."""
	filters = {"qty": 1, "is_finished": 0}
	if location:
		filters["location"] = location
	out = []
	for b in frappe.get_all(
		"Order Bag", filters=filters,
		fields=["name", "design", "customer", "location", "act_gross_weight", "act_nett_weight"],
		order_by="modified desc", limit=300,
	):
		gold = flt(get_bag_contents(b.name)["gold_grams"])
		if gold <= 0:
			continue
		b["gold"] = round(gold, 3)
		out.append(b)
	return out


@frappe.whitelist()
def get_make_product_card(order_bag):
	"""Make Products scan: validate a card BEFORE it joins the queue — must exist,
	not already a product, qty exactly 1, and it must have ACTUAL weight (weighed
	in). Returns the card's queue row; throws a specific error otherwise."""
	if not frappe.db.exists("Order Bag", order_bag):
		frappe.throw(frappe._("Card {0} not found.").format(order_bag))
	b = frappe.db.get_value("Order Bag", order_bag,
		["name", "design", "image", "qty", "location", "is_finished", "stock_status",
			"customer", "salesman", "job_order", "due_date"], as_dict=True)
	if b.is_finished:
		frappe.throw(frappe._("{0} is already a product.").format(order_bag))
	if b.stock_status in ("Cancelled", "Sold"):
		frappe.throw(frappe._("{0} is {1}.").format(order_bag, b.stock_status))
	if int(b.qty or 0) != 1:
		frappe.throw(frappe._("{0} holds qty {1} — extract/split it to single pieces first.").format(order_bag, b.qty))
	p = _actual_profile(order_bag)
	if flt(p["gross"]) <= 0:
		frappe.throw(frappe._("{0} has NO actual weight — weigh the piece in before making it a product.").format(order_bag))
	return {
		"name": b.name, "design": b.design, "image": b.image, "location": b.location,
		"customer": b.customer, "salesman": b.salesman, "job_order": b.job_order,
		"due_date": str(b.due_date or ""),
		"gross": flt(p["gross"]), "nett": flt(p["nett"]),
		"dmd_ct": flt(p["dmd_weight"]) + flt(p["pdmd_weight"]),
	}


@frappe.whitelist()
def get_extraction_cards(party=None, job_order=None, design_type=None, salesman=None):
	"""The Bag Extraction pool for the Make Products page: every unfinished card
	sitting at BAG EXTRACTION with its actual weights and whether it's READY
	(qty 1 + actual weight). Filter lists ride along for the pills."""
	filters = {"location": "BAG EXTRACTION", "is_finished": 0,
		"stock_status": ["not in", ["Cancelled", "Sold"]]}
	if party:
		filters["customer"] = party
	if job_order:
		filters["job_order"] = job_order
	if salesman:
		filters["salesman"] = salesman
	bags = frappe.get_all("Order Bag", filters=filters,
		fields=["name", "design", "qty", "customer", "salesman", "job_order", "due_date"],
		order_by="due_date asc, name asc", limit_page_length=0)
	dt_of = {}
	designs = list({b.design for b in bags if b.design})
	if designs:
		for d in frappe.get_all("Design", filters={"name": ["in", designs]}, fields=["name", "design_type"]):
			dt_of[d.name] = d.design_type
	rows = []
	for b in bags:
		b["design_type"] = dt_of.get(b.design) or ""
		if design_type and b["design_type"] != design_type:
			continue
		prof = _actual_profile(b.name)
		b["gross"] = flt(prof["gross"])
		b["nett"] = flt(prof["nett"])
		b["dmd_ct"] = flt(prof["dmd_weight"]) + flt(prof["pdmd_weight"])
		b["ready"] = 1 if (int(b.qty or 0) == 1 and b["gross"] > 0) else 0
		b["blocker"] = "" if b["ready"] else (
			frappe._("qty {0} — split first").format(b.qty) if int(b.qty or 0) != 1 else frappe._("no actual weight"))
		b["due_date"] = str(b.due_date or "")
		rows.append(b)
	# the full (unfiltered-by-me) pools for the filter pills
	pool = frappe.get_all("Order Bag", filters={"location": "BAG EXTRACTION", "is_finished": 0,
		"stock_status": ["not in", ["Cancelled", "Sold"]]},
		fields=["customer", "salesman", "job_order", "design"], limit_page_length=0)
	all_designs = list({x.design for x in pool if x.design})
	all_dts = set()
	if all_designs:
		all_dts = {d.design_type for d in frappe.get_all("Design",
			filters={"name": ["in", all_designs]}, fields=["design_type"]) if d.design_type}
	return {"rows": rows,
		"parties": sorted({x.customer for x in pool if x.customer}),
		"job_orders": sorted({x.job_order for x in pool if x.job_order}),
		"salesmen": sorted({x.salesman for x in pool if x.salesman}),
		"design_types": sorted(all_dts)}


@frappe.whitelist()
def make_products(bags):
	"""Turn selected qty-1 bags into finished stock products: consume their materials
	(gold In Bags -> Finished Goods), freeze the actual weights, set held_by (the
	order's customer, else JD Stock) and stock_status = In Stock."""
	if isinstance(bags, str):
		bags = json.loads(bags or "[]")
	jd = _jd_stock_customer()
	done, errors = [], []
	for nm in bags or []:
		try:
			bag = frappe.db.get_value("Order Bag", nm, ["qty", "is_finished", "customer"], as_dict=True)
			if not bag:
				errors.append({"name": nm, "error": frappe._("Not found")})
				continue
			if bag.is_finished:
				errors.append({"name": nm, "error": frappe._("Already a product")})
				continue
			if int(bag.qty or 0) != 1:
				errors.append({"name": nm, "error": frappe._("Qty must be 1 — extract/split it first")})
				continue
			convert_to_ornament(nm)  # consume materials -> product + freeze actuals + is_finished
			frappe.db.set_value("Order Bag", nm, {
				"stock_status": "In Stock", "held_by": bag.customer or jd,
				"in_stock_on": frappe.utils.now_datetime(),
			})
			done.append(nm)
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	frappe.db.commit()
	return {"count": len(done), "done": done, "errors": errors}


# ---------------------------------------------------------------------------
# Import Stock (Delivery) — bring pre-existing finished pieces into the system.
# ---------------------------------------------------------------------------
# Stone Type -> the bag's plan/actual bucket prefix.
_BUCKET_OF_STONE_TYPE = {
	"Diamond": "dmd", "Precious Stone": "ps", "Color Stone": "cs",
	"Cubic Zirconia": "cz", "CVD": "cvd", "Party Diamond": "pdmd", "Party Other": "poth",
}


@frappe.whitelist()
def import_finished_stock(payload):
	"""Delivery > Import Stock: vault/opening pieces become finished products in one
	shot — Job Order (type Import) + one Order Bag per piece, ledger In + Convert
	rows, frozen actuals, In Stock + held_by, HUID/cert/charge tags.

	The weight coming in MUST be backed by a stock entry (no unbacked bags):
	  mode "issue"    — materials consumed from Gold Issue / Stone Issue (already
	                    purchased there); one Material Transfer moves them to
	                    Finished Goods. Shortfalls abort with a per-item list.
	  mode "purchase" — a submitted Purchase Receipt straight into Finished Goods.

	payload = {mode, customer, supplier?, remarks?, pieces: [{design, karat,
	gold (g), gross (g), size?, huid?, certifications?, tags?: [category],
	stones?: [{item, pcs, ct}]}]}"""
	from jewelima.setup import GOLD_ISSUE_WAREHOUSE, STONE_ISSUE_WAREHOUSE

	p = frappe.parse_json(payload)
	mode = (p.get("mode") or "issue").lower()
	if mode not in ("issue", "purchase"):
		frappe.throw(frappe._("Unknown mode: {0}").format(mode))
	customer = p.get("customer")
	if not customer or not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("Pick the party holding these pieces (JD Stock = ourselves)."))
	pieces = p.get("pieces") or []
	if not pieces:
		frappe.throw(frappe._("Add at least one piece."))

	# ---- validate everything BEFORE touching stock or creating anything -------
	totals = {}  # item -> {"qty": weight, "pcs": n, "is_stone": bool}
	for i, pc in enumerate(pieces, 1):
		tag = frappe._("Row {0}").format(i)
		design, karat = pc.get("design"), pc.get("karat")
		if not design or not frappe.db.exists("Design", design):
			frappe.throw(frappe._("{0}: Design {1} not found — create it first.").format(tag, design or "?"))
		if not karat or not frappe.db.exists("Item", karat):
			frappe.throw(frappe._("{0}: gold item {1} not found.").format(tag, karat or "?"))
		if frappe.db.get_value("Item", karat, "stone_type"):
			frappe.throw(frappe._("{0}: {1} is a stone — pick the karat gold item.").format(tag, karat))
		gold, gross = flt(pc.get("gold")), flt(pc.get("gross"))
		if gold <= 0 or gross <= 0:
			frappe.throw(frappe._("{0}: enter the gold weight and the gross weight.").format(tag))
		if gross + 0.0005 < gold:
			frappe.throw(frappe._("{0}: gross ({1} g) can't be below the gold weight ({2} g).").format(tag, gross, gold))
		for s in pc.get("stones") or []:
			it, ct, pcs = s.get("item"), flt(s.get("ct")), cint(s.get("pcs"))
			if not it or not frappe.db.exists("Item", it):
				frappe.throw(frappe._("{0}: stone item {1} not found.").format(tag, it or "?"))
			if not frappe.db.get_value("Item", it, "stone_type"):
				frappe.throw(frappe._("{0}: {1} is not a stone.").format(tag, it))
			if ct <= 0 or pcs <= 0:
				frappe.throw(frappe._("{0}: {1} needs carats and a piece count.").format(tag, it))
			t = totals.setdefault(it, {"qty": 0.0, "pcs": 0, "is_stone": True})
			t["qty"] += ct
			t["pcs"] += pcs
		for tname in pc.get("tags") or []:
			if not frappe.db.exists("Charge Category", tname):
				frappe.throw(frappe._("{0}: unknown charge category {1}.").format(tag, tname))
		t = totals.setdefault(karat, {"qty": 0.0, "pcs": 0, "is_stone": False})
		t["qty"] += gold

	fg = _wh("Finished Goods")
	if not fg:
		frappe.throw(frappe._("Finished Goods warehouse is missing."))

	# ---- the backing stock entry (the gate: no entry, no bags) ----------------
	if mode == "issue":
		gold_wh, stone_wh = _wh(GOLD_ISSUE_WAREHOUSE), _wh(STONE_ISSUE_WAREHOUSE)
		short = []
		for it, t in totals.items():
			src = stone_wh if t["is_stone"] else gold_wh
			have = flt(frappe.db.get_value("Bin", {"item_code": it, "warehouse": src}, "actual_qty"))
			if have + 0.0005 < t["qty"]:
				short.append("{0}: need {1}, have {2} in {3}".format(it, round(t["qty"], 3), round(have, 3), src))
		if short:
			frappe.throw(frappe._("Not enough stock in the issue warehouses:<br>{0}<br><br>Purchase it first, or use New Purchase mode.").format("<br>".join(short)))
		se = frappe.get_doc({
			"doctype": "Stock Entry",
			"stock_entry_type": "Material Transfer",
			"company": _company(),
			"items": [{
				"item_code": it, "qty": t["qty"],
				"uom": frappe.db.get_value("Item", it, "stock_uom") or "Gram",
				"s_warehouse": stone_wh if t["is_stone"] else gold_wh,
				"t_warehouse": fg,
				"allow_zero_valuation_rate": 1,
			} for it, t in totals.items()],
		})
		se.flags.ignore_permissions = True
		se.insert()
		se.submit()
		stock_doc = se.name
	else:
		supplier = p.get("supplier") or "JD Stock"
		items = [{
			"item": it, "weight": round(t["qty"], 3), "count": t["pcs"],
			"purity": flt(frappe.db.get_value("Item", it, "purity_percentage")) if not t["is_stone"] else 0,
			"rate": 0,
		} for it, t in totals.items()]
		stock_doc = post_raw_material_purchase(supplier, fg, items=items)["name"]

	# ---- Job Order (type Import) + one finished bag per piece -----------------
	if not frappe.db.exists("Order Type", "Import"):
		frappe.get_doc({"doctype": "Order Type", "order_type_name": "Import"}).insert(ignore_permissions=True)
	jo = frappe.get_doc({
		"doctype": "Job Order",
		"order_date": frappe.utils.today(),
		"customer": customer,
		"order_type": "Import",
	})
	jo.insert(ignore_permissions=True)

	made = []
	for pc in pieces:
		karat, gold, gross = pc.get("karat"), flt(pc.get("gold")), flt(pc.get("gross"))
		purity = flt(frappe.db.get_value("Item", karat, "purity_percentage"))
		plan = {"gross_weight": gross, "nett_weight": gold, "purity": purity}
		for s in pc.get("stones") or []:
			b = _BUCKET_OF_STONE_TYPE.get(frappe.db.get_value("Item", s["item"], "stone_type") or "")
			if b:
				plan[f"{b}_no"] = cint(plan.get(f"{b}_no")) + cint(s.get("pcs"))
				plan[f"{b}_weight"] = flt(plan.get(f"{b}_weight")) + flt(s.get("ct"))
		bag = frappe.get_doc({
			"doctype": "Order Bag",
			"job_order": jo.name, "design": pc.get("design"), "qty": 1,
			"size": pc.get("size"), "customer": customer, "order_type": "Import",
			"order_date": jo.order_date, "narration": p.get("remarks"),
			"huid": (pc.get("huid") or "").strip(), "certifications": (pc.get("certifications") or "").strip(),
			"charge_categories": [{"charge_category": t} for t in pc.get("tags") or []],
			**plan,
		})
		bag.insert(ignore_permissions=True)
		_bag_ledger(bag.name, karat, "In", gold, "Gold Issue", remarks="Import", reference=stock_doc)
		for s in pc.get("stones") or []:
			_bag_ledger(bag.name, s["item"], "In", flt(s.get("ct")), "Stone Issue",
			            remarks="Import", reference=stock_doc, pcs=cint(s.get("pcs")))
		convert_to_ornament(bag.name, move_stock=False)  # the batch entry already landed stock in FG
		frappe.db.set_value("Order Bag", bag.name, {
			"stock_status": "In Stock", "held_by": customer,
			"in_stock_on": frappe.utils.now_datetime(),
			"act_gross_weight": gross,  # the physical scale weight wins over the material sum
		})
		made.append(bag.name)

	frappe.db.commit()
	return {"job_order": jo.name, "bags": made, "stock_doc": stock_doc, "mode": mode}


# ---------------------------------------------------------------------------
# Certification (Delivery > Certification) — send finished pieces to IGI /
# hallmarking, receive them back with HUID / certificate numbers.
# ---------------------------------------------------------------------------
def _bag_convert_materials(bags):
	"""{bag: {item: qty}} from the finished bags' Convert (Out) rows — the frozen
	composition every finished-piece stock move is denominated in."""
	out = {b: {} for b in bags}
	if not bags:
		return out
	for bag, item, qty in frappe.db.sql(
		"""select order_bag, item, sum(qty) from `tabBag Material Ledger`
		   where order_bag in %(bags)s and entry_type = 'Convert' and direction = 'Out'
		   group by order_bag, item""",
		{"bags": bags},
	):
		if flt(qty) > 0:
			out[bag][item] = flt(qty)
	return out


def _stock_move_many(item_qty, source, target):
	"""One Material Transfer moving {item: qty} from source -> target. Returns the
	SE name (None when there is nothing to move)."""
	rows = [{
		"item_code": it, "qty": q,
		"uom": frappe.db.get_value("Item", it, "stock_uom") or "Gram",
		"s_warehouse": source, "t_warehouse": target,
		"allow_zero_valuation_rate": 1,
	} for it, q in item_qty.items() if flt(q) > 0]
	if not rows or not source or not target:
		return None
	se = frappe.get_doc({"doctype": "Stock Entry", "stock_entry_type": "Material Transfer",
	                     "company": _company(), "items": rows})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	return se.name


# ---------------------------------------------------------------------------
# Loss Report (Reports > Stock Reports) — the '<Stage> -LOSS' warehouses rolled
# up: what recoverable loss sits where, and how much PURE gold it holds
# (qty x the item's purity%). Mostly karat golds; anything else shows honestly.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_loss_report():
	whs = frappe.get_all("Warehouse", filters={"custom_is_loss": 1, "is_group": 0},
	                     fields=["name", "warehouse_name"], order_by="warehouse_name")
	if not whs:
		return {"items": [], "warehouses": [], "totals": {}}
	label_of = {w.name: (w.warehouse_name or w.name).replace(" -LOSS", "") for w in whs}

	cells, meta = {}, {}
	for b in frappe.get_all("Bin", filters={"warehouse": ["in", [w.name for w in whs]]},
	                        fields=["item_code", "warehouse", "actual_qty"]):
		qty = flt(b.actual_qty)
		if qty <= 0.0005:
			continue
		cells.setdefault(b.item_code, {})[b.warehouse] = round(cells.get(b.item_code, {}).get(b.warehouse, 0) + qty, 3)
	for it in frappe.get_all("Item", filters={"name": ["in", list(cells) or [""]]},
	                         fields=["name", "item_group", "purity_percentage", "stone_type"]):
		meta[it.name] = it

	wh_stats = {}
	items = []
	for item in sorted(cells, key=lambda i: ((meta.get(i) or {}).get("item_group") or "", i)):
		m = meta.get(item) or {}
		purity = flt(m.get("purity_percentage"))
		total = round(sum(cells[item].values()), 3)
		pure = round(total * purity / 100.0, 3)
		items.append({
			"item": item, "group": m.get("item_group") or "", "purity": purity,
			"is_stone": bool(m.get("stone_type")),
			"total": total, "pure": pure, "cells": cells[item],
		})
		for wh, qty in cells[item].items():
			d = wh_stats.setdefault(wh, {"gross": 0.0, "pure": 0.0})
			d["gross"] += qty
			d["pure"] += qty * purity / 100.0

	warehouses = [{
		"warehouse": w.name, "label": label_of[w.name],
		"gross": round(wh_stats[w.name]["gross"], 3), "pure": round(wh_stats[w.name]["pure"], 3),
	} for w in whs if w.name in wh_stats]
	return {
		"items": items,
		"warehouses": warehouses,
		"totals": {
			"gross": round(sum(x["gross"] for x in warehouses), 3),
			"pure": round(sum(x["pure"] for x in warehouses), 3),
			"warehouses": len(warehouses),
			"materials": len(items),
		},
	}


# ---------------------------------------------------------------------------
# User Roles (Setup > Employee) — who holds which roles, at a glance.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_user_roles():
	"""Every enabled system user with their roles. Jewelima roles (+ the two
	ERPNext roles that gate our pages) come as matrix columns; the rest as chips."""
	frappe.only_for(("System Manager",))
	users = frappe.get_all(
		"User",
		filters={"enabled": 1, "user_type": "System User", "name": ["not in", ["Guest"]]},
		fields=["name", "full_name", "role_profile_name", "last_active"],
		order_by="full_name",
	)
	role_rows = frappe.get_all("Has Role", filters={"parenttype": "User", "parent": ["in", [u.name for u in users] or [""]]},
	                           fields=["parent", "role"])
	by_user = {}
	for r in role_rows:
		by_user.setdefault(r.parent, set()).add(r.role)
	# employee names, where the user account belongs to one
	emp = {e.user_id: e.employee_name for e in frappe.get_all(
		"Employee", filters={"user_id": ["in", [u.name for u in users] or [""]]},
		fields=["user_id", "employee_name"])}

	ours = sorted(frappe.get_all("Role", filters={"role_name": ["like", "Jewelima%"]}, pluck="name"))
	columns = ours + ["System Manager", "Stock Manager"]
	hide = set(columns) | {"All", "Guest", "Desk User"}
	out = []
	for u in users:
		roles = by_user.get(u.name, set())
		out.append({
			"user": u.name, "full_name": u.full_name or u.name,
			"employee": emp.get(u.name, ""), "role_profile": u.role_profile_name or "",
			"last_active": str(u.last_active or ""),
			"has": {c: (c in roles) for c in columns},
			"others": sorted(r for r in roles if r not in hide),
		})
	return {"columns": columns, "users": out}


# ---------------------------------------------------------------------------
# Add User (Setup > Employee) — desk logins are created ONLY from the Employee
# list, following the import_users conventions: username from the name, email =
# username@jd.in (record id only — login is by USERNAME), Jewelima Only
# module profile, Employee.user_id linked. Passwords are NEVER set here.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_employees_without_user():
	"""Active employees who don't have a working desk login yet."""
	frappe.only_for(("System Manager",))
	from jewelima.jewelima.imports.import_users import _username

	rows = frappe.get_all("Employee", filters={"status": "Active"},
	                      fields=["name", "employee_name", "designation", "user_id"], order_by="employee_name")
	out = []
	for e in rows:
		if e.user_id and frappe.db.exists("User", e.user_id):
			continue
		uname = _username(e.employee_name)
		out.append({
			"employee": e.name, "employee_name": e.employee_name, "designation": e.designation or "",
			"username": uname, "email": uname.lower() + "@jd.in",
		})
	return {
		"employees": out,
		"roles": sorted(frappe.get_all("Role", filters={"role_name": ["like", "Jewelima%"]}, pluck="name")),
	}


@frappe.whitelist()
def create_employee_users(payload):
	"""Create desk users for the picked employees (no passwords — set those with
	your own set_passwords step). rows = [{employee, username}], roles = [...]"""
	frappe.only_for(("System Manager",))
	from jewelima.jewelima.imports.import_users import _free_username, _username

	p = frappe.parse_json(payload)
	rows = p.get("rows") or []
	roles = [r for r in (p.get("roles") or []) if frappe.db.exists("Role", r)]
	if not roles:
		roles = ["Jewelima"]  # a role-less user gets demoted to Website User and can't log in to the desk
	if not rows:
		frappe.throw(frappe._("Pick at least one employee."))
	frappe.db.set_single_value("System Settings", "allow_login_using_user_name", 1)

	made, skipped = [], []
	for row in rows:
		emp = frappe.db.get_value("Employee", row.get("employee"), ["name", "employee_name", "user_id"], as_dict=True)
		if not emp:
			frappe.throw(frappe._("Employee {0} not found.").format(row.get("employee") or "?"))
		if emp.user_id and frappe.db.exists("User", emp.user_id):
			skipped.append(emp.employee_name)
			continue
		base = (row.get("username") or "").strip().upper() or _username(emp.employee_name)
		base = _username(base)
		if not base:
			frappe.throw(frappe._("{0}: username came out empty.").format(emp.employee_name))
		email = base.lower() + "@jd.in"
		if frappe.db.exists("User", email):
			user = frappe.get_doc("User", email)
		else:
			full = (emp.employee_name or base).split()
			user = frappe.get_doc({
				"doctype": "User", "email": email,
				"first_name": (full[0].title() if full else base),
				"last_name": " ".join(x.title() for x in full[1:]),
				"user_type": "System User", "send_welcome_email": 0,
			}).insert(ignore_permissions=True)
		user.username = _free_username(base, user.name)
		have = {x.role for x in user.get("roles")}
		for r in roles:
			if r not in have:
				user.append("roles", {"role": r})
		if frappe.db.exists("Module Profile", "Jewelima Only") and user.module_profile != "Jewelima Only":
			user.module_profile = "Jewelima Only"
			mp = frappe.get_doc("Module Profile", "Jewelima Only")
			user.set("block_modules", [{"module": m.module} for m in mp.block_modules])
		user.save(ignore_permissions=True)
		frappe.db.set_value("Employee", emp.name, "user_id", email)
		made.append({"employee": emp.employee_name, "username": user.username, "email": email})
	frappe.db.commit()
	return {"created": made, "skipped": skipped}


# ---------------------------------------------------------------------------
# Add Employee (Setup > Employee) — a lean intake matching the importer's
# conventions: full name into first_name, department/designation created on the
# fly, DOB/DOJ stay optional. Benches picked here land on the rosters.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_employee_form_data():
	frappe.only_for(("System Manager",))
	return {
		"departments": sorted({d.department_name for d in frappe.get_all(
			"Department", filters={"is_group": 0}, fields=["department_name"])
			if d.department_name and d.department_name != "All Departments"}),
		"designations": frappe.get_all("Designation", pluck="name", order_by="name"),
		"genders": frappe.get_all("Gender", pluck="name", order_by="name"),
		"benches": frappe.get_all("Bench", pluck="name", order_by="name"),
	}


@frappe.whitelist()
def create_employee(payload):
	frappe.only_for(("System Manager",))
	from jewelima.jewelima.imports.import_employees import _company, _ensure_department, _ensure_designation

	p = frappe.parse_json(payload)
	full_name = (p.get("full_name") or "").strip()
	if not full_name:
		frappe.throw(frappe._("Enter the employee's name."))
	if not (p.get("gender") or "").strip():
		frappe.throw(frappe._("Pick the gender — Employee requires it."))
	if frappe.db.exists("Employee", {"employee_name": full_name, "status": "Active"}):
		frappe.throw(frappe._("An active employee named {0} already exists.").format(full_name))
	company = _company()
	dept = (p.get("department") or "").strip()
	desig = (p.get("designation") or "").strip()
	emp = frappe.get_doc({
		"doctype": "Employee",
		"first_name": full_name,  # full name — matches the importer's convention
		"gender": (p.get("gender") or "").strip() or None,
		"department": _ensure_department(dept, company) if dept else None,
		"designation": _ensure_designation(desig) if desig else None,
		"company": company,
		"status": "Active",
	}).insert(ignore_permissions=True)

	allotted = []
	for bench in p.get("benches") or []:
		if not frappe.db.exists("Bench", bench):
			continue
		b = frappe.get_doc("Bench", bench)
		if not any(r.employee == emp.name for r in b.employees):
			b.append("employees", {"employee": emp.name})
			b.save(ignore_permissions=True)  # validate refreshes the roster
			allotted.append(bench)
	frappe.db.commit()
	return {"employee": emp.name, "employee_name": emp.employee_name, "benches": allotted}


# ---------------------------------------------------------------------------
# Reset Password (Setup > Employee, SYSTEM MANAGER only) — no real mailboxes on
# the floor accounts, so the admin sets passwords directly and hands them over.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_login_accounts(user=None):
	"""Every desk account with its login handle, last login and live session count.
	`user` narrows it to one account (the Employee form's Login Details section)."""
	frappe.only_for(("System Manager",))
	filters = {"user_type": "System User", "name": ["!=", "Guest"]}
	if user:
		filters["name"] = user
	users = frappe.get_all("User", filters=filters,
		fields=["name", "username", "full_name", "enabled", "last_login", "last_active"],
		order_by="enabled desc, username asc, name asc", limit_page_length=0)

	# tabSessions is a plain table, not a DocType — count the live ones per user
	live = {}
	for r in frappe.db.sql("""SELECT user, COUNT(*) n, MAX(lastupdate) seen
		FROM tabSessions WHERE status='Active' GROUP BY user""", as_dict=True):
		live[r.user] = r

	emps = {}
	for e in frappe.get_all("Employee", filters={"user_id": ["is", "set"]},
		fields=["name", "employee_name", "user_id", "designation", "department"], limit_page_length=0):
		emps[e.user_id] = e

	roles = {}
	for r in frappe.get_all("Has Role", filters={"parenttype": "User"},
		fields=["parent", "role"], limit_page_length=0):
		roles.setdefault(r.parent, []).append(r.role)

	out = []
	for u in users:
		s = live.get(u.name) or {}
		e = emps.get(u.name) or {}
		out.append({
			"user": u.name,
			"username": u.username or "",
			"full_name": e.get("employee_name") or u.full_name or "",
			"employee": e.get("name") or "",
			"designation": e.get("designation") or "",
			"department": e.get("department") or "",
			"enabled": u.enabled,
			"last_login": u.last_login,
			"last_active": u.last_active,
			"sessions": s.get("n") or 0,
			"session_seen": s.get("seen"),
			"roles": sorted(r for r in roles.get(u.name, []) if r),
			"never_logged_in": 0 if u.last_login else 1,
		})
	return out


@frappe.whitelist()
def end_user_sessions(user):
	"""Log a user out of every device. Doesn't touch their password."""
	frappe.only_for(("System Manager",))
	if not user or not frappe.db.exists("User", user) or user == "Guest":
		frappe.throw(frappe._("Pick a real user."))
	from frappe.sessions import clear_sessions

	clear_sessions(user=user, force=True)
	frappe.db.commit()
	return {"user": user}


@frappe.whitelist()
def admin_reset_password(user, new_password):
	frappe.only_for(("System Manager",))
	if not user or not frappe.db.exists("User", user) or user == "Guest":
		frappe.throw(frappe._("Pick a real user."))
	pwd = (new_password or "").strip()
	if len(pwd) < 6:
		frappe.throw(frappe._("Password must be at least 6 characters."))
	from frappe.utils.password import update_password

	update_password(user=user, pwd=pwd)
	# old sessions die with the old password
	from frappe.sessions import clear_sessions

	clear_sessions(user=user, force=True)
	# audit trail on the User record (never the password itself)
	frappe.get_doc("User", user).add_comment(
		"Comment", frappe._("Password reset by {0} via the Reset Password page.").format(frappe.session.user))
	frappe.db.commit()
	return {"user": user, "username": frappe.db.get_value("User", user, "username")}


# ---------------------------------------------------------------------------
# Loss Collection / Write-off (Stock) — Option B: recovered pure gold is minused
# from the loss warehouses per purity (grams = pure ÷ purity%); the dust never
# leaves the house. Residue is only removed by management on the write-off page.
# ---------------------------------------------------------------------------
def _validate_loss_lines(lines):
	"""Common guard: every line must be real loss-warehouse stock with enough qty.
	Returns [(item, warehouse, grams, purity)]."""
	if isinstance(lines, str):
		lines = json.loads(lines or "[]")
	out = []
	for l in lines or []:
		item, wh, grams = l.get("item"), l.get("warehouse"), flt(l.get("grams"))
		if grams <= 0:
			continue
		if not frappe.db.get_value("Warehouse", wh, "custom_is_loss"):
			frappe.throw(frappe._("{0} is not a loss warehouse.").format(wh))
		have = flt(frappe.db.get_value("Bin", {"item_code": item, "warehouse": wh}, "actual_qty"))
		if have + 0.0005 < grams:
			frappe.throw(frappe._("{0} at {1}: only {2} g there (asked {3} g).").format(item, wh, round(have, 3), grams))
		purity = flt(frappe.db.get_value("Item", item, "purity_percentage"))
		out.append((item, wh, grams, purity))
	if not out:
		frappe.throw(frappe._("Pick at least one loss line."))
	return out


@frappe.whitelist()
def collect_loss(payload):
	"""Book a refining recovery: consume loss-warehouse dust (grams per line, as
	allocated on the page) and produce the recovered standard gold — ONE Repack
	Stock Entry. Conservation enforced: the dust's pure content must equal the
	recovered gold's pure content (small rounding tolerance)."""
	p = frappe.parse_json(payload)
	out_item, got, out_wh = p.get("output_item"), flt(p.get("got_grams")), p.get("warehouse")
	if not out_item or not frappe.db.exists("Item", out_item):
		frappe.throw(frappe._("Pick the recovered gold item."))
	if frappe.db.get_value("Item", out_item, "item_group") != "GOLD STANDARD":
		frappe.throw(frappe._("{0} is not a standard gold — recoveries land as Standard gold.").format(out_item))
	if got <= 0:
		frappe.throw(frappe._("Enter the recovered grams."))
	if not out_wh or not frappe.db.exists("Warehouse", out_wh):
		frappe.throw(frappe._("Pick the warehouse the recovered gold goes to."))
	lines = _validate_loss_lines(p.get("lines"))
	out_purity = flt(frappe.db.get_value("Item", out_item, "purity_percentage"))
	need_pure = got * out_purity / 100.0
	give_pure = sum(g * pur / 100.0 for _, _, g, pur in lines)
	if abs(need_pure - give_pure) > 0.01:
		frappe.throw(frappe._("Pure gold doesn't balance: recovering {0} g pure but deducting {1} g pure from loss.").format(
			round(need_pure, 3), round(give_pure, 3)))

	se = frappe.get_doc({
		"doctype": "Stock Entry", "stock_entry_type": "Repack", "company": _company(),
		"items": [{
			"item_code": item, "qty": grams,
			"uom": frappe.db.get_value("Item", item, "stock_uom") or "Gram",
			"s_warehouse": wh, "allow_zero_valuation_rate": 1,
		} for item, wh, grams, _ in lines] + [{
			"item_code": out_item, "qty": got,
			"uom": frappe.db.get_value("Item", out_item, "stock_uom") or "Gram",
			"t_warehouse": out_wh, "allow_zero_valuation_rate": 1, "is_finished_item": 1,
		}],
		"remarks": "Loss collection: {0} g {1} recovered from dust. {2}".format(got, out_item, p.get("remarks") or "").strip(),
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	frappe.db.commit()
	return {"stock_entry": se.name, "pure": round(need_pure, 3), "lines": len(lines)}


@frappe.whitelist()
def writeoff_loss(payload):
	"""MANAGEMENT ONLY: write unrecoverable dust out of the loss warehouses —
	one Material Issue, reason required."""
	frappe.only_for("System Manager")
	p = frappe.parse_json(payload)
	reason = (p.get("reason") or "").strip()
	if not reason:
		frappe.throw(frappe._("A reason is required to write loss off."))
	lines = _validate_loss_lines(p.get("lines"))
	se = frappe.get_doc({
		"doctype": "Stock Entry", "stock_entry_type": "Material Issue", "company": _company(),
		"items": [{
			"item_code": item, "qty": grams,
			"uom": frappe.db.get_value("Item", item, "stock_uom") or "Gram",
			"s_warehouse": wh, "allow_zero_valuation_rate": 1,
		} for item, wh, grams, _ in lines],
		"remarks": "Loss write-off: {0}".format(reason),
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()
	frappe.db.commit()
	pure = round(sum(g * pur / 100.0 for _, _, g, pur in lines), 3)
	return {"stock_entry": se.name, "pure": pure, "lines": len(lines)}


# ---------------------------------------------------------------------------
# Parties (Setup > Party) — "Party" is our word for ERPNext's Customer.
# Every party carries a STRUCTURED name built from four master codes:
#     GROUP-ZONE-STATE[-SPECIAL]      e.g. JOS-TCR-KL-PTY, EDI-CHE-TN
# so who/where is readable at a glance and everything is queryable by part.
# The ONLY exemptions are the internal stock holders (JD Stock / BTQ Stock).
# ---------------------------------------------------------------------------
PARTY_EXEMPT = ("JD Stock", "BTQ Stock")


def _party_code(dt, name):
	if not name:
		return None
	code = frappe.db.get_value(dt, name, "code")
	if not code:
		frappe.throw(frappe._("{0} '{1}' not found.").format(dt, name))
	return code


def _party_name_from(group, zone, state, special=None):
	"""JOS + TCR + KL (+ PTY) -> 'JOS-TCR-KL-PTY'. All parts are master codes."""
	g = _party_code("Party Group", group)
	z = _party_code("Party Zone", zone)
	s = _party_code("Party State", state)
	if not (g and z and s):
		frappe.throw(frappe._("Pick the group, zone and state — they build the party name."))
	parts = [g, z, s]
	sp = _party_code("Party Special", special) if special else None
	if sp:
		parts.append(sp)
	return "-".join(parts)


# voucher types ride on the same Masters page (usage = purchase records)
VOUCHER_MASTER = ("Voucher Type", "title", None)

PARTY_MASTERS = {
	# kind -> (doctype, label field, customer link field)
	"group": ("Party Group", "group_name", "party_group"),
	"zone": ("Party Zone", "zone_name", "party_zone"),
	"state": ("Party State", "state_name", "party_state"),
	"special": ("Party Special", "special_name", "party_special"),
}


@frappe.whitelist()
def get_party_masters():
	"""Setup > Masters: every party master with its code, full name and how many
	customers carry it — one call paints the whole page."""
	custs = frappe.get_all("Customer",
		fields=["party_group", "party_zone", "party_state", "party_special"], limit_page_length=0)
	out = {}
	for kind, (dt, label_field, cust_field) in PARTY_MASTERS.items():
		rows = frappe.get_all(dt, fields=["name", label_field], order_by="name")
		counts = {}
		for c in custs:
			v = c.get(cust_field)
			if v:
				counts[v] = counts.get(v, 0) + 1
		out[kind] = [{"code": r.name, "label": r.get(label_field), "customers": counts.get(r.name, 0)} for r in rows]
	if frappe.db.exists("DocType", "Certification Type"):
		ct = frappe.get_all("Certification Type", fields=["name", "title"], order_by="name")
		ccnt = {}
		for r in frappe.get_all("Certification Center", fields=["certification_type"], limit_page_length=0):
			ccnt[r.certification_type] = ccnt.get(r.certification_type, 0) + 1
		out["cert"] = [{"code": r.name, "label": r.title, "customers": ccnt.get(r.name, 0)} for r in ct]
	if frappe.db.exists("DocType", "Voucher Type"):
		vt = frappe.get_all("Voucher Type", fields=["name", "title"], order_by="name")
		vcnt = {}
		if frappe.db.exists("DocType", "Purchase Record"):
			for r in frappe.get_all("Purchase Record", fields=["voucher_type"], limit_page_length=0):
				vcnt[r.voucher_type] = vcnt.get(r.voucher_type, 0) + 1
		out["voucher"] = [{"code": r.name, "label": r.title, "customers": vcnt.get(r.name, 0)} for r in vt]
	return out


@frappe.whitelist()
def get_master_customers(kind, code):
	"""Every customer under one master value (the drill-down on Setup > Masters)."""
	if kind not in PARTY_MASTERS:
		frappe.throw(frappe._("Unknown master kind."))
	dt, _label, cust_field = PARTY_MASTERS[kind]
	if not frappe.db.exists(dt, code):
		frappe.throw(frappe._("{0} '{1}' not found.").format(dt, code))
	return {"customers": frappe.get_all("Customer", filters={cust_field: code},
		fields=["name", "party_group", "party_zone", "party_state", "party_special",
			"default_salesman", "disabled"], order_by="name", limit_page_length=0)}


@frappe.whitelist()
def add_party_master(kind, code, label):
	"""Add a value to one of the party masters or the voucher types (codes
	validated by each doctype)."""
	if kind == "cert":
		dt, label_field = "Certification Type", "title"
	elif kind == "voucher":
		dt, label_field = "Voucher Type", "title"
	elif kind in PARTY_MASTERS:
		dt, label_field, _cust = PARTY_MASTERS[kind]
	else:
		frappe.throw(frappe._("Unknown master kind."))
	code, label = (code or "").strip().upper(), (label or "").strip()
	if not code or not label:
		frappe.throw(frappe._("Enter both the code and the full name."))
	if frappe.db.exists(dt, code):
		frappe.throw(frappe._("{0} {1} already exists.").format(dt, code))
	d = frappe.get_doc({"doctype": dt, "code": code, label_field: label}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name}


@frappe.whitelist()
def get_party_directory():
	"""Every party with its identity parts + defaults, plus the master lists —
	one call paints the whole Parties page."""
	rows = frappe.get_all("Customer",
		fields=["name", "customer_name", "party_group", "party_zone", "party_state",
			"party_special", "default_salesman", "default_price_chart", "disabled"],
		order_by="name", limit_page_length=0)
	for r in rows:
		r["exempt"] = 1 if r.name in PARTY_EXEMPT else 0
		r["classified"] = 1 if (r.party_group and r.party_zone and r.party_state) else 0
	masters = {
		"groups": frappe.get_all("Party Group", fields=["name", "group_name"], order_by="name"),
		"zones": frappe.get_all("Party Zone", fields=["name", "zone_name"], order_by="name"),
		"states": frappe.get_all("Party State", fields=["name", "state_name"], order_by="name"),
		"specials": frappe.get_all("Party Special", fields=["name", "special_name"], order_by="name"),
		"salesmen": frappe.get_all("Sales Person", filters={"is_group": 0}, pluck="name", order_by="name"),
		"price_charts": frappe.get_all("Price Chart", filters={"status": "Active"}, pluck="name", order_by="name")
			if frappe.db.exists("DocType", "Price Chart") else [],
	}
	return {"parties": rows, "masters": masters,
		"unclassified": len([r for r in rows if not r["classified"] and not r["exempt"]])}


@frappe.whitelist()
def make_party(group, zone, state, special=None, salesman=None, price_chart=None):
	"""Create a NEW party — the name IS the code combo (JOS-TCR-KL[-PTY]).
	One party per exact combo; use a different zone (e.g. Chennai 2) for a
	second store in the same city."""
	nm = _party_name_from(group, zone, state, special)
	if frappe.db.exists("Customer", nm):
		frappe.throw(frappe._("{0} already exists — one party per combo (add a zone like 'Chennai 2' for a second store).").format(nm))
	cg = frappe.db.get_value("Customer Group", {"is_group": 0}, "name")
	terr = frappe.db.get_value("Territory", {"is_group": 0}, "name")
	doc = frappe.get_doc({
		"doctype": "Customer", "customer_name": nm, "customer_group": cg, "territory": terr,
		"party_group": group, "party_zone": zone, "party_state": state, "party_special": special or None,
		"default_salesman": salesman or None, "default_price_chart": price_chart or None,
	}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name}


@frappe.whitelist()
def classify_party(customer, group, zone, state, special=None):
	"""Give an EXISTING party its structured identity: set the four links and
	RENAME the record to the generated code name. Every Link field pointing at
	the customer (orders, bags, sales) follows the rename automatically."""
	if not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("{0} not found.").format(customer))
	if customer in PARTY_EXEMPT:
		frappe.throw(frappe._("{0} is an internal stock holder — it stays outside the naming scheme.").format(customer))
	nm = _party_name_from(group, zone, state, special)
	if nm != customer and frappe.db.exists("Customer", nm):
		frappe.throw(frappe._("{0} already exists — one party per combo.").format(nm))
	frappe.db.set_value("Customer", customer, {
		"party_group": group, "party_zone": zone, "party_state": state,
		"party_special": special or None,
	})
	if nm != customer:
		frappe.rename_doc("Customer", customer, nm, force=True)
		frappe.db.set_value("Customer", nm, "customer_name", nm)
	frappe.db.commit()
	return {"name": nm, "renamed": nm != customer}


@frappe.whitelist()
def update_party_defaults(customer, salesman=None, price_chart=None):
	"""The per-party defaults the sale/order flow prefills from."""
	if not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("{0} not found.").format(customer))
	if salesman and not frappe.db.exists("Sales Person", salesman):
		frappe.throw(frappe._("Salesman {0} not found.").format(salesman))
	if price_chart and not frappe.db.exists("Price Chart", price_chart):
		frappe.throw(frappe._("Price Chart {0} not found.").format(price_chart))
	frappe.db.set_value("Customer", customer, {
		"default_salesman": salesman or None, "default_price_chart": price_chart or None,
	})
	frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def get_party_detail(customer):
	"""Everything the Parties page shows for one party: identity, defaults, and
	the numbers that matter (orders, cards on the floor, stock, sales)."""
	if not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("{0} not found.").format(customer))
	d = frappe.db.get_value("Customer", customer,
		["name", "customer_name", "party_group", "party_zone", "party_state", "party_special",
		 "default_salesman", "default_price_chart", "disabled", "creation"], as_dict=True)
	labels = {
		"group_label": frappe.db.get_value("Party Group", d.party_group, "group_name") if d.party_group else "",
		"zone_label": frappe.db.get_value("Party Zone", d.party_zone, "zone_name") if d.party_zone else "",
		"state_label": frappe.db.get_value("Party State", d.party_state, "state_name") if d.party_state else "",
		"special_label": frappe.db.get_value("Party Special", d.party_special, "special_name") if d.party_special else "",
	}
	stats = {
		"job_orders": frappe.db.count("Job Order", {"customer": customer}),
		"last_order": str((frappe.get_all("Job Order", filters={"customer": customer},
			order_by="creation desc", limit=1, pluck="creation") or [""])[0])[:10],
		"bags_in_production": frappe.db.count("Order Bag", {"customer": customer, "stock_status": "In Production"}),
		"products_in_stock": frappe.db.count("Order Bag", {"customer": customer, "is_finished": 1, "stock_status": "In Stock"}),
		"sold": frappe.db.count("Order Bag", {"customer": customer, "stock_status": "Sold"}),
	}
	recent = frappe.get_all("Job Order", filters={"customer": customer},
		fields=["name", "creation", "salesman", "order_type"], order_by="creation desc", limit=8)
	for r in recent:
		r["creation"] = str(r.creation)[:10]
	return {**d, **labels, "stats": stats, "recent_orders": recent, "exempt": 1 if customer in PARTY_EXEMPT else 0}


# ---------------------------------------------------------------------------
# Price Charts (Setup) — the dedicated editor + the customer-facing PDF letter.
# The doctype stays the storage (the Sell engine prices against it; saving an
# Active chart auto-supersedes the previous one of the same name); these APIs
# and the price-charts page replace the raw ERPNext form as the face.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_price_chart_list():
	"""Charts grouped by name: the Active one + its superseded history."""
	rows = frappe.get_all("Price Chart",
		fields=["name", "chart_name", "chart_date", "status", "modified"],
		order_by="chart_name asc, chart_date desc, creation desc")
	groups = {}
	for r in rows:
		g = groups.setdefault(r.chart_name or r.name, {"chart_name": r.chart_name, "active": None, "history": []})
		r["chart_date"] = str(r.chart_date or "")
		if r.status == "Active" and not g["active"]:
			g["active"] = r
		else:
			g["history"].append(r)
	return {"groups": list(groups.values())}


@frappe.whitelist()
def get_price_chart(name):
	"""One chart, every block — feeds the editor and the PDF preview."""
	d = frappe.get_doc("Price Chart", name)
	return {
		"name": d.name, "chart_name": d.chart_name, "chart_date": str(d.chart_date or ""),
		"status": d.status,
		"diamond_rates": [{"sieve_label": r.sieve_label, "from_ct": r.from_ct, "to_ct": r.to_ct,
			"quality": r.quality, "rate": r.rate} for r in d.diamond_rates],
		"solitaire_min_ct": flt(d.get("solitaire_min_ct")) or 0.07,
		"solitaire_rates": [{"from_ct": r.from_ct, "to_ct": r.to_ct, "quality": r.quality, "rate": r.rate}
			for r in (d.get("solitaire_rates") or [])],
		"certification_charges": [{"certification": r.certification, "basis": r.basis or "Per Piece",
			"rate": r.rate, "min_amount": r.min_amount} for r in (d.get("certification_charges") or [])],
		"precious_stone_rates": [{"stone": r.stone, "rate": r.rate} for r in (d.get("precious_stone_rates") or [])],
		"cs_rates": [{"from_ct": r.from_ct, "to_ct": r.to_ct, "basis": r.basis or "Per Ct", "rate": r.rate} for r in (d.get("cs_rates") or [])],
		"cz_rates": [{"from_ct": r.from_ct, "to_ct": r.to_ct, "basis": r.basis or "Per Ct", "rate": r.rate} for r in (d.get("cz_rates") or [])],
		"cvd_rates": [{"from_ct": r.from_ct, "to_ct": r.to_ct, "basis": r.basis or "Per Ct", "rate": r.rate} for r in (d.get("cvd_rates") or [])],
		"making_rate": flt(d.making_rate), "making_min_grams": flt(d.making_min_grams),
		"making_rules": [{"design_type": r.design_type or "", "basis": r.basis or "Per Gram",
			"rate": r.rate, "min_per_piece": r.min_per_piece} for r in (d.get("making_rules") or [])],
		"payment_terms": d.payment_terms or "", "terms": d.terms or "",
		"signatory": d.signatory or "", "signatory_phone": d.signatory_phone or "",
	}


@frappe.whitelist()
def save_price_chart(payload):
	"""Save = a NEW Active version (the controller supersedes the previous Active
	chart of the same name — history is never edited)."""
	p = frappe.parse_json(payload) if isinstance(payload, str) else payload
	if not (p.get("chart_name") or "").strip():
		frappe.throw(frappe._("Give the chart a name — usually the party's name."))
	doc = frappe.new_doc("Price Chart")
	doc.chart_name = p["chart_name"].strip().upper()
	doc.chart_date = p.get("chart_date") or frappe.utils.today()
	doc.status = "Active"
	for r in p.get("diamond_rates") or []:
		if r.get("rate"):
			doc.append("diamond_rates", {"sieve_label": r.get("sieve_label"), "from_ct": flt(r.get("from_ct")),
				"to_ct": flt(r.get("to_ct")), "quality": (r.get("quality") or "").strip(), "rate": flt(r.get("rate"))})
	doc.solitaire_min_ct = flt(p.get("solitaire_min_ct")) or 0.07
	for r in p.get("solitaire_rates") or []:
		if r.get("rate"):
			doc.append("solitaire_rates", {"from_ct": flt(r.get("from_ct")), "to_ct": flt(r.get("to_ct")),
				"quality": (r.get("quality") or "").strip(), "rate": flt(r.get("rate"))})
	cert_rows_in = [(r.get("certification") or "").strip().upper()
		for r in (p.get("certification_charges") or []) if (r.get("certification") or "").strip()]
	# ALL LABS is the GROUP price for every lab certification (hallmarking stays
	# its own row) — group and individual lab rows never coexist
	labs_in = [c for c in cert_rows_in if c not in ("HALL", "HALLMARKING", "ALL LABS")]
	if "ALL LABS" in cert_rows_in and labs_in:
		frappe.throw(frappe._("Use EITHER the ALL LABS group price OR individual lab rows — not both ({0}).").format(", ".join(labs_in)))
	for r in p.get("certification_charges") or []:
		if (r.get("certification") or "").strip():
			cert_code = r.get("certification").strip().upper()
			basis = r.get("basis") or "Per Piece"
			if cert_code in ("HALL", "HALLMARKING") and basis == "Per Ct":
				frappe.throw(frappe._("HALLMARKING is always per piece (per HUID) — Per Ct doesn't apply."))
			doc.append("certification_charges", {"certification": cert_code, "basis": basis,
				"rate": flt(r.get("rate")), "min_amount": flt(r.get("min_amount"))})
	for r in p.get("precious_stone_rates") or []:
		if (r.get("stone") or "").strip():
			doc.append("precious_stone_rates", {"stone": r.get("stone").strip(), "rate": flt(r.get("rate"))})
	for field in ("cs_rates", "cz_rates", "cvd_rates"):
		for r in p.get(field) or []:
			if flt(r.get("rate")):
				doc.append(field, {"from_ct": flt(r.get("from_ct")), "to_ct": flt(r.get("to_ct")),
					"basis": r.get("basis") or "Per Ct", "rate": flt(r.get("rate"))})
	doc.making_rate = flt(p.get("making_rate"))
	doc.making_min_grams = flt(p.get("making_min_grams"))
	for r in p.get("making_rules") or []:
		if flt(r.get("rate")):
			doc.append("making_rules", {"design_type": r.get("design_type") or None,
				"basis": r.get("basis") or "Per Gram", "rate": flt(r.get("rate")),
				"min_per_piece": flt(r.get("min_per_piece"))})
	doc.payment_terms = p.get("payment_terms") or ""
	doc.terms = p.get("terms") or ""
	doc.signatory = p.get("signatory") or ""
	doc.signatory_phone = p.get("signatory_phone") or ""
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "chart_name": doc.chart_name}


def _price_chart_letter_html(d):
	"""The customer-facing rate letter (A4). d = get_price_chart payload."""
	def money(v):
		"""Indian grouping: 185000 -> 1,85,000."""
		n = flt(v)
		neg = n < 0
		whole = int(abs(n))
		frac = abs(n) - whole
		sw = str(whole)
		if len(sw) > 3:
			head, tail = sw[:-3], sw[-3:]
			parts = []
			while len(head) > 2:
				parts.insert(0, head[-2:])
				head = head[:-2]
			if head:
				parts.insert(0, head)
			sw = ",".join(parts + [tail])
		out = sw + ("{:.2f}".format(frac)[1:] if frac >= 0.005 else "")
		return ("-" if neg else "") + out
	def bracket(r):
		if flt(r["to_ct"]):
			return "{0} – {1} ct".format(r["from_ct"], r["to_ct"])
		return "{0} ct & above".format(r["from_ct"]) if flt(r["from_ct"]) else "any size"
	dmd = "".join("<tr><td>{0}</td><td>{1}</td><td><b>{2}</b></td><td class='r'>₹ {3}</td></tr>".format(
		frappe.utils.escape_html(r["sieve_label"] or ""), bracket(r),
		frappe.utils.escape_html(r["quality"] or "All"), money(r["rate"])) for r in d["diamond_rates"])
	sol = "".join("<tr><td>{0}</td><td><b>{1}</b></td><td class='r'>₹ {2}</td></tr>".format(
		bracket(r), frappe.utils.escape_html(r["quality"] or "All"), money(r["rate"])) for r in d.get("solitaire_rates", []))
	certs = "".join("<tr><td>{0}</td><td class='r'>{1}</td></tr>".format(
		frappe.utils.escape_html(r["certification"]),
		("₹ {0} / ct (min ₹ {1})".format(money(r["rate"]), money(r.get("min_amount"))) if flt(r.get("min_amount"))
			else "₹ {0} / ct".format(money(r["rate"]))) if r.get("basis") == "Per Ct"
		else ("₹ " + money(r["rate"]) + " / piece" if flt(r["rate"]) else "Included"))
		for r in d.get("certification_charges", []))
	psr = "".join("<tr><td>{0}</td><td class='r'>₹ {1} / ct</td></tr>".format(
		frappe.utils.escape_html(r["stone"]), money(r["rate"])) for r in d.get("precious_stone_rates", []))
	mkr = "".join("<tr><td>{0}</td><td>{1}</td><td class='r'>₹ {2}{3}</td></tr>".format(
		frappe.utils.escape_html(r["design_type"] or "All designs (default)"), r["basis"],
		money(r["rate"]) + ("/g" if r["basis"] == "Per Gram" else "/pc"),
		" · min ₹ " + money(r["min_per_piece"]) if flt(r.get("min_per_piece")) else "")
		for r in d.get("making_rules", []))
	def brk(rows):
		return "".join("<tr><td>{0}</td><td class='r'>₹ {1} / {2}</td></tr>".format(
			("any weight" if not flt(r["from_ct"]) and not flt(r["to_ct"])
				else ("{0} – {1} ct".format(r["from_ct"], r["to_ct"]) if flt(r["to_ct"])
				else "{0} ct & above".format(r["from_ct"]))), money(r["rate"]),
			"pc" if r.get("basis") == "Per Piece" else "ct") for r in rows)
	cs = brk(d.get("cs_rates", []))
	cz = brk(d.get("cz_rates", []))
	cvd = brk(d.get("cvd_rates", []))
	flats = []
	flat_rows = "".join("<tr><td>{0}</td><td class='r'>{1}</td></tr>".format(k, v) for k, v in flats)
	sec = lambda title, table_head, body: (
		"<div class='sec'><div class='st'>{0}</div><table>{1}<tbody>{2}</tbody></table></div>".format(
			title, table_head, body) if body else "")
	esc = frappe.utils.escape_html
	import base64 as _b64
	logo_html = ""
	try:
		lp = frappe.get_app_path("jewelima", "public", "images", "jewelima-letterhead.png")
		logo_html = "<img src='data:image/png;base64,{0}'>".format(_b64.b64encode(open(lp, "rb").read()).decode())
	except Exception:
		logo_html = "<div style='font-size:21px;font-weight:800;color:#1f4e5f;'>JEWELIMA</div>"
	return """<!doctype html><html><head><meta charset='utf-8'><style>
		@page {{ size: A4; margin: 18mm 16mm; }}
		body {{ font-family: Helvetica, Arial, sans-serif; color: #1a1a1a; font-size: 12.5px; }}
		.head {{ border-bottom: 3px solid #1f4e5f; padding-bottom: 10px; margin-bottom: 18px; }}
		.head img {{ max-height: 64px; max-width: 320px; }}
		.foot {{ margin-top: 30px; text-align: center; }}
		.foot .rule {{ border-top: 1px solid #1f4e5f; margin-bottom: 8px; }}
		.foot .tag {{ font-size: 12px; letter-spacing: .35em; color: #1f4e5f; text-transform: lowercase; }}
		.doc {{ font-size: 13px; color: #666; margin-top: 2px; }}
		.meta {{ margin: 10px 0 4px; }}
		.meta b {{ font-size: 16px; }}
		.meta span {{ float: right; color: #666; }}
		.qnote {{ color: #444; font-size: 12px; }}
		.sec {{ margin: 14px 0; page-break-inside: avoid; }}
		.st {{ font-size: 11px; font-weight: 700; letter-spacing: .08em; text-transform: uppercase;
			color: #1f4e5f; border-bottom: 1px solid #1f4e5f; padding-bottom: 3px; margin-bottom: 6px; }}
		table {{ width: 100%; border-collapse: collapse; }}
		th {{ text-align: left; font-size: 10px; text-transform: uppercase; letter-spacing: .05em;
			color: #888; padding: 4px 8px; border-bottom: 1px solid #ddd; }}
		td {{ padding: 5px 8px; border-bottom: 1px solid #eee; }}
		td.r, th.r {{ text-align: right; white-space: nowrap; }}
		.terms {{ margin-top: 16px; font-size: 11.5px; color: #444; white-space: pre-wrap; }}
		.sign {{ margin-top: 34px; display: flex; justify-content: space-between; align-items: flex-end; }}
		.sign .who {{ font-weight: 700; }}
		.sign .line {{ border-top: 1px solid #999; padding-top: 4px; width: 220px; text-align: center; color: #666; font-size: 11px; }}
	</style></head><body>
		<div class='head'>{logo}<div class='doc'>Rate Chart</div></div>
		<div class='meta'><b>{chart_name}</b><span>{chart_date}</span></div>
		{qnote}
		{dmd_sec}{sol_sec}{ps_sec}{cs_sec}{cz_sec}{cvd_sec}{mk_sec}{cert_sec}
		{payment}{terms}
		<div class='sign'>
			<div><div class='who'>{signatory}</div><div>{signatory_phone}</div></div>
			<div class='line'>Authorised Signatory</div>
		</div>
		<div class='foot'><div class='rule'></div><div class='tag'>crafting &mdash; for &mdash; you</div></div>
	</body></html>""".format(
		logo=logo_html,
		chart_name=esc(d["chart_name"]), chart_date=esc(d["chart_date"]),
		qnote="",
		dmd_sec=sec("Diamond Rates", "<thead><tr><th>Sieves</th><th>Size</th><th>Quality</th><th class='r'>Rate / ct</th></tr></thead>", dmd),
		sol_sec=sec("Solitaire Rates (per-stone above {0} ct)".format(d.get("solitaire_min_ct", 0.07)),
			"<thead><tr><th>Per-stone size</th><th>Quality</th><th class='r'>Rate / ct</th></tr></thead>", sol),
		cert_sec=sec("Certification Charges", "", certs),
		ps_sec=sec("Precious Stone Rates", "", psr),
		mk_sec=sec("Making Charges", "<thead><tr><th>Design</th><th>Basis</th><th class='r'>Rate</th></tr></thead>", mkr),
		cs_sec=sec("Colour Stone Rates", "", cs),
		cz_sec=sec("CZ Rates", "", cz),
		cvd_sec=sec("CVD Rates", "", cvd),
		payment="<div class='sec'><div class='st'>Payment Terms</div><div class='terms'>{0}</div></div>".format(esc(d["payment_terms"])) if d["payment_terms"] else "",
		terms="<div class='terms'>{0}</div>".format(esc(d["terms"])) if d["terms"] else "",
		signatory=esc(d["signatory"]), signatory_phone=esc(d["signatory_phone"]))


@frappe.whitelist()
def price_chart_letter(name):
	"""The rate letter as HTML — the page prints it in place (hidden iframe)."""
	return {"html": _price_chart_letter_html(get_price_chart(name))}


@frappe.whitelist()
def export_price_chart_pdf(name):
	"""The rate-chart letter as a PDF — clean enough to send straight to the party."""
	from frappe.utils.pdf import get_pdf
	d = get_price_chart(name)
	frappe.local.response.filename = "RateChart-{0}-{1}.pdf".format(
		(d["chart_name"] or "chart").replace(" ", "-"), d["chart_date"])
	frappe.local.response.filecontent = get_pdf(_price_chart_letter_html(d))
	frappe.local.response.type = "download"


# ---------------------------------------------------------------------------
# Design Bank CARD BUILDER — the Photoshop replacement. Composes the standard
# card PNG (photo + design no + GW/DW + stones + extras) with Pillow; the page
# previews live (base64, nothing stored) and Save renders the real card into
# the record's image, keeping the raw photo for future re-renders.
# ---------------------------------------------------------------------------
def _card_font(size, bold=False):
	"""Card typography: Cantarell (thin, elegant), SHIPPED IN THE APP so every
	deployment renders identically; DejaVu is only the last-resort fallback."""
	from PIL import ImageFont
	# bold -> Bold, "regular" -> Regular (printable weight for the stone/note
	# lines), everything else -> Light
	name = {"bold": "Cantarell-Bold.otf", "regular": "Cantarell-Regular.otf"}.get(
		"bold" if bold is True else bold or "", "Cantarell-Light.otf")
	shipped = frappe.get_app_path("jewelima", "public", "fonts", name)
	for cand in (shipped,
			"/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold
			else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"):
		try:
			return ImageFont.truetype(cand, size)
		except Exception:
			pass
	return ImageFont.load_default()


def _card_compose(p):
	"""payload -> PIL image of the standard design card (portrait 900x1200)."""
	from PIL import Image, ImageDraw
	# BLACK AND WHITE by design — these cards go to a mono printer. Divisions are
	# plain rules; header = design no left / design type right; footer = JEWELIMA.
	W, H = 900, 1200
	img = Image.new("RGB", (W, H), "#ffffff")
	d = ImageDraw.Draw(img)
	d.text((30, 20), (p.get("design_no") or "—").upper(), font=_card_font(40, True), fill="#000000")
	if p.get("design_type"):
		d.text((W - 30, 30), p["design_type"].upper(), font=_card_font(26, True), fill="#000000", anchor="ra")
	d.line([30, 80, W - 30, 80], fill="#000000", width=3)
	# photo box
	top, bottom = 100, 760
	ph = _cad_image_any(p.get("photo") or "")
	if ph:
		ph = ph.convert("RGB")
		ph.thumbnail((W - 60, bottom - top), Image.LANCZOS)
		img.paste(ph, ((W - ph.width) // 2, top + (bottom - top - ph.height) // 2))
	else:
		d.rectangle([30, top, W - 30, bottom], outline="#cccccc", width=2)
		d.text((W // 2, (top + bottom) // 2), "PHOTO", font=_card_font(36), fill="#bbbbbb", anchor="mm")
	y = bottom + 30
	d.line([30, y - 12, W - 30, y - 12], fill="#000000", width=3)
	big = _card_font(36, True)
	if flt(p.get("gross_weight")):
		d.text((30, y), "GW  {0} gm".format(p["gross_weight"]), font=big, fill="#111111")
	if flt(p.get("diamond_weight")):
		d.text((W - 30, y), "DW  {0} ct".format(p["diamond_weight"]), font=big, fill="#111111", anchor="ra")
	y += 58
	# stone/note lines print too faint in Light — Regular, sized to actually use
	# the info band the card reserves for them
	f_line = _card_font(30, "regular")
	for st in (p.get("stones") or []):
		if not (st.get("stone") or st.get("sieve")):
			continue
		bits = [x for x in (st.get("stone"), st.get("sieve"),
			("{0} pc".format(st["pcs"]) if cint(st.get("pcs")) else ""),
			("{0} ct".format(st["ct"]) if flt(st.get("ct")) else "")) if x]
		d.text((30, y), "  •  ".join(str(b) for b in bits), font=f_line, fill="#111111")
		y += 42
	if p.get("note"):
		d.text((30, y), str(p["note"]), font=f_line, fill="#111111")
		y += 42
	for ln in (p.get("extra_lines") or "").split("\n"):
		if ln.strip():
			d.text((30, y), ln.strip(), font=f_line, fill="#111111")
			y += 40
	d.line([30, H - 52, W - 30, H - 52], fill="#000000", width=2)
	d.text((W // 2, H - 40), "JEWELIMA", font=_card_font(22, True), fill="#000000", anchor="ma")
	return img


@frappe.whitelist()
def design_card_autocrop(name):
	"""Pull the PRODUCT photo out of a legacy scanned card: score every pixel by
	colour saturation / darkness, drop the printed top+bottom text bands, take
	the dense bounding box with padding. Heuristic — the live preview shows the
	result and a manual upload always wins. Returns base64 (nothing stored)."""
	import base64
	from io import BytesIO
	d = frappe.get_doc("Design Bank", name)
	src = _cad_image_any(d.photo or d.image)
	if not src:
		frappe.throw(frappe._("No image on {0}.").format(d.design_no or name))
	img = src.convert("RGB")
	# some scans are ENORMOUS — cap before the pixel walk or the batch OOMs
	if max(img.size) > 2400:
		img.thumbnail((2400, 2400))
	w, h = img.size
	small = img.resize((max(1, w // 4), max(1, h // 4)))
	sw, sh = small.size
	px = small.load()
	# per-row/col density of "photo-ish" pixels (saturated OR clearly dark)
	rows = [0] * sh
	cols = [0] * sw
	for y in range(sh):
		for x in range(sw):
			r, g, b = px[x, y]
			mx, mn = max(r, g, b), min(r, g, b)
			sat = mx - mn
			lum = (r + g + b) // 3
			if sat > 26 or lum < 120:
				rows[y] += 1
				cols[x] += 1
	# ignore the printed bands: top 12% (design no) and bottom 20% (GW/DW text)
	y0_lim, y1_lim = int(sh * 0.12), int(sh * 0.80)
	def span(vals, lo, hi, need):
		idx = [i for i in range(lo, hi) if vals[i] >= need]
		return (idx[0], idx[-1]) if idx else (lo, hi - 1)
	ry0, ry1 = span(rows, y0_lim, y1_lim, max(3, sw // 25))
	cx0, cx1 = span(cols, 0, sw, max(3, sh // 25))
	pad = max(2, sw // 40)
	box = (max(0, (cx0 - pad) * 4), max(0, (ry0 - pad) * 4),
		min(w, (cx1 + pad) * 4), min(h, (ry1 + pad) * 4))
	if box[2] - box[0] < w // 6 or box[3] - box[1] < h // 8:
		frappe.throw(frappe._("Couldn't find a clear photo region on this card — upload the photo manually."))
	crop = img.crop(box)
	buf = BytesIO()
	crop.save(buf, "PNG")
	return {"image": "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()}


def _db_img_name(design_no, kind, ext="png"):
	"""The three slots carry the DESIGN CODE as the filename: '<code>.photo.png',
	'<code>.info.png', '<code>.customer.png' (raw keeps its imported name)."""
	safe = re.sub(r"[^A-Za-z0-9 ._-]", "-", (design_no or "design").strip())
	return "{0}.{1}.{2}".format(safe, kind, ext)


@frappe.whitelist()
def design_card_preview(payload):
	"""Live preview: compose and hand back base64 — nothing touches the record."""
	import base64
	from io import BytesIO
	p = frappe.parse_json(payload) if isinstance(payload, str) else payload
	buf = BytesIO()
	_card_compose(p).save(buf, "PNG")
	return {"image": "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()}


@frappe.whitelist()
def design_card_remove_raw(name):
	"""Drop the RAW photo off a card (e.g. while approving) — the rendered card
	image stays; only the making-of source goes."""
	d = frappe.get_doc("Design Bank", name)
	if d.photo:
		for f in frappe.get_all("File", filters={"attached_to_doctype": "Design Bank",
				"attached_to_name": name, "file_url": d.photo}, pluck="name"):
			frappe.delete_doc("File", f, force=True, ignore_permissions=True)
		d.photo = ""
		d.save(ignore_permissions=True)
		frappe.db.commit()
	return {"ok": 1}


@frappe.whitelist()
def check_design_code(code):
	"""The new-code guard: exact match against EVERYTHING (live + retired)."""
	code = (code or "").strip()
	hit = frappe.db.get_value("Design Bank", {"design_no": code}, ["name", "status"], as_dict=True)
	return {"taken": bool(hit), "status": hit.status if hit else "", "record": hit.name if hit else ""}


@frappe.whitelist()
def next_design_code(prefix):
	"""Next free number in an in-house series: 'JS' -> 'JS 0001' / first gap-free top."""
	prefix = (prefix or "").strip().upper()
	if not prefix:
		frappe.throw(frappe._("Give the series prefix (JS, JN, ...)."))
	top = 0
	for dn in frappe.get_all("Design Bank", filters={"design_no": ["like", prefix + " %"]}, pluck="design_no"):
		m = re.match(re.escape(prefix) + r"\s+(\d+)$", (dn or "").strip().upper())
		if m:
			top = max(top, int(m.group(1)))
	return {"code": "{0} {1:04d}".format(prefix, top + 1)}


@frappe.whitelist()
def get_design_card(name):
	d = frappe.get_doc("Design Bank", name)
	return {"name": d.name, "design_no": d.design_no, "status": d.status,
		"design_type": d.design_type or "", "gross_weight": flt(d.gross_weight),
		"diamond_weight": flt(d.diamond_weight), "note": d.note or "",
		"extra_lines": d.extra_lines or "", "photo": d.photo or "", "image": d.image or "",
		"photoupdate": d.photoupdate, "duplicate_review": d.duplicate_review,
		"stones": [{"stone": r.stone, "sieve": r.sieve, "pcs": r.pcs, "ct": r.ct} for r in d.stones]}


@frappe.whitelist()
def save_design_card(payload):
	"""Create or update a card. New records pass the code guard; Save re-renders
	the card PNG into the record's image (deterministic CARD-<name>.png, prior
	render replaced), stores the raw photo, and clears photoupdate."""
	p = frappe.parse_json(payload) if isinstance(payload, str) else payload
	code = (p.get("design_no") or "").strip()
	if not code:
		frappe.throw(frappe._("Give the design number."))
	if p.get("name"):
		d = frappe.get_doc("Design Bank", p["name"])
		other = frappe.db.get_value("Design Bank", {"design_no": code, "name": ["!=", d.name]}, "name")
		if other:
			frappe.throw(frappe._("{0} is already used by another card.").format(code))
	else:
		if frappe.db.exists("Design Bank", {"design_no": code}):
			hit = check_design_code(code)
			frappe.throw(frappe._("{0} is already taken ({1}) — codes are never reused.").format(code, hit["status"]))
		d = frappe.new_doc("Design Bank")
		d.status = "Pending"
	d.design_no = code
	d.design_type = p.get("design_type") or None
	d.gross_weight = flt(p.get("gross_weight"))
	d.diamond_weight = flt(p.get("diamond_weight"))
	d.note = p.get("note") or ""
	d.extra_lines = p.get("extra_lines") or ""
	d.set("stones", [{"stone": r.get("stone"), "sieve": r.get("sieve"),
		"pcs": cint(r.get("pcs")), "ct": flt(r.get("ct"))} for r in (p.get("stones") or [])
		if r.get("stone") or r.get("sieve")])
	if d.is_new():
		d.insert(ignore_permissions=True)
	d.photo = _cad_store_image_generic(p.get("photo"), "Design Bank", d.name,
		fname=_db_img_name(code, "photo")) or d.photo
	# render the card; deterministic name so the old render dies with the new save
	from io import BytesIO
	rp = dict(p)
	rp["photo"] = d.photo
	buf = BytesIO()
	_card_compose(rp).save(buf, "PNG")
	info_name = _db_img_name(code, "info")
	for old in frappe.get_all("File", filters={"attached_to_doctype": "Design Bank",
			"attached_to_name": d.name,
			"file_name": ["in", [info_name, "CARD-{0}.png".format(d.name)]]}, pluck="name"):
		frappe.delete_doc("File", old, force=True, ignore_permissions=True)
	fdoc = frappe.get_doc({"doctype": "File", "file_name": info_name, "content": buf.getvalue(),
		"is_private": 0, "attached_to_doctype": "Design Bank", "attached_to_name": d.name}).insert(ignore_permissions=True)
	d.image = fdoc.file_url
	d.photoupdate = 0
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name, "design_no": d.design_no, "image": d.image}


def _cad_store_image_generic(ref, doctype, name, fname=None):
	"""data-URL -> stored File on any doc; an existing /files/ url passes through."""
	if not ref:
		return ""
	if not ref.startswith("data:"):
		return ref
	import base64
	head, b64 = ref.split(",", 1)
	ext = "jpg" if ("jpeg" in head or "jpg" in head) else "png"
	if fname:
		for old_f in frappe.get_all("File", filters={"attached_to_doctype": doctype,
				"attached_to_name": name, "file_name": ["like", fname.rsplit(".", 1)[0] + "%"]}, pluck="name"):
			frappe.delete_doc("File", old_f, force=True, ignore_permissions=True)
	f = frappe.get_doc({"doctype": "File",
		"file_name": fname or "photo-{0}.{1}".format(frappe.generate_hash(length=8), ext),
		"content": base64.b64decode(b64), "is_private": 0,
		"attached_to_doctype": doctype, "attached_to_name": name}).insert(ignore_permissions=True)
	return f.file_url


# ---------------------------------------------------------------------------
# Prepare for Sale — the two-step sale: build the priced list (per-line manual
# overrides recorded: chart price vs final price vs who), export the
# confirmation excel for the party, and only THEN Sell (stock moves, bags Sold).
# ---------------------------------------------------------------------------
PREP_VALUE_FIELDS = ("gold_value", "diamond_value", "stone_value", "labour_value", "charges_value")


def _prep_totals(doc):
	doc.grand_total = round(sum(flt(r.piece_total) for r in doc.items), 2)


@frappe.whitelist()
def get_sale_preparations():
	return {"preps": frappe.get_all("Sale Preparation",
		fields=["name", "customer", "price_chart", "status", "grand_total", "modified", "sale"],
		order_by="creation desc", limit=50)}


@frappe.whitelist()
def get_sale_preparation(name):
	d = frappe.get_doc("Sale Preparation", name)
	return {
		"name": d.name, "customer": d.customer, "price_chart": d.price_chart,
		"gold_rate": flt(d.gold_rate), "status": d.status, "sale": d.sale,
		"remarks": d.remarks or "", "grand_total": flt(d.grand_total),
		"items": [{"row": r.name, "order_bag": r.order_bag, "design": r.design,
			"design_type": r.design_type, "nett": flt(r.nett), "dmd_ct": flt(r.dmd_ct),
			"solitaire_ct": flt(r.solitaire_ct), "ostone_ct": flt(r.ostone_ct),
			**{f: flt(r.get(f)) for f in PREP_VALUE_FIELDS},
			**{"chart_" + f.replace("_value", ""): flt(r.get("chart_" + f.replace("_value", ""))) for f in PREP_VALUE_FIELDS},
			"piece_total": flt(r.piece_total), "overridden": r.overridden,
			"override_remark": r.override_remark or "", "changed_by": r.changed_by or ""} for r in d.items],
	}


@frappe.whitelist()
def create_sale_preparation(customer, price_chart, gold_rate=0):
	if not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("Pick the party."))
	if not frappe.db.exists("Price Chart", price_chart):
		frappe.throw(frappe._("Pick the price chart."))
	d = frappe.get_doc({"doctype": "Sale Preparation", "customer": customer,
		"price_chart": price_chart, "gold_rate": flt(gold_rate), "status": "Draft"})
	d.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name}


@frappe.whitelist()
def prep_add_piece(name, barcode):
	"""Scan a piece into the preparation — priced by the chart engine (all its
	guards apply: quality/solitaire/certification not on chart = scan denied)."""
	d = frappe.get_doc("Sale Preparation", name)
	if d.status not in ("Draft", "Sent"):
		frappe.throw(frappe._("{0} is {1} — no more edits.").format(name, d.status))
	nm = (barcode or "").strip()
	if any(r.order_bag == nm for r in d.items):
		frappe.throw(frappe._("{0} is already on this list.").format(nm))
	pc = get_sale_piece(nm, d.price_chart, d.gold_rate)
	vals = {f: flt(pc[f]) for f in PREP_VALUE_FIELDS}
	d.append("items", {
		"order_bag": nm, "design": pc["design"], "design_type": pc["design_type"],
		"nett": pc["nett"], "dmd_ct": pc["dmd_ct"], "solitaire_ct": pc.get("solitaire_ct", 0),
		"ostone_ct": pc["ostone_ct"],
		**{"chart_" + f.replace("_value", ""): v for f, v in vals.items()},
		**vals, "piece_total": round(sum(vals.values()), 2),
	})
	_prep_totals(d)
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return get_sale_preparation(name)


@frappe.whitelist()
def prep_set_line(name, row, field, value, remark=None):
	"""A manual price change on one line — ALWAYS recorded: the chart price stays
	in the chart_* column, the final in the value column, plus who changed it."""
	if field not in PREP_VALUE_FIELDS:
		frappe.throw(frappe._("Only the value columns can be changed."))
	d = frappe.get_doc("Sale Preparation", name)
	if d.status not in ("Draft", "Sent"):
		frappe.throw(frappe._("{0} is {1} — no more edits.").format(name, d.status))
	r = next((x for x in d.items if x.name == row), None)
	if not r:
		frappe.throw(frappe._("Line not found."))
	r.set(field, flt(value))
	r.piece_total = round(sum(flt(r.get(f)) for f in PREP_VALUE_FIELDS), 2)
	changed = any(abs(flt(r.get(f)) - flt(r.get("chart_" + f.replace("_value", "")))) > 0.005 for f in PREP_VALUE_FIELDS)
	r.overridden = 1 if changed else 0
	r.changed_by = frappe.session.user if changed else ""
	if remark is not None:
		r.override_remark = remark
	_prep_totals(d)
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return get_sale_preparation(name)


@frappe.whitelist()
def prep_remove_line(name, row):
	d = frappe.get_doc("Sale Preparation", name)
	if d.status not in ("Draft", "Sent"):
		frappe.throw(frappe._("{0} is {1} — no more edits.").format(name, d.status))
	d.set("items", [x for x in d.items if x.name != row])
	_prep_totals(d)
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return get_sale_preparation(name)


@frappe.whitelist()
def prep_set_status(name, status):
	"""Draft <-> Sent (the excel went to the party) / Cancelled. Sold only happens
	through sell_preparation."""
	if status not in ("Draft", "Sent", "Cancelled"):
		frappe.throw(frappe._("Bad status."))
	d = frappe.get_doc("Sale Preparation", name)
	if d.status == "Sold":
		frappe.throw(frappe._("{0} is already sold.").format(name))
	d.status = status
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return {"status": d.status}


@frappe.whitelist()
def export_sale_prep_xlsx(name):
	"""The default sale-confirmation excel for the party. (Company-specific
	templates plug in later, one exporter per format.)"""
	from io import BytesIO
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill
	d = get_sale_preparation(name)
	wb = Workbook()
	ws = wb.active
	ws.title = "Sale Confirmation"
	head_font, head_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E5F")
	ws.append(["Sale Confirmation", "", d["customer"], "", "Chart: " + (d["price_chart"] or ""),
		"", "Gold rate: " + str(d["gold_rate"])])
	ws.append([])
	cols = ["Card", "Design", "Type", "Nett g", "Dmd ct", "Solitaire ct", "Stone ct",
		"Gold", "Diamond", "Stone", "Labour", "Charges", "Total"]
	ws.append(cols)
	for c in ws[3]:
		c.font, c.fill = head_font, head_fill
	for r in d["items"]:
		ws.append([r["order_bag"], r["design"], r["design_type"], r["nett"], r["dmd_ct"],
			r["solitaire_ct"], r["ostone_ct"], r["gold_value"], r["diamond_value"],
			r["stone_value"], r["labour_value"], r["charges_value"], r["piece_total"]])
	ws.append([])
	ws.append(["", "", "", "", "", "", "", "", "", "", "", "GRAND TOTAL", d["grand_total"]])
	ws[ws.max_row][12].font = Font(bold=True)
	for i, w in enumerate([16, 14, 12, 9, 9, 11, 9, 11, 11, 10, 10, 10, 12], 1):
		ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = w
	buf = BytesIO()
	wb.save(buf)
	frappe.local.response.filename = "SaleConfirmation-{0}-{1}.xlsx".format(d["customer"].replace(" ", "-"), name)
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "download"


@frappe.whitelist()
def sell_preparation(name):
	"""The actual sale: hands the preparation's FINAL values to create_product_sale
	(stock write-off, bags Sold, holder transfer) and locks the prep as Sold."""
	d = frappe.get_doc("Sale Preparation", name)
	if d.status == "Sold":
		frappe.throw(frappe._("{0} is already sold.").format(name))
	if not d.items:
		frappe.throw(frappe._("Nothing on the list."))
	lines = []
	for r in d.items:
		lines.append({"order_bag": r.order_bag, "design": r.design, "design_type": r.design_type,
			"held_by": frappe.db.get_value("Order Bag", r.order_bag, "held_by"),
			"nett": flt(r.nett), "dmd_ct": flt(r.dmd_ct), "ostone_ct": flt(r.ostone_ct),
			**{f: flt(r.get(f)) for f in PREP_VALUE_FIELDS}})
	overrides = [r for r in d.items if r.overridden]
	remark_bits = ["{0}: chart {1} -> sold {2} (by {3}){4}".format(
		r.order_bag,
		round(sum(flt(r.get("chart_" + f.replace("_value", ""))) for f in PREP_VALUE_FIELDS), 2),
		flt(r.piece_total), r.changed_by,
		" — " + r.override_remark if r.override_remark else "") for r in overrides]
	res = create_product_sale(json.dumps({
		"customer": d.customer, "price_chart": d.price_chart, "gold_rate": flt(d.gold_rate),
		"remarks": ((d.remarks or "") + ("\nManual price changes:\n" + "\n".join(remark_bits) if remark_bits else "")).strip(),
		"lines": lines,
	}))
	d.status = "Sold"
	d.sale = res["name"]
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return {**res, "prep": name}


# ---------------------------------------------------------------------------
# Sell (Sales) — price scanned pieces against a Price Chart, record the sale,
# write the stock off. The verified costing math: gold = nett x effective rate;
# diamonds = ct x cents-bracket rate; labour per rule (gram+min / piece /
# purity-percent); pass-through charges per piece.
# ---------------------------------------------------------------------------
@frappe.whitelist()
def get_sale_piece(barcode, price_chart, gold_rate=0):
	"""Price one scanned piece against the chart. Guards: finished + In Stock;
	every diamond quality in the piece must have a chart row (quality-blank rows
	accept any). Values are suggestions — the Sell page keeps them editable."""
	nm = (barcode or "").strip()
	if not frappe.db.exists("Order Bag", nm):
		frappe.throw(frappe._("{0} not found.").format(nm or "?"))
	b = frappe.db.get_value("Order Bag", nm, [
		"name", "design", "held_by", "stock_status", "is_finished", "huid", "qty", "certifications",
		"act_gross_weight", "act_nett_weight", "act_dmd_weight", "act_dmd_no",
		"act_ps_weight", "act_cs_weight", "act_cz_weight", "act_cvd_weight", "act_pdmd_weight", "act_poth_weight",
		"act_cs_no", "act_cz_no", "act_cvd_no",
	], as_dict=True)
	if not b.is_finished:
		frappe.throw(frappe._("{0} is not a product yet.").format(nm))
	if b.stock_status != "In Stock":
		frappe.throw(frappe._("{0} is {1} — only pieces In Stock can be sold.").format(nm, b.stock_status))
	chart = frappe.get_doc("Price Chart", price_chart)
	gold_rate = flt(gold_rate)

	design_type = (frappe.db.get_value("Design", b.design, "design_type") if b.design else "") or ""
	tags = [r.charge_category for r in frappe.get_all(
		"Order Bag Charge Category", filters={"parent": nm}, fields=["charge_category"])]

	# ---- diamonds ---------------------------------------------------------------
	# Per ITEM line: per-stone ct = the line's carats divided equally by its piece
	# count. Above the chart's solitaire threshold the line is a SOLITAIRE — pulled
	# out of the diamond totals and priced from the solitaire brackets (missing
	# bracket = scan denied). Qualities map through the global Diamond Quality Map
	# (VVS1-EF rates as VVS-EF) before the chart lookup.
	qmap = _diamond_qmap()
	bom_pcs = {}
	bag_doc = frappe.get_doc("Order Bag", nm)
	for r in bag_doc.bag_bom:
		bom_pcs[r.item] = bom_pcs.get(r.item, 0) + cint(flt(r.qty) * (bag_doc.qty or 1))
	sol_min = flt(getattr(chart, "solitaire_min_ct", 0)) or 0.07

	qual_ct, sol_lines = {}, []
	for item, qty in _bag_convert_materials([nm])[nm].items():
		st, grp = frappe.db.get_value("Item", item, ["stone_type", "item_group"]) or ("", "")
		if st != "Diamond":
			continue
		q = (grp or "").replace("DIAMOND ", "")
		q = qmap.get(q, q)
		pcs = bom_pcs.get(item, 0)
		line_per_stone = (flt(qty) / pcs) if pcs else 0
		if line_per_stone > sol_min and getattr(chart, "solitaire_rates", None) is not None:
			sol_lines.append({"item": item, "quality": q, "ct": flt(qty), "pcs": pcs, "per_stone": line_per_stone})
		else:
			qual_ct[q] = qual_ct.get(q, 0) + flt(qty)

	diamond_value = 0.0
	dmd_detail = []
	per_stone = (flt(b.act_dmd_weight) / cint(b.act_dmd_no)) if cint(b.act_dmd_no) else 0
	for q, ct in qual_ct.items():
		rows = [r for r in chart.diamond_rates if (r.quality or "") in (q, "")]
		if not rows:
			frappe.throw(frappe._("{0}: quality {1} is not on chart {2} — scan denied.").format(nm, q, chart.chart_name))
		exact = [r for r in rows if (r.quality or "") == q] or rows
		row = None
		if per_stone:
			for r in exact:
				if flt(r.from_ct) <= per_stone and (not flt(r.to_ct) or per_stone < flt(r.to_ct)):
					row = r
					break
		row = row or sorted(exact, key=lambda r: flt(r.from_ct))[0]
		diamond_value += ct * flt(row.rate)
		dmd_detail.append({"quality": q, "ct": round(ct, 3), "rate": flt(row.rate)})

	# solitaires: their own brackets, their own line on the bill
	solitaire_value, solitaire_ct, sol_detail = 0.0, 0.0, []
	for s in sol_lines:
		rows = [r for r in (chart.get("solitaire_rates") or [])
			if (r.quality or "") in (s["quality"], "")
			and flt(r.from_ct) <= s["per_stone"] and (not flt(r.to_ct) or s["per_stone"] < flt(r.to_ct))]
		if not rows:
			frappe.throw(frappe._("{0}: solitaire {1} ct/stone ({2}) has no solitaire bracket on chart {3} — scan denied.").format(
				nm, round(s["per_stone"], 3), s["quality"], chart.chart_name))
		exact = [r for r in rows if (r.quality or "") == s["quality"]] or rows
		rate = flt(exact[0].rate)
		solitaire_value += s["ct"] * rate
		solitaire_ct += s["ct"]
		sol_detail.append({"quality": s["quality"], "pcs": s["pcs"], "per_stone": round(s["per_stone"], 3),
			"ct": round(s["ct"], 3), "rate": rate})

	# ---- coloured buckets: CS / CZ / CVD each price from their OWN bracket
	# table (blank-range row = flat). A bucket present on the piece with no
	# rows = scan denied. Precious stones: per-stone rows only, same law.
	def bucket_value(label, field, ct, pcs):
		"""Bracket picked by total carats; the row's basis decides the maths —
		Per Ct = ct x rate, Per Piece = stone count x rate."""
		ct = flt(ct)
		if ct <= 0:
			return 0.0
		rows = list(chart.get(field) or [])
		if not rows:
			frappe.throw(frappe._("{0} carries {1} ct {2} but chart {3} has no {2} rates — scan denied.").format(
				nm, round(ct, 3), label, chart.chart_name))
		row = next((r for r in rows if flt(r.from_ct) <= ct and (not flt(r.to_ct) or ct < flt(r.to_ct))), None)
		if not row:
			frappe.throw(frappe._("{0}: {1} ct {2} falls outside every {2} bracket on chart {3} — scan denied.").format(
				nm, round(ct, 3), label, chart.chart_name))
		if (row.basis or "Per Ct") == "Per Piece":
			if not cint(pcs):
				frappe.throw(frappe._("{0}: {1} is priced per piece but the piece count is 0 — scan denied.").format(nm, label))
			return cint(pcs) * flt(row.rate)
		return ct * flt(row.rate)

	ps_rows = {(r.stone or "").upper(): flt(r.rate) for r in (chart.get("precious_stone_rates") or [])}
	ps_value = 0.0
	ps_detail = []
	if flt(b.act_ps_weight) > 0 and not ps_rows:
		frappe.throw(frappe._("{0} carries precious stones but chart {1} has no Precious Stone rates — scan denied.").format(
			nm, chart.chart_name))
	if ps_rows:
		for item, qty in _bag_convert_materials([nm])[nm].items():
			if frappe.db.get_value("Item", item, "stone_type") != "Precious Stone":
				continue
			rate = ps_rows.get(item.upper())
			if rate is None:
				frappe.throw(frappe._("{0} carries {1} but chart {2} has no rate for it — scan denied.").format(
					nm, item, chart.chart_name))
			ps_value += flt(qty) * rate
			ps_detail.append({"stone": item, "ct": round(flt(qty), 3), "rate": rate})
	stone_value = (bucket_value("CS", "cs_rates", b.act_cs_weight, b.act_cs_no)
		+ bucket_value("CZ", "cz_rates", b.act_cz_weight, b.act_cz_no)
		+ bucket_value("CVD", "cvd_rates", b.act_cvd_weight, b.act_cvd_no) + ps_value)
	ostone_ct = flt(b.act_cs_weight) + flt(b.act_cz_weight) + flt(b.act_ps_weight) + flt(b.act_cvd_weight) + flt(b.act_poth_weight)
	# party diamonds have no pricing route since the flat job-work rate was
	# retired — a piece carrying them can't be priced yet
	if flt(b.act_pdmd_weight) > 0:
		frappe.throw(frappe._("{0} carries party diamonds — no job-work pricing is configured yet.").format(nm))
	job_work = 0.0

	# ---- gold + making --------------------------------------------------------
	nett = flt(b.act_nett_weight)
	gold_value = nett * gold_rate
	labour = 0.0
	rule_desc = ""
	making_rules = list(chart.get("making_rules") or [])
	if making_rules:
		# per design type, blank row = DEFAULT; the minimum is a RUPEE floor
		# (rate 1500/g but anything under 1250 bills as 1250)
		row = next((r for r in making_rules if (r.design_type or "") == design_type and design_type), None) \
			or next((r for r in making_rules if not r.design_type), None)
		if not row:
			frappe.throw(frappe._("{0} ({1}) has no making rule on chart {2} and no DEFAULT row — scan denied.").format(
				nm, design_type or "no type", chart.chart_name))
		# making is PER GRAM, always — nett x rate, floored at the minimum rupees
		labour = nett * flt(row.rate)
		rule_desc = "{0} g x {1}/g".format(round(nett, 3), flt(row.rate))
		if flt(row.min_per_piece) and labour < flt(row.min_per_piece):
			labour = flt(row.min_per_piece)
			rule_desc += " (floored to {0})".format(flt(row.min_per_piece))
		if row.design_type:
			rule_desc += " [{0}]".format(row.design_type)
		else:
			rule_desc += " [default]"
	elif flt(chart.making_rate):
		# legacy flat rule: under the minimum grams bills AS the minimum
		min_g = flt(chart.making_min_grams) or 1
		billed_g = max(nett, min_g)
		labour = billed_g * flt(chart.making_rate)
		rule_desc = "{0} g x {1}/g".format(round(billed_g, 3), flt(chart.making_rate))
		if nett < min_g:
			rule_desc += " (min {0} g)".format(min_g)
	labour += job_work
	# certifications price ONLY through the chart's certification rows;
	# HALLMARKING rows scale per HUID (a stud pair carries two and pays twice)
	huids = [x for x in re.split(r"[,/\s]+", b.huid or "") if x]
	pieces = max(len(huids), 1)
	charges = 0.0

	# certifications the bag ACTUALLY carries, charged per the chart's rows. A
	# certification the chart hasn't priced BLOCKS the scan (rate 0 = free is fine).
	cert_detail = []
	trail = [x.strip().upper() for x in (b.certifications or "").split(",") if x.strip()]
	cert_rows = {(r.certification or "").upper(): r for r in (chart.get("certification_charges") or [])}
	if trail and not cert_rows:
		frappe.throw(frappe._("{0} is certified ({1}) but chart {2} has no Certification Charges — scan denied.").format(
			nm, ", ".join(trail), chart.chart_name))
	if cert_rows or trail:
		for token in trail:
			if cert_rows:
				is_hall = token in ("HALL", "HALLMARKING")
				# labs may be priced individually OR through the ALL LABS group
				# row (never both — save blocks it); hallmarking only ever
				# matches its own row
				row = cert_rows.get(token)
				if row is None and not is_hall:
					row = cert_rows.get("ALL LABS")
				if row is None:
					frappe.throw(frappe._("{0} is {1} certified but chart {2} has no price for {1} — scan denied.").format(
						nm, token, chart.chart_name))
				if not is_hall and (row.basis or "Per Piece") == "Per Ct":
					# lab certs may bill on the piece's DMD carats, floored at
					# the row's minimum (600/ct but never under 150)
					val = flt(b.act_dmd_weight) * flt(row.rate)
					if flt(row.min_amount) and val < flt(row.min_amount):
						val = flt(row.min_amount)
					charges += val
					cert_detail.append({"certification": token, "rate": flt(row.rate),
						"basis": "Per Ct", "ct": flt(b.act_dmd_weight), "value": round(val, 2),
						"via": "ALL LABS" if token not in cert_rows else token})
				else:
					# per piece; HALLMARKING scales per HUID (a stud pair pays twice)
					mult = pieces if is_hall else 1
					charges += flt(row.rate) * mult
					cert_detail.append({"certification": token, "rate": flt(row.rate), "pieces": mult,
						"via": "ALL LABS" if (token not in cert_rows and not is_hall) else token})
			else:
				# legacy chart without the table: the old flat charge covers everything
				pass

	# solitaires ride inside diamond_value but stay OUT of the piece's diamond ct
	dmd_ct_regular = flt(b.act_dmd_weight) + flt(b.act_pdmd_weight) - solitaire_ct
	return {
		"order_bag": nm, "design": b.design or "", "design_type": design_type,
		"held_by": b.held_by or "", "huid": b.huid or "",
		"gross": flt(b.act_gross_weight), "nett": nett,
		"dmd_ct": round(dmd_ct_regular, 3), "solitaire_ct": round(solitaire_ct, 3),
		"ostone_ct": round(ostone_ct, 3),
		"gold_value": round(gold_value, 2),
		"diamond_value": round(diamond_value + solitaire_value, 2),
		"stone_value": round(stone_value, 2), "labour_value": round(labour, 2),
		"charges_value": round(charges, 2),
		"dmd_detail": dmd_detail, "sol_detail": sol_detail, "cert_detail": cert_detail,
		"ps_detail": ps_detail,
		"labour_rule": rule_desc,
	}


@frappe.whitelist()
def create_product_sale(payload):
	"""Record the sale: Product Sale doc + ONE Material Issue writing the pieces'
	materials out of Finished Goods + bags -> Sold (kept for returns), held_by ->
	the buyer (logged as a Holder Transfer)."""
	p = frappe.parse_json(payload)
	customer = p.get("customer")
	if not customer or not frappe.db.exists("Customer", customer):
		frappe.throw(frappe._("Pick who you are selling to."))
	lines = p.get("lines") or []
	if not lines:
		frappe.throw(frappe._("Scan at least one piece."))
	bags = [l.get("order_bag") for l in lines]
	if len(set(bags)) != len(bags):
		frappe.throw(frappe._("Duplicate pieces on the bill."))
	for nm in bags:
		b = frappe.db.get_value("Order Bag", nm, ["is_finished", "stock_status"], as_dict=True)
		if not b or not b.is_finished or b.stock_status != "In Stock":
			frappe.throw(frappe._("{0} is not a piece In Stock.").format(nm))

	# the write-off: everything the pieces hold leaves Finished Goods
	totals = {}
	for mats in _bag_convert_materials(bags).values():
		for it, q in mats.items():
			totals[it] = totals.get(it, 0) + q
	fg = _wh("Finished Goods")
	se = frappe.get_doc({
		"doctype": "Stock Entry", "stock_entry_type": "Material Issue", "company": _company(),
		"items": [{
			"item_code": it, "qty": q,
			"uom": frappe.db.get_value("Item", it, "stock_uom") or "Gram",
			"s_warehouse": fg, "allow_zero_valuation_rate": 1,
		} for it, q in totals.items() if flt(q) > 0],
	})
	se.flags.ignore_permissions = True
	se.insert()
	se.submit()

	sums = {k: 0.0 for k in ("gold_value", "diamond_value", "stone_value", "labour_value", "charges_value")}
	rows = []
	for l in lines:
		nm = l["order_bag"]
		vals = {k: flt(l.get(k)) for k in sums}
		total = round(sum(vals.values()), 2)
		for k in sums:
			sums[k] += vals[k]
		rows.append({
			"order_bag": nm, "design": l.get("design"), "design_type": l.get("design_type"),
			"holder_at_sale": l.get("held_by") or None,
			"nett": flt(l.get("nett")), "dmd_ct": flt(l.get("dmd_ct")), "ostone_ct": flt(l.get("ostone_ct")),
			**vals, "piece_total": total,
		})
	sale = frappe.get_doc({
		"doctype": "Product Sale",
		"customer": customer, "sale_date": frappe.utils.today(), "status": "Completed",
		"price_chart": p.get("price_chart"), "gold_rate": flt(p.get("gold_rate")),
		"remarks": p.get("remarks"), "stock_entry": se.name, "items": rows,
		**{k: round(v, 2) for k, v in sums.items()},
		"grand_total": round(sum(sums.values()), 2),
	})
	sale.insert(ignore_permissions=True)

	now = frappe.utils.now_datetime()
	for l in lines:
		nm = l["order_bag"]
		old_holder = frappe.db.get_value("Order Bag", nm, "held_by")
		if (old_holder or "") != customer:
			ht = frappe.get_doc({
				"doctype": "Holder Transfer", "order_bag": nm, "from_holder": old_holder,
				"to_holder": customer, "transfer_time": now, "transferred_by": frappe.session.user,
				"reason": "Sold via {0}".format(sale.name),
			})
			ht.flags.ignore_permissions = True
			ht.insert(ignore_permissions=True)
		frappe.db.set_value("Order Bag", nm, {"stock_status": "Sold", "held_by": customer})
	frappe.db.commit()
	return {"name": sale.name, "grand_total": sale.grand_total, "stock_entry": se.name, "count": len(rows)}


# ---------------------------------------------------------------------------
# Transfer Holder (Delivery) — move a piece's reservation to another customer,
# every move written to a Holder Transfer record (the full paper trail).
# ---------------------------------------------------------------------------
_BUCKET_LABELS = ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth")


@frappe.whitelist()
def get_holder_piece(barcode):
	"""Resolve a scanned card for the Transfer Holder page: current holder, when it
	(re)entered stock, and its frozen weights (gross / pure / per stone bucket)."""
	nm = (barcode or "").strip()
	if not frappe.db.exists("Order Bag", nm):
		frappe.throw(frappe._("{0} not found.").format(nm or "?"))
	b = frappe.db.get_value("Order Bag", nm, [
		"name", "design", "held_by", "stock_status", "is_finished", "in_stock_on",
		"act_gross_weight", "act_pure_weight",
	] + [f"act_{x}_weight" for x in _BUCKET_LABELS], as_dict=True)
	if not b.is_finished:
		frappe.throw(frappe._("{0} is not a product yet — it's still on the floor.").format(nm))
	if b.stock_status != "In Stock":
		frappe.throw(frappe._("{0} is {1} — only pieces In Stock can change holder.").format(nm, b.stock_status))
	return {
		"order_bag": b.name, "design": b.design or "",
		"design_type": (frappe.db.get_value("Design", b.design, "design_type") if b.design else "") or "",
		"held_by": b.held_by or "", "in_stock_on": str(b.in_stock_on or ""),
		"gross": flt(b.act_gross_weight), "pure": flt(b.act_pure_weight),
		"buckets": {x: flt(b.get(f"act_{x}_weight")) for x in _BUCKET_LABELS},
	}


@frappe.whitelist()
def transfer_holder(bags, to_customer, reason=None):
	"""Move the hold on the given pieces to `to_customer` (JD Stock = release to
	free stock). One Holder Transfer record per piece; held_by updated."""
	if isinstance(bags, str):
		bags = json.loads(bags or "[]")
	bags = [b for b in bags if b]
	if not bags:
		frappe.throw(frappe._("Scan at least one piece."))
	if not to_customer or not frappe.db.exists("Customer", to_customer):
		frappe.throw(frappe._("Pick the new holder (JD Stock = our own shelf)."))
	rows = []
	for nm in bags:
		b = frappe.db.get_value("Order Bag", nm, ["is_finished", "stock_status", "held_by"], as_dict=True)
		if not b or not b.is_finished or b.stock_status != "In Stock":
			frappe.throw(frappe._("{0} is not a piece In Stock.").format(nm))
		if (b.held_by or "") == to_customer:
			frappe.throw(frappe._("{0} is already held by {1}.").format(nm, to_customer))
		rows.append((nm, b.held_by or None))
	now = frappe.utils.now_datetime()
	made = []
	for nm, from_holder in rows:
		ht = frappe.get_doc({
			"doctype": "Holder Transfer",
			"order_bag": nm, "from_holder": from_holder, "to_holder": to_customer,
			"transfer_time": now, "transferred_by": frappe.session.user, "reason": reason,
		})
		ht.flags.ignore_permissions = True
		ht.insert(ignore_permissions=True)
		frappe.db.set_value("Order Bag", nm, "held_by", to_customer)
		made.append(ht.name)
	frappe.db.commit()
	return {"count": len(made), "transfers": made, "to": to_customer}


@frappe.whitelist()
def get_recent_holder_transfers(limit=25):
	"""Freshest holder moves — the Transfer Holder page's side feed."""
	return frappe.get_all(
		"Holder Transfer",
		fields=["name", "order_bag", "from_holder", "to_holder", "transfer_time", "reason"],
		order_by="transfer_time desc", limit=cint(limit) or 25,
	)


@frappe.whitelist()
def get_certifiable_pieces(search=None):
	"""Finished pieces In Stock — the pool the Certification desk picks from."""
	bags = frappe.get_all(
		"Order Bag", filters={"is_finished": 1, "stock_status": "In Stock"},
		fields=["name", "design", "held_by", "act_gross_weight", "act_dmd_weight", "act_dmd_no", "huid"],
		order_by="modified desc", limit=500,
	)
	designs = list({b.design for b in bags if b.design})
	dtype = {d.name: (d.design_type or "") for d in frappe.get_all(
		"Design", filters={"name": ["in", designs or [""]]}, fields=["name", "design_type"])}
	out = []
	s = (search or "").strip().lower()
	for b in bags:
		row = {
			"order_bag": b.name, "design": b.design or "", "design_type": dtype.get(b.design, ""),
			"held_by": b.held_by or "", "gross": flt(b.act_gross_weight), "dmd_ct": flt(b.act_dmd_weight),
			"dmd_no": cint(b.act_dmd_no), "huid": b.huid or "",
		}
		if s and s not in " ".join([row["order_bag"], row["design"], row["design_type"], row["held_by"]]).lower():
			continue
		out.append(row)
	return out


# ---------------------------------------------------------------------------
# Certification PREP (the reworked desk): pick the certification + center,
# lock the rules/format, scan products in (rejections recorded client-side),
# then SEND from the Send Certifications page. The prep IS the batch — it is
# named by the certification code series (IGI-0001) the moment it's created.
# ---------------------------------------------------------------------------
def _diamond_qmap():
	"""quality -> parent quality. TWO sources, tree first: a DIAMOND leaf sitting
	under an 'X GROUP' Item Group resolves to X (drop a new batch-leaf under the
	parent and it just works); Diamond Quality Map rows override/extend."""
	qmap = {}
	for g in frappe.get_all("Item Group",
			filters={"name": ["like", "DIAMOND %"], "is_group": 0},
			fields=["name", "parent_item_group"]):
		if (g.parent_item_group or "").endswith(" GROUP"):
			qmap[g.name.replace("DIAMOND ", "")] = g.parent_item_group[:-6].strip()
	if frappe.db.exists("DocType", "Diamond Quality Map"):
		for r in frappe.get_all("Diamond Quality Map", fields=["name", "parent_quality"]):
			qmap[r.name] = r.parent_quality
	return qmap


def _bag_diamond_qualities(nm):
	"""The DISTINCT parent-mapped diamond qualities a finished piece carries."""
	qmap = _diamond_qmap()
	quals = set()
	for item in _bag_convert_materials([nm])[nm]:
		st, grp = frappe.db.get_value("Item", item, ["stone_type", "item_group"]) or ("", "")
		if st == "Diamond":
			q = (grp or "").replace("DIAMOND ", "")
			quals.add(qmap.get(q, q))
	return sorted(quals)


@frappe.whitelist()
def get_cert_prep_context():
	"""The certify page's setup: certifications with centers + requirements, and
	the quality options (parent-mapped diamond item groups) for the IGI lock."""
	types = frappe.get_all("Certification Type", fields=["name", "title", "excel_requirements"], order_by="name")
	centers = frappe.get_all("Certification Center",
		fields=["name", "certification_type", "center_name"], order_by="center_name")
	qmap = _diamond_qmap()
	groups = frappe.get_all("Item Group", filters={"name": ["like", "DIAMOND %"], "is_group": 0}, pluck="name")
	quals = sorted({qmap.get(g.replace("DIAMOND ", ""), g.replace("DIAMOND ", "")) for g in groups})
	return {"types": types, "centers": centers, "qualities": quals}


def _cert_validate_piece(cert_type, quality, nm, taken=None):
	"""All the scan guards, WITHOUT writing anything. Returns the piece's basics."""
	if not frappe.db.exists("Order Bag", nm):
		frappe.throw(frappe._("{0} does not exist.").format(nm or "?"))
	b = frappe.db.get_value("Order Bag", nm,
		["is_finished", "stock_status", "design", "act_gross_weight", "act_dmd_weight"], as_dict=True)
	if not b.is_finished:
		frappe.throw(frappe._("{0} is not a product yet — MAKE IT A PRODUCT first (Make Products).").format(nm))
	if b.stock_status != "In Stock":
		frappe.throw(frappe._("{0} is {1} — only pieces In Stock can go out.").format(nm, b.stock_status))
	if taken and nm in taken:
		frappe.throw(frappe._("{0} is already on this list.").format(nm))
	other = frappe.db.sql("""select i.parent from `tabCertification Item` i
		join `tabCertification` c on c.name = i.parent
		where i.order_bag = %s and c.status = 'Prepared' limit 1""", (nm,))
	if other:
		frappe.throw(frappe._("{0} is already on prepared batch {1}.").format(nm, other[0][0]))
	if cert_type == "IGI":
		quals = _bag_diamond_qualities(nm)
		if len(quals) > 1:
			frappe.throw(frappe._("{0} carries MIXED diamond qualities ({1}) — IGI batches take one only.").format(nm, ", ".join(quals)))
		if not quals:
			frappe.throw(frappe._("{0} has no diamonds — nothing for IGI to certify.").format(nm))
		if quals[0] != quality:
			frappe.throw(frappe._("{0} is {1} — this batch is locked to {2}.").format(nm, quals[0], quality))
	return b


def _cert_format_row(cert_type, quality, nm, b):
	row = {"order_bag": nm, "design": b.design or "",
		"design_type": (frappe.db.get_value("Design", b.design, "design_type") if b.design else "") or "",
		"gross": flt(b.act_gross_weight), "dmd_ct": flt(b.act_dmd_weight)}
	if cert_type == "IGI":
		color = "Gold"
		for it in _bag_convert_materials([nm])[nm]:
			if not frappe.db.get_value("Item", it, "stone_type"):
				g = re.search(r"(\d{2}K)(YG|WG|PG)$", it)
				if g:
					color = _IGI_METAL_COLOR[g.group(2)]
					break
		row.update({"style_no": "{0} {1}".format(b.design or "", nm).strip(),
			"metal_color": color,
			"color": (_IGI_QUALITY.get(quality or "", ("", "")))[0],
			"clarity": (_IGI_QUALITY.get(quality or "", ("", "")))[1],
			"shape": "Round Brilliant"})
	return row


@frappe.whitelist()
def get_bag_bom_summary(order_bag):
	"""Hover summary: the piece's ACTUAL frozen composition (Convert rows), one
	line per item — '22KPG 8g' / 'VVS-EF 0.5-2 10pc 0.9ct'."""
	nm = (order_bag or "").strip()
	if not frappe.db.exists("Order Bag", nm):
		return {"lines": []}
	pcs_by_item = {}
	for r in frappe.get_all("Bag Material Ledger",
			filters={"order_bag": nm, "entry_type": "Convert", "direction": "Out"},
			fields=["item", "pcs"]):
		pcs_by_item[r.item] = pcs_by_item.get(r.item, 0) + int(r.pcs or 0)
	lines = []
	for item, qty in sorted(_bag_convert_materials([nm])[nm].items()):
		st = frappe.db.get_value("Item", item, "stone_type")
		if st:
			pcs = pcs_by_item.get(item, 0)
			lines.append("{0}{1} {2}ct".format(item, " {0}pc".format(pcs) if pcs else "", round(flt(qty), 3)))
		else:
			lines.append("{0} {1}g".format(item, round(flt(qty), 3)))
	return {"lines": lines}


@frappe.whitelist()
def cert_draft_scan(cert_type, quality, barcode, existing=None):
	"""Validate ONE scan for the local (unsaved) draft list — nothing is written.
	Throws with the reason on any rejection; returns the format row otherwise."""
	if isinstance(existing, str):
		existing = json.loads(existing or "[]")
	nm = (barcode or "").strip()
	try:
		b = _cert_validate_piece(cert_type, (quality or "").strip(), nm, set(existing or []))
	except frappe.ValidationError as e:
		frappe.local.message_log = []  # no modal — the page logs it in scan history
		return {"rejected": str(e)}
	return _cert_format_row(cert_type, quality, nm, b)


@frappe.whitelist()
def cert_prep_create_full(cert_type, center=None, quality=None, bags=None):
	"""PREP: the draft list becomes the real batch in one shot — re-validated
	piece by piece, then named by the code series (IGI-0001)."""
	if isinstance(bags, str):
		bags = json.loads(bags or "[]")
	bags = [b for b in (bags or []) if b]
	if not bags:
		frappe.throw(frappe._("Scan at least one piece before prepping."))
	if not frappe.db.exists("Certification Type", cert_type):
		frappe.throw(frappe._("Pick the certification."))
	quality = (quality or "").strip()
	if cert_type == "IGI" and not quality:
		frappe.throw(frappe._("IGI batches carry ONE colour+clarity — pick it first."))
	rows, seen = [], set()
	for nm in bags:
		b = _cert_validate_piece(cert_type, quality, nm, seen)
		seen.add(nm)
		rows.append({"order_bag": nm, "design": b.design,
			"design_type": (frappe.db.get_value("Design", b.design, "design_type") if b.design else "") or "",
			"gross": flt(b.act_gross_weight), "dmd_ct": flt(b.act_dmd_weight)})
	legacy = {"HALL": "HALLMARKING", "SGL": "SGL", "IDT": "IDT", "GIG": "GIG"}.get(cert_type)
	d = frappe.get_doc({"doctype": "Certification", "cert_type": cert_type,
		"center": center or None, "quality": quality, "status": "Prepared",
		"prepared_on": frappe.utils.today(), "certification_type": legacy, "items": rows})
	d.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name, "count": len(rows)}


@frappe.whitelist()
def create_cert_prep(cert_type, center=None, quality=None):
	"""Start a prep — it takes its FINAL outgoing name now (IGI-0001)."""
	if not frappe.db.exists("Certification Type", cert_type):
		frappe.throw(frappe._("Pick the certification."))
	if cert_type == "IGI" and not (quality or "").strip():
		frappe.throw(frappe._("IGI batches carry ONE colour+clarity — pick it first."))
	# legacy Select filled only where the old option list has the value (receive
	# flow reads it for HUID handling); the new cert_type is the real key
	legacy = {"HALL": "HALLMARKING", "SGL": "SGL", "IDT": "IDT", "GIG": "GIG"}.get(cert_type)
	d = frappe.get_doc({"doctype": "Certification", "cert_type": cert_type,
		"center": center or None, "quality": (quality or "").strip(),
		"status": "Prepared", "prepared_on": frappe.utils.today(),
		"certification_type": legacy})
	d.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": d.name}


@frappe.whitelist()
def cert_prep_scan(name, barcode):
	"""Scan a piece into the prep. Rejections THROW with the reason (the page
	logs them in its scan history): not found / not a product yet (make product
	first) / not In Stock / already on this or another open batch / IGI quality
	mismatch or mixed-quality piece."""
	d = frappe.get_doc("Certification", name)
	if d.status != "Prepared":
		frappe.throw(frappe._("{0} is {1} — no more scanning.").format(name, d.status))
	nm = (barcode or "").strip()
	if not frappe.db.exists("Order Bag", nm):
		frappe.throw(frappe._("{0} does not exist.").format(nm or "?"))
	b = frappe.db.get_value("Order Bag", nm,
		["is_finished", "stock_status", "design", "act_gross_weight", "act_dmd_weight", "huid"], as_dict=True)
	if not b.is_finished:
		frappe.throw(frappe._("{0} is not a product yet — MAKE IT A PRODUCT first (Make Products).").format(nm))
	if b.stock_status != "In Stock":
		frappe.throw(frappe._("{0} is {1} — only pieces In Stock can go out.").format(nm, b.stock_status))
	if any(r.order_bag == nm for r in d.items):
		frappe.throw(frappe._("{0} is already on this batch.").format(nm))
	other = frappe.db.sql("""select i.parent from `tabCertification Item` i
		join `tabCertification` c on c.name = i.parent
		where i.order_bag = %s and c.status = 'Prepared' and c.name != %s limit 1""", (nm, name))
	if other:
		frappe.throw(frappe._("{0} is already on prepared batch {1}.").format(nm, other[0][0]))
	# IGI: every diamond line must resolve to THE locked quality; mixed = error
	if d.cert_type == "IGI":
		quals = _bag_diamond_qualities(nm)
		if len(quals) > 1:
			frappe.throw(frappe._("{0} carries MIXED diamond qualities ({1}) — IGI batches take one only.").format(nm, ", ".join(quals)))
		if quals and quals[0] != d.quality:
			frappe.throw(frappe._("{0} is {1} — this batch is locked to {2}.").format(nm, quals[0], d.quality))
		if not quals:
			frappe.throw(frappe._("{0} has no diamonds — nothing for IGI to certify.").format(nm))
	d.append("items", {"order_bag": nm, "design": b.design,
		"design_type": (frappe.db.get_value("Design", b.design, "design_type") if b.design else "") or "",
		"gross": flt(b.act_gross_weight), "dmd_ct": flt(b.act_dmd_weight)})
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return get_cert_prep(name)


@frappe.whitelist()
def get_cert_prep(name):
	d = frappe.get_doc("Certification", name)
	rows = []
	for r in d.items:
		row = {"row": r.name, "order_bag": r.order_bag, "design": r.design or "",
			"design_type": r.design_type or "", "gross": flt(r.gross), "dmd_ct": flt(r.dmd_ct)}
		if d.cert_type == "IGI":
			# metal colour comes off the gold ITEM CODE suffix (…22KYG -> Yellow
			# Gold), exactly like the IGI excel export does
			color = "Gold"
			for it in _bag_convert_materials([r.order_bag])[r.order_bag]:
				if not frappe.db.get_value("Item", it, "stone_type"):
					g = re.search(r"(\d{2}K)(YG|WG|PG)$", it)
					if g:
						color = _IGI_METAL_COLOR[g.group(2)]
						break
			row.update({"style_no": "{0} {1}".format(r.design or "", r.order_bag).strip(),
				"metal_color": color,
				"color": (_IGI_QUALITY.get(d.quality or "", ("", "")))[0],
				"clarity": (_IGI_QUALITY.get(d.quality or "", ("", "")))[1],
				"shape": "Round Brilliant"})
		rows.append(row)
	return {"name": d.name, "cert_type": d.cert_type, "center": d.center or "",
		"quality": d.quality or "", "status": d.status, "prepared_on": str(d.prepared_on or ""),
		"remarks": d.remarks or "", "rows": rows, "count": len(rows),
		"gross": round(sum(x["gross"] for x in rows), 3), "dmd_ct": round(sum(x["dmd_ct"] for x in rows), 3)}


def _cert_excel_bytes(prep):
	"""The submission excel for a batch as bytes: IGI = the shipped template
	(reuses export_igi_xlsx's fill); others = the basic table."""
	bags = [r["order_bag"] for r in prep["rows"]]
	if prep["cert_type"] == "IGI":
		export_igi_xlsx(json.dumps(bags))
		content = frappe.local.response.filecontent
		fname = frappe.local.response.filename
		frappe.local.response.filecontent = frappe.local.response.filename = frappe.local.response.type = None
		return fname, content
	from io import BytesIO
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill
	wb = Workbook()
	ws = wb.active
	ws.title = prep["cert_type"]
	ws.append(["Card", "Design", "Type", "Gross (g)", "Diamond (ct)"])
	for c in ws[1]:
		c.font, c.fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E5F")
	for r in prep["rows"]:
		ws.append([r["order_bag"], r["design"], r["design_type"], r["gross"], r["dmd_ct"]])
	ws.append(["TOTAL", "", "", prep["gross"], prep["dmd_ct"]])
	ws[ws.max_row][0].font = Font(bold=True)
	for i, w in enumerate([16, 14, 14, 11, 12], 1):
		ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
	buf = BytesIO()
	wb.save(buf)
	return "{0}-{1}.xlsx".format(prep["cert_type"], prep["name"]), buf.getvalue()


@frappe.whitelist()
def get_cert_mail_defaults(name):
	"""What the email prompt prefills: the center's address + its subject/body
	templates with {batch}/{count}/{date} already filled in."""
	p = get_cert_prep(name)
	c = frappe.db.get_value("Certification Center", p["center"],
		["email", "mail_subject", "mail_body", "center_name"], as_dict=True) if p["center"] else None
	# the batch summary the mail carries: pieces by design type + totals
	by_type = {}
	for r in p["rows"]:
		t = r["design_type"] or "UNTYPED"
		by_type[t] = by_type.get(t, 0) + 1
	types = ", ".join("{0} {1}".format(n, t) for t, n in sorted(by_type.items()))
	ph = {"batch": p["name"], "count": p["count"], "date": frappe.utils.today(),
		"types": types, "gross": p["gross"], "dmd": p["dmd_ct"]}
	def render(t, default):
		t = (t or default)
		for k, v in ph.items():
			t = t.replace("{" + k + "}", str(v))
		return t
	return {
		"recipient": (c and c.email) or "",
		"center_name": (c and c.center_name) or "",
		"subject": render(c and c.mail_subject, "Jewelima submission {batch} — {count} piece(s)"),
		"body": render(c and c.mail_body,
			"Dear team,\n\nPlease find attached our submission {batch} dated {date}.\n\n"
			"{count} piece(s) — {types}\nTotal gross weight: {gross} g\nTotal diamond weight: {dmd} ct\n\n"
			"Regards,\nJewelima"),
	}


@frappe.whitelist()
def email_cert_excel(name, recipient, subject, body, cc=None):
	"""Send the batch's submission excel to the center (or whoever the prompt
	says), with optional CCs — every address format-checked before anything
	goes out. Uses the default outgoing account (system@jewelima.com)."""
	from frappe.utils import validate_email_address
	recipient = (recipient or "").strip()
	if not recipient:
		frappe.throw(frappe._("Enter the recipient email."))
	if not validate_email_address(recipient):
		frappe.throw(frappe._("{0} is not a valid email address.").format(recipient))
	cc_list = [x.strip() for x in re.split(r"[,;\s]+", cc or "") if x.strip()]
	for a in cc_list:
		if not validate_email_address(a):
			frappe.throw(frappe._("CC address {0} is not a valid email.").format(a))
	p = get_cert_prep(name)
	if not p["rows"]:
		frappe.throw(frappe._("Nothing on the batch."))
	fname, content = _cert_excel_bytes(p)
	frappe.sendmail(recipients=[recipient], cc=cc_list or None, subject=subject or p["name"],
		message=(body or "").replace("\n", "<br>"),
		attachments=[{"fname": fname, "fcontent": content}], now=True)
	return {"sent_to": recipient, "cc": cc_list, "attachment": fname}


@frappe.whitelist()
def cert_prep_remove(name, row):
	d = frappe.get_doc("Certification", name)
	if d.status != "Prepared":
		frappe.throw(frappe._("{0} is {1} — no more edits.").format(name, d.status))
	d.set("items", [x for x in d.items if x.name != row])
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return get_cert_prep(name)


@frappe.whitelist()
def cert_prep_cancel(name):
	d = frappe.get_doc("Certification", name)
	if d.status != "Prepared":
		frappe.throw(frappe._("Only a Prepared batch cancels — {0} is {1}.").format(name, d.status))
	d.status = "Cancelled"
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return {"status": "Cancelled"}


@frappe.whitelist()
def get_cert_preps():
	"""The Send Certifications page: every Prepared batch, plus recent sent."""
	out = {"prepared": [], "recent": []}
	for r in frappe.get_all("Certification",
			filters={"status": ["in", ["Prepared", "Sent", "Cancelled"]], "cert_type": ["is", "set"]},
			fields=["name", "cert_type", "center", "quality", "status", "prepared_on", "sent_on"],
			order_by="creation desc", limit=40):
		r["pieces"] = frappe.db.count("Certification Item", {"parent": r.name})
		(out["prepared"] if r.status == "Prepared" else out["recent"]).append(r)
	return out


@frappe.whitelist()
def get_certifications_out():
	"""Certification Out board: every SENT batch — days out, piece count split by
	design type, weights. No per-piece rows; collection is whole-batch."""
	out = []
	for c in frappe.get_all("Certification",
			filters={"status": "Sent"},
			fields=["name", "cert_type", "certification_type", "center", "quality", "sent_on"],
			order_by="sent_on asc, creation asc"):
		items = frappe.get_all("Certification Item", filters={"parent": c.name},
			fields=["design_type", "gross", "dmd_ct"])
		by_type = {}
		for r in items:
			t = r.design_type or "UNTYPED"
			by_type[t] = by_type.get(t, 0) + 1
		out.append({
			"name": c.name, "cert_type": c.cert_type or c.certification_type or "?",
			"center": c.center or "", "quality": c.quality or "",
			"sent_on": str(c.sent_on or ""),
			"days_out": frappe.utils.date_diff(frappe.utils.today(), c.sent_on) if c.sent_on else 0,
			"pieces": len(items),
			"by_type": [{"design_type": t, "count": n} for t, n in sorted(by_type.items())],
			"gross": round(sum(flt(r.gross) for r in items), 3),
			"dmd_ct": round(sum(flt(r.dmd_ct) for r in items), 3),
		})
	return {"batches": out, "total_pieces": sum(b["pieces"] for b in out)}


@frappe.whitelist()
def collect_certification(name):
	"""The packet came back: count the pieces and COLLECT the whole batch — stock
	moves At Certification -> Finished Goods, bags flip back In Stock, status ->
	Collected. NOT confirmed yet: HUID / certificate confirmation happens on its
	own page later."""
	from jewelima.setup import CERTIFICATION_WAREHOUSE
	d = frappe.get_doc("Certification", name)
	if d.status != "Sent":
		frappe.throw(frappe._("{0} is {1} — only Sent batches collect.").format(name, d.status))
	bags = [r.order_bag for r in d.items]
	totals = {}
	for mats in _bag_convert_materials(bags).values():
		for it, q in mats.items():
			totals[it] = totals.get(it, 0) + q
	se = _stock_move_many(totals, _wh(CERTIFICATION_WAREHOUSE), _wh("Finished Goods"))
	now = frappe.utils.now_datetime()
	for nm in bags:
		frappe.db.set_value("Order Bag", nm, {"stock_status": "In Stock", "in_stock_on": now})
	d.status = "Collected"
	d.collected_on = frappe.utils.today()
	d.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": name, "pieces": len(bags), "stock_entry": se}


@frappe.whitelist()
def get_confirm_pool():
	"""The Confirm page's pool: every piece on a COLLECTED batch with where it
	stands (pending / confirmed / rejected) — grouped by batch, live-refreshable."""
	batches = []
	for c in frappe.get_all("Certification",
			filters={"status": ["in", ["Collected", "Partially Received"]]},
			fields=["name", "cert_type", "certification_type", "quality", "collected_on"],
			order_by="collected_on asc, creation asc"):
		items = frappe.get_all("Certification Item", filters={"parent": c.name},
			fields=["name", "order_bag", "design_type", "received", "rejected", "confirmed_by"],
			order_by="idx")
		batches.append({"name": c.name, "cert_type": c.cert_type or c.certification_type or "?",
			"quality": c.quality or "", "collected_on": str(c.collected_on or ""),
			"pieces": [{"order_bag": i.order_bag, "design_type": i.design_type or "",
				"state": "confirmed" if i.received else ("rejected" if i.rejected else "pending"),
				"by": i.confirmed_by or ""} for i in items]})
	pend = sum(1 for b in batches for p in b["pieces"] if p["state"] == "pending")
	return {"batches": batches, "pending": pend}


@frappe.whitelist()
def confirm_cert_scan(barcode, mode="accept"):
	"""ONE scan on the Confirm page — lightweight and race-safe for several
	scanners at once. Finds the piece on a Collected batch and marks it
	confirmed (accept) or sends it to the reject queue (reject). Every failure
	comes back as data for the scan history, never a modal."""
	nm = (barcode or "").strip()
	if mode not in ("accept", "reject"):
		mode = "accept"
	row = frappe.db.sql("""select i.name, i.parent, i.received, i.rejected, i.confirmed_by
		from `tabCertification Item` i join `tabCertification` c on c.name = i.parent
		where i.order_bag = %s and c.status in ('Collected', 'Partially Received')
		order by c.creation desc limit 1""", nm, as_dict=True)
	if not row:
		if not frappe.db.exists("Order Bag", nm):
			return {"rejected_scan": frappe._("This card doesn't exist")}
		return {"rejected_scan": frappe._("Not on any collected batch")}
	r = row[0]
	if r.received:
		return {"rejected_scan": frappe._("Already confirmed by {0}").format(r.confirmed_by or "?")}
	if r.rejected:
		return {"rejected_scan": frappe._("Already in the reject queue")}
	now = frappe.utils.now_datetime()
	user = frappe.session.user
	# ATOMIC claim: the guarded UPDATE wins for exactly one scanner — a second
	# simultaneous scan of the same card affects 0 rows and reports the loser
	field = "received" if mode == "accept" else "rejected"
	frappe.db.sql("""update `tabCertification Item`
		set {0} = 1, received_on = %s, confirmed_by = %s
		where name = %s and received = 0 and rejected = 0""".format(field), (now, user, r.name))
	claimed = frappe.db.sql("select confirmed_by from `tabCertification Item` where name = %s", r.name)[0][0]
	if claimed != user:
		return {"rejected_scan": frappe._("Already confirmed by {0}").format(claimed or "?")}
	if mode == "accept":
		ct = frappe.db.get_value("Certification", r.parent, "cert_type") or frappe.db.get_value(
			"Certification", r.parent, "certification_type")
		frappe.db.set_value("Order Bag", nm, "certifications", _stamp_certification(nm, ct))
	# batch rollup: all processed -> Received; some -> Partially Received
	left = frappe.db.sql("""select count(*) from `tabCertification Item`
		where parent = %s and received = 0 and rejected = 0""", r.parent)[0][0]
	frappe.db.set_value("Certification", r.parent, "status",
		"Received" if left == 0 else "Partially Received", update_modified=False)
	frappe.db.commit()
	return {"ok": 1, "mode": mode, "batch": r.parent, "batch_done": left == 0}


@frappe.whitelist()
def send_cert_prep(name):
	"""The actual SEND: one stock move Finished Goods -> At Certification for
	everything the pieces hold, bags flip At Certification, status -> Sent."""
	from jewelima.setup import CERTIFICATION_WAREHOUSE
	d = frappe.get_doc("Certification", name)
	if d.status != "Prepared":
		frappe.throw(frappe._("{0} is {1} — only Prepared batches send.").format(name, d.status))
	if not d.items:
		frappe.throw(frappe._("Nothing on the batch."))
	bags = [r.order_bag for r in d.items]
	for nm in bags:
		b = frappe.db.get_value("Order Bag", nm, ["is_finished", "stock_status"], as_dict=True)
		if not b or not b.is_finished or b.stock_status != "In Stock":
			frappe.throw(frappe._("{0} is no longer In Stock — remove it from the batch.").format(nm))
	totals = {}
	for mats in _bag_convert_materials(bags).values():
		for it, q in mats.items():
			totals[it] = totals.get(it, 0) + q
	se = _stock_move_many(totals, _wh("Finished Goods"), _wh(CERTIFICATION_WAREHOUSE))
	d.stock_entry = se
	d.status = "Sent"
	d.sent_on = frappe.utils.today()
	if not d.certification_type:
		d.certification_type = "HALLMARKING" if d.cert_type == "HALL" else None
	d.save(ignore_permissions=True)
	for nm in bags:
		frappe.db.set_value("Order Bag", nm, "stock_status", "At Certification")
	frappe.db.commit()
	return {"name": d.name, "count": len(bags), "stock_entry": se}


@frappe.whitelist()
def send_certification(payload):
	"""Create a Certification batch: snapshot the pieces, move their materials
	Finished Goods -> At Certification (one Stock Entry), flip the bags to
	At Certification. payload = {certification_type, lab, remarks, bags: [...]}"""
	from jewelima.setup import CERTIFICATION_WAREHOUSE

	p = frappe.parse_json(payload)
	ctype = (p.get("certification_type") or "").upper()
	if ctype not in ("HALLMARKING", "IGL", "DHSC", "SGL", "IDT", "GIG"):
		frappe.throw(frappe._("Unknown certification type: {0}").format(ctype or "?"))
	bags = [b for b in (p.get("bags") or []) if b]
	if not bags:
		frappe.throw(frappe._("Pick at least one piece."))
	rows = []
	for nm in bags:
		b = frappe.db.get_value("Order Bag", nm,
		                        ["is_finished", "stock_status", "design", "act_gross_weight", "act_dmd_weight"], as_dict=True)
		if not b:
			frappe.throw(frappe._("{0} not found.").format(nm))
		if not b.is_finished or b.stock_status != "In Stock":
			frappe.throw(frappe._("{0} is {1} — only finished pieces In Stock can be sent.").format(
				nm, "not a product yet" if not b.is_finished else b.stock_status))
		rows.append({
			"order_bag": nm, "design": b.design,
			"design_type": frappe.db.get_value("Design", b.design, "design_type") if b.design else "",
			"gross": flt(b.act_gross_weight), "dmd_ct": flt(b.act_dmd_weight),
		})

	# the batch's stock move backs the send — everything the pieces hold
	totals = {}
	for mats in _bag_convert_materials(bags).values():
		for it, q in mats.items():
			totals[it] = totals.get(it, 0) + q
	se = _stock_move_many(totals, _wh("Finished Goods"), _wh(CERTIFICATION_WAREHOUSE))

	doc = frappe.get_doc({
		"doctype": "Certification",
		"certification_type": ctype, "status": "Sent", "sent_on": frappe.utils.today(),
		"lab": p.get("lab"), "remarks": p.get("remarks"), "stock_entry": se,
		"items": rows,
	})
	doc.insert(ignore_permissions=True)
	for nm in bags:
		frappe.db.set_value("Order Bag", nm, "stock_status", "At Certification")
	frappe.db.commit()
	return {"name": doc.name, "count": len(bags), "stock_entry": se}


@frappe.whitelist()
def receive_certification(name, rows):
	"""Receive pieces back from a Certification batch: stamp HUID / certificate no
	on the row AND the Order Bag, move their materials At Certification ->
	Finished Goods (one Stock Entry per receive), flip the bags back In Stock.
	rows = [{row (child name), huid}] — HUIDs only come from HALLMARKING; stone
	labs are received by count via receive_certification_all."""
	from jewelima.setup import CERTIFICATION_WAREHOUSE

	if isinstance(rows, str):
		rows = json.loads(rows or "[]")
	doc = frappe.get_doc("Certification", name)
	by_name = {r.name: r for r in doc.items}
	picked = []
	for r in rows or []:
		child = by_name.get(r.get("row"))
		if not child:
			frappe.throw(frappe._("Row {0} not found on {1}.").format(r.get("row") or "?", name))
		if child.received:
			frappe.throw(frappe._("{0} is already received.").format(child.order_bag))
		picked.append((child, (r.get("huid") or "").strip().upper()))
	if not picked:
		frappe.throw(frappe._("Pick at least one piece to receive."))

	totals = {}
	for mats in _bag_convert_materials([c.order_bag for c, _ in picked]).values():
		for it, q in mats.items():
			totals[it] = totals.get(it, 0) + q
	se = _stock_move_many(totals, _wh(CERTIFICATION_WAREHOUSE), _wh("Finished Goods"))

	now = frappe.utils.now_datetime()
	for child, huid in picked:
		child.received = 1
		child.huid = huid
		child.received_on = now
		vals = {"stock_status": "In Stock", "in_stock_on": now,
			"certifications": _stamp_certification(child.order_bag, doc.get("cert_type") or doc.certification_type)}
		if huid:
			vals["huid"] = huid
		frappe.db.set_value("Order Bag", child.order_bag, vals)
	doc.status = "Received" if all(r.received for r in doc.items) else "Partially Received"
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": name, "received": len(picked), "status": doc.status, "stock_entry": se}


def _stamp_certification(order_bag, cert_type):
	"""Append the batch's type to the bag's certifications trail (no duplicates)."""
	cur = [x.strip() for x in (frappe.db.get_value("Order Bag", order_bag, "certifications") or "").split(",") if x.strip()]
	if cert_type and cert_type not in cur:
		cur.append(cert_type)
	return ", ".join(cur)


@frappe.whitelist()
def receive_certification_all(name):
	"""Stone-lab collection: we COUNT the packet and mark the whole batch
	collected — no per-piece scanning, no certificate numbers."""
	doc = frappe.get_doc("Certification", name)
	rows = [{"row": r.name, "huid": ""} for r in doc.items if not r.received]
	if not rows:
		frappe.throw(frappe._("Everything on {0} is already received.").format(name))
	return receive_certification(name, rows)


# IGI bulk-submission export — the user's own template shipped in-app; we only
# type values into it, so IGI's formatting/validation sheet ride along untouched.
_IGI_METAL_COLOR = {"YG": "Yellow Gold", "WG": "White Gold", "PG": "Pink Gold"}
# our diamond quality group -> (Color Criteria, Clarity Criteria) in IGI's wording
_IGI_QUALITY = {
	"SI-IJ": ("IJ", "SI"), "VS-FG": ("FG", "VS"), "VS-IJ": ("IJ", "VS"),
	"VVS-EF": ("EF", "VVS"), "VVS/VS-GH": ("GH", "VS/VVS"), "VVS1-EF": ("EF", "VVS1"),
	"VVS2": ("", "VVS2"),
}


@frappe.whitelist()
def export_igi_xlsx(bags, metal_type=None):
	"""Download the IGI bulk-submission workbook for the selected finished pieces —
	the shipped template filled from row 3. Style Number = design + card barcode;
	Diamond Weight = natural DMD + party diamonds (CVD excluded — lab-grown goes
	on its own submission); Jewelry Description via the IGI Description Map."""
	import re
	from io import BytesIO

	import openpyxl

	if isinstance(bags, str):
		bags = json.loads(bags or "[]")
	bags = [b for b in bags if b]
	if not bags:
		frappe.throw(frappe._("Pick at least one piece."))

	rows_out, unmapped = [], set()
	mats = _bag_convert_materials(bags)
	for nm in bags:
		b = frappe.db.get_value("Order Bag", nm, [
			"is_finished", "stock_status", "design", "huid", "act_gross_weight",
			"act_dmd_weight", "act_dmd_no", "act_pdmd_weight", "act_pdmd_no"], as_dict=True)
		if not b or not b.is_finished or b.stock_status not in ("In Stock", "At Certification"):
			frappe.throw(frappe._("{0} is not a finished piece In Stock / At Certification.").format(nm))
		dtype = frappe.db.get_value("Design", b.design, "design_type") if b.design else ""
		desc = frappe.db.get_value("IGI Description Map", dtype, "igi_description") if dtype else None
		if not desc:
			unmapped.add(dtype or "(no design type)")

		# walk the piece's frozen materials: gold -> metal color/karat; diamonds ->
		# dominant quality; coloured/precious -> names + carats
		karat = color = ""
		qual_ct, colored_names, colored_ct = {}, [], 0.0
		items = mats.get(nm, {})
		imeta = {i.name: i for i in frappe.get_all(
			"Item", filters={"name": ["in", list(items) or [""]]}, fields=["name", "stone_type", "item_group"])}
		for it, qty in items.items():
			m = imeta.get(it)
			if not m:
				continue
			if not m.stone_type:
				g = re.search(r"(\d{2}K)(YG|WG|PG)$", it)
				if g:
					karat, color = g.group(1), _IGI_METAL_COLOR[g.group(2)]
				elif not color:
					color = "Gold"
			elif m.stone_type in ("Diamond", "Party Diamond"):
				q = (m.item_group or "").replace("DIAMOND ", "") if m.stone_type == "Diamond" else ""
				if q:
					qual_ct[q] = qual_ct.get(q, 0) + qty
			elif m.stone_type in ("Color Stone", "Cubic Zirconia", "Precious Stone", "Party Other"):
				colored_names.append(it.split(" ")[0] if it[-1].isdigit() else it)
				colored_ct += qty
		best_q = max(qual_ct, key=qual_ct.get) if qual_ct else ""
		jk = _IGI_QUALITY.get(best_q, ("", ""))
		dmd_ct = round(flt(b.act_dmd_weight) + flt(b.act_pdmd_weight), 3)
		dmd_no = cint(b.act_dmd_no) + cint(b.act_pdmd_no)
		has_dmd = dmd_ct > 0 or dmd_no > 0
		rows_out.append({
			"B": "{0} / {1}".format(b.design or "", nm), "C": flt(b.act_gross_weight),
			"D": dmd_ct, "E": dmd_no, "F": desc or "",
			"G": color, "H": metal_type or "", "I": "Round Brilliant" if has_dmd else "",
			"J": jk[0] if has_dmd else "", "K": jk[1] if has_dmd else "", "L": "Very Good" if has_dmd else "",
			"M": ("{0} Hallmarked".format(karat) if b.huid else karat) if karat else "",
			"V": ", ".join(sorted(set(colored_names))), "Z": round(colored_ct, 3) or "",
		})
	if unmapped:
		frappe.throw(frappe._("No IGI wording mapped for design type(s): {0}.<br>Add them in the IGI Description Map list first.").format(
			", ".join(sorted(unmapped))))

	wb = openpyxl.load_workbook(frappe.get_app_path("jewelima", "data", "igi_template.xlsx"))
	ws = wb["Sheet1"]
	# wipe the template's sample rows, then fill ours from row 3
	# (ws.cell(..., value=None) would be a no-op — assign .value explicitly)
	for r in range(3, max(10, 3 + len(rows_out) + 3)):
		for c in range(1, 29):
			ws.cell(row=r, column=c).value = None
	for i, row in enumerate(rows_out):
		for col, val in row.items():
			if val != "" and val is not None:
				ws[f"{col}{3 + i}"] = val
	buf = BytesIO()
	wb.save(buf)
	frappe.local.response.filename = "IGI-{0}-{1}pc.xlsx".format(frappe.utils.today(), len(rows_out))
	frappe.local.response.filecontent = buf.getvalue()
	frappe.local.response.type = "binary"


@frappe.whitelist()
def get_certification_batches():
	"""Open batches (plus the freshest few received) for the Certification Out
	board. Items carry pure gold + total stone ct (frozen actuals) so the header
	can total what's physically out; summary covers UNRECEIVED pieces only."""
	names = frappe.get_all("Certification", filters={"status": ["!=", "Received"]},
	                       order_by="creation desc", pluck="name")
	names += frappe.get_all("Certification", filters={"status": "Received"},
	                        order_by="modified desc", limit=3, pluck="name")
	all_bags = set()
	docs = []
	for nm in names:
		d = frappe.get_doc("Certification", nm)
		docs.append(d)
		all_bags.update(r.order_bag for r in d.items)
	acts = {b.name: b for b in frappe.get_all(
		"Order Bag", filters={"name": ["in", list(all_bags) or [""]]},
		fields=["name", "act_pure_weight", "act_dmd_weight", "act_ps_weight", "act_cs_weight",
		        "act_cz_weight", "act_cvd_weight", "act_pdmd_weight", "act_poth_weight"])}

	def stones_of(bag):
		a = acts.get(bag)
		return round(sum(flt(a.get(f"act_{x}_weight")) for x in ("dmd", "ps", "cs", "cz", "cvd", "pdmd", "poth")), 3) if a else 0

	out, summary = [], {"pieces_out": 0, "pure_gold": 0.0, "stones_ct": 0.0, "batches_out": 0}
	for d in docs:
		items = [{
			"row": r.name, "order_bag": r.order_bag, "design": r.design or "", "design_type": r.design_type or "",
			"gross": flt(r.gross), "dmd_ct": flt(r.dmd_ct), "received": cint(r.received),
			"pure": flt((acts.get(r.order_bag) or {}).get("act_pure_weight")), "stones_ct": stones_of(r.order_bag),
			"huid": r.huid or "",
		} for r in d.items]
		if d.status != "Received":
			summary["batches_out"] += 1
			for it in items:
				if not it["received"]:
					summary["pieces_out"] += 1
					summary["pure_gold"] += it["pure"]
					summary["stones_ct"] += it["stones_ct"]
		out.append({
			"name": d.name, "certification_type": d.certification_type, "status": d.status,
			"sent_on": str(d.sent_on or ""), "lab": d.lab or "", "remarks": d.remarks or "",
			"total": len(d.items), "back": sum(1 for r in d.items if r.received),
			"items": items,
		})
	summary["pure_gold"] = round(summary["pure_gold"], 3)
	summary["stones_ct"] = round(summary["stones_ct"], 3)
	return {"batches": out, "summary": summary}


@frappe.whitelist()
def get_finished_items(status=None, held_by=None):
	"""The finished-goods register: every bag made into a product, with its frozen
	weights, holder, design + design type and stock status."""
	filters = {"is_finished": 1}
	if status:
		filters["stock_status"] = status
	if held_by:
		filters["held_by"] = held_by
	rows = frappe.get_all(
		"Order Bag", filters=filters,
		fields=[
			"name", "design", "held_by", "stock_status",
			"act_gross_weight", "act_nett_weight", "act_pure_weight", "act_purity",
			"act_dmd_no", "act_dmd_weight", "act_cs_no", "act_cs_weight",
		],
		order_by="modified desc", limit=2000,
	)
	designs = list({r.design for r in rows if r.design})
	dtype = {}
	if designs:
		dtype = {d.name: d.design_type for d in frappe.get_all("Design", filters={"name": ["in", designs]}, fields=["name", "design_type"])}
	for r in rows:
		r["design_type"] = dtype.get(r.design)
	return rows


@frappe.whitelist()
def get_print_branding():
	"""Company + logo for the shared print header (logo ships with the app; company
	name/address/contact come from the ERPNext Company so they're editable)."""
	company = _company()
	c = {}
	if company:
		c = frappe.db.get_value("Company", company, ["company_name", "phone_no", "email", "tax_id", "website"], as_dict=True) or {}
	addr = ""
	if company:
		links = frappe.get_all("Dynamic Link", filters={"link_doctype": "Company", "link_name": company, "parenttype": "Address"}, fields=["parent"], limit=1)
		if links:
			a = frappe.db.get_value("Address", links[0].parent, ["address_line1", "address_line2", "city", "state", "pincode"], as_dict=True) or {}
			loc = ", ".join([x for x in [a.get("city"), a.get("state"), a.get("pincode")] if x])
			addr = ", ".join([x for x in [a.get("address_line1"), a.get("address_line2"), loc] if x])
	return {
		"company": c.get("company_name") or company or "Jewelima",
		"address": addr, "phone": c.get("phone_no") or "", "email": c.get("email") or "",
		"gstin": c.get("tax_id") or "", "website": c.get("website") or "",
		"logo_url": "/assets/jewelima/images/jewelima-letterhead.png",
	}


@frappe.whitelist()
def get_card_passport(order_bag):
	"""Everything about a card for the lookup/print view: header, plan + actual
	weights, current contents, the transfer trail and the bench stage history."""
	bag = frappe.db.get_value(
		"Order Bag", order_bag,
		[
			"name", "design", "qty", "size", "location", "tree", "stock_status", "held_by", "customer", "salesman",
			"order_type", "order_date", "due_date", "customer_date", "is_finished", "narration", "image", "job_order",
			"gross_weight", "nett_weight", "purity", "dmd_no", "dmd_weight", "ps_no", "ps_weight", "cs_no", "cs_weight",
			"act_gross_weight", "act_nett_weight", "act_pure_weight", "act_purity",
			"act_dmd_no", "act_dmd_weight", "act_ps_no", "act_ps_weight", "act_cs_no", "act_cs_weight",
		],
		as_dict=True,
	)
	if not bag:
		return {}
	if bag.design:
		d = frappe.db.get_value("Design", bag.design, ["design_type", "item"], as_dict=True) or {}
		bag["design_type"] = d.get("design_type")
		bag["item"] = d.get("item")
	# material issues into the card — who issued what stones/gold and when
	issues = frappe.db.sql("""
		SELECT l.item, i.stone_type, l.entry_type, l.direction, l.qty, l.pcs, l.datetime,
			IFNULL(e.employee_name, IFNULL(l.employee, '')) who, l.remarks
		FROM `tabBag Material Ledger` l
		JOIN `tabItem` i ON i.name = l.item
		LEFT JOIN `tabEmployee` e ON e.name = l.employee
		WHERE l.order_bag = %s AND l.entry_type IN ('Stone Issue', 'Gold Issue')
		ORDER BY l.datetime""", order_bag, as_dict=True)
	return {
		"bag": bag,
		"contents": get_bag_contents(order_bag),
		"transfers": frappe.get_all(
			"Order Bag Transfer", filters={"order_bag": order_bag},
			fields=["from_location", "to_location", "transfer_time", "transferred_by"], order_by="transfer_time asc",
		),
		"stages": get_bag_stage_history(order_bag),
		"issues": [{"item": r.item, "stone_type": r.stone_type or "", "entry_type": r.entry_type,
			"direction": r.direction, "qty": flt(r.qty), "pcs": cint(r.pcs),
			"datetime": str(r.datetime or ""), "who": r.who, "remarks": r.remarks or ""} for r in issues],
	}


@frappe.whitelist()
def get_cards_at_location(location):
	"""All ACTIVE production cards currently at a location, with each card's current bench
	status (In Queue / Issued / Completed …) — the Transfer page's Cards picker."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	loc = (location or "").strip().upper()
	if not loc:
		return []
	bags = frappe.get_all(
		"Order Bag",
		filters={"location": loc, "is_finished": 0, "stock_status": ["not in", ["Cancelled", "Sold"]]},
		fields=["name", "design", "qty", "due_date"], order_by="name",
	)
	dt = BENCH_DOCTYPE.get(loc)
	smap = {}
	if bags and dt and frappe.db.exists("DocType", dt):
		for r in frappe.get_all(
			dt, filters={"order_bag": ["in", [b.name for b in bags]], "bench": loc},
			fields=["order_bag", "status"], order_by="creation asc",
		):
			smap[r.order_bag] = r.status  # last record per bag wins
	for b in bags:
		b["status"] = smap.get(b.name) or "In Queue"
	return bags


@frappe.whitelist()
def transfer_order_bags(names, to_location, remarks=None):
	"""Transfer a batch of Order Bags (all collected at one source) to a destination."""
	if isinstance(names, str):
		names = json.loads(names or "[]")
	done, errors = [], []
	for nm in names or []:
		try:
			transfer_order_bag(nm, to_location, remarks)
			done.append(nm)
		except Exception as e:
			errors.append({"name": nm, "error": str(e)})
	return {"transferred": done, "count": len(done), "errors": errors}


def _qr_data_uri(text):
	"""Standard (full) QR as a PNG data-URI via segno. make_qr() forces a real QR with all
	three finder patterns — segno.make() emits a Micro QR for short codes, which most phone
	cameras / barcode scanners can't read. None if segno is unavailable."""
	try:
		import segno

		return segno.make_qr(str(text), error="m").png_data_uri(scale=4, border=2)
	except Exception:
		return None


@frappe.whitelist()
def get_barcode_card(order_bag):
	"""Label data for the Print Barcode page. Weights come from ACTUAL (never the BOM/plan);
	`actual_empty` flags cards with no actual weight yet (the page warns). Errors if qty > 1 —
	a multi-piece card must be extracted into singles first. No status restriction: a card can
	be barcode-printed at any stage."""
	if not order_bag or not frappe.db.exists("Order Bag", order_bag):
		return {"error": "No Order Bag {0}.".format(order_bag)}
	b = frappe.get_doc("Order Bag", order_bag)
	if int(b.qty or 0) > 1:
		return {"error": "{0}: qty is {1} — extract it into single pieces before printing.".format(b.name, b.qty)}
	gw = flt(b.act_gross_weight)
	return {
		"name": b.name,
		"design": b.design,
		"design_type": frappe.db.get_value("Design", b.design, "design_type") if b.design else None,
		"size": b.size,
		"qty": int(b.qty or 0),
		"gw": gw,
		"nett": flt(b.act_nett_weight),
		"purity": flt(b.act_purity),
		"dmd_no": int(b.act_dmd_no or 0), "dmd_wt": flt(b.act_dmd_weight),
		"ps_no": int(b.act_ps_no or 0), "ps_wt": flt(b.act_ps_weight),
		"cs_no": int(b.act_cs_no or 0), "cs_wt": flt(b.act_cs_weight),
		"actual_empty": not gw,
		"qr": _qr_data_uri(b.name),
	}


@frappe.whitelist()
def get_job_order_status(job_order):
	"""Where every piece of a Job Order is right now: current location, whether it's
	assigned/issued (and to whom), and when it entered that location. Accepts a Job Order
	name or any of its card codes. Used by the Job Order Status page."""
	from collections import Counter

	from jewelima.jewelima.benches import bench_doctype

	if not job_order:
		return {"error": "Enter a Job Order or a card code."}
	# accept a card code too -> resolve to its job order
	if not frappe.db.exists("Job Order", job_order):
		jo2 = frappe.db.get_value("Order Bag", job_order, "job_order")
		if jo2:
			job_order = jo2
		else:
			return {"error": "No Job Order (or card) '{0}'.".format(job_order)}

	bags = frappe.get_all(
		"Order Bag", filters={"job_order": job_order},
		fields=["name", "design", "qty", "location", "is_finished", "stock_status", "act_gross_weight"],
		order_by="name asc",
	)
	emp_names, out = {}, []
	for b in bags:
		loc = b.location
		status = employee = entered = None
		dt = bench_doctype(loc)
		if dt and frappe.db.exists("DocType", dt):
			recs = frappe.get_all(
				dt, filters={"order_bag": b.name},
				fields=["status", "employee", "time_in", "issued_at"], order_by="creation desc", limit=1,
			)
			if recs:
				status, employee = recs[0].status, recs[0].employee
				entered = recs[0].issued_at or recs[0].time_in
		if not entered and loc:
			entered = frappe.db.get_value(
				"Order Bag Transfer", {"order_bag": b.name, "to_location": loc}, "transfer_time", order_by="transfer_time desc"
			)
		if not entered:
			entered = frappe.db.get_value("Order Bag", b.name, "creation")
		if employee and employee not in emp_names:
			emp_names[employee] = frappe.db.get_value("Employee", employee, "employee_name") or employee
		out.append({
			"name": b.name, "design": b.design, "qty": b.qty,
			"location": loc or "—", "is_finished": b.is_finished, "stock_status": b.stock_status,
			"gross": flt(b.act_gross_weight),
			"status": status, "employee": employee, "employee_name": emp_names.get(employee),
			"entered": str(entered) if entered else None,
		})

	header = frappe.db.get_value(
		"Job Order", job_order, ["customer", "salesman", "order_type", "order_date", "due_date"], as_dict=True
	) or {}
	return {
		"job_order": job_order, "header": header, "bags": out, "total": len(out),
		"by_location": dict(Counter(x["location"] for x in out)),
	}


@frappe.whitelist()
def get_employee_performance(days=30):
	"""Per-employee shop-floor performance for the live TV dashboard: pieces completed,
	gold handled, loss + loss%, pieces today, work in hand now, and current held weight.
	Aggregated across every bench doctype over the last `days`."""
	from jewelima.jewelima.benches import BENCH_DOCTYPE

	days = int(days or 30)
	since = frappe.utils.add_days(frappe.utils.nowdate(), -days)
	today = frappe.utils.nowdate()
	agg = {}
	for dt in dict.fromkeys(BENCH_DOCTYPE.values()):
		if not frappe.db.exists("DocType", dt):
			continue
		for r in frappe.get_all(
			dt,
			filters={"employee": ["is", "set"], "creation": [">=", since]},
			fields=["employee", "status", "weight_out", "loss", "receipted_at", "creation"],
		):
			a = agg.setdefault(r.employee, {"pieces": 0, "gold": 0.0, "loss": 0.0, "today": 0, "active": 0})
			if r.status in ("Receipted", "Completed"):
				a["pieces"] += 1
				a["gold"] += flt(r.weight_out)
				a["loss"] += flt(r.loss)
				if str(r.receipted_at or r.creation)[:10] == today:
					a["today"] += 1
			elif r.status in ("Issued", "Ongoing"):
				a["active"] += 1

	rows = []
	for emp, a in agg.items():
		rows.append({
			"employee": emp,
			"name": frappe.db.get_value("Employee", emp, "employee_name") or emp,
			"pieces": a["pieces"], "gold": round(a["gold"], 3), "loss": round(a["loss"], 3),
			"loss_pct": round(a["loss"] / a["gold"] * 100, 2) if a["gold"] else 0,
			"today": a["today"], "active": a["active"],
			"holding": round(flt(frappe.db.get_value("Employee Metal Balance", emp, "current_weight")), 3),
		})
	rows.sort(key=lambda x: (-x["pieces"], x["loss_pct"]))
	return {"period_days": days, "as_of": frappe.utils.now_datetime().strftime("%d %b %H:%M"), "rows": rows}


@frappe.whitelist()
def get_tv_overview():
	"""Headline shop KPIs for the at-a-glance TV board: orders today/this month, pieces in
	production, pieces completed + loss today, workers active now, gold in hand, and a
	breakdown of work-in-progress by stage."""
	from collections import Counter

	from jewelima.jewelima.benches import BENCH_DOCTYPE

	today = frappe.utils.nowdate()
	month_start = today[:8] + "01"
	done_today, loss_today, active = 0, 0.0, set()
	for dt in dict.fromkeys(BENCH_DOCTYPE.values()):
		if not frappe.db.exists("DocType", dt):
			continue
		for r in frappe.get_all(dt, filters={"receipted_at": [">=", today]}, fields=["loss"]):
			done_today += 1
			loss_today += flt(r.loss)
		for r in frappe.get_all(dt, filters={"status": ["in", ["Issued", "Ongoing"]], "employee": ["is", "set"]}, fields=["employee"]):
			active.add(r.employee)
	gold = sum(flt(b.current_weight) for b in frappe.get_all("Employee Metal Balance", fields=["current_weight"]))
	by_stage = Counter((b.location or "—") for b in frappe.get_all("Order Bag", filters={"is_finished": 0}, fields=["location"]))
	return {
		"as_of": frappe.utils.now_datetime().strftime("%d %b %H:%M"),
		"orders_today": frappe.db.count("Job Order", {"order_date": today}),
		"orders_month": frappe.db.count("Job Order", {"order_date": [">=", month_start]}),
		"in_production": frappe.db.count("Order Bag", {"is_finished": 0}),
		"done_today": done_today,
		"active_workers": len(active),
		"gold_in_hand": round(gold, 1),
		"loss_today": round(loss_today, 3),
		"by_stage": [{"stage": k, "n": v} for k, v in by_stage.most_common(10)],
	}


@frappe.whitelist()
def get_orders_taken(days=30):
	"""Order intake from the ORIGINAL Order Bags (extraction pieces excluded via split_of),
	focused on TODAY: today's totals + karat split (22K/18K/14K) + top customers, plus a
	daily pure-gold trend over `days` for context. Plan/order-time weights; pure gold =
	nett grams x purity%."""
	from collections import defaultdict

	days = int(days or 30)
	today = frappe.utils.nowdate()
	since = frappe.utils.add_days(today, -days)
	bands = [("22K", 88, 96), ("18K", 70, 80), ("14K", 54, 63)]

	def karat(p):
		for k, lo, hi in bands:
			if lo <= p <= hi:
				return k
		return "Other"

	bags = frappe.get_all(
		"Order Bag",
		filters={"split_of": ["is", "not set"], "order_date": [">=", since]},
		fields=["order_date", "purity", "gross_weight", "nett_weight", "dmd_weight", "ps_weight", "cs_weight", "qty", "customer"],
	)
	tbk = defaultdict(lambda: {"orders": 0, "qty": 0, "gross": 0.0, "pure": 0.0})   # today, by karat
	ttot = {"orders": 0, "qty": 0, "gross": 0.0, "pure": 0.0, "dmd": 0.0, "ps": 0.0, "cs": 0.0}
	tcust = defaultdict(float)                                                      # today, by customer
	day = defaultdict(lambda: {"22K": 0.0, "18K": 0.0, "14K": 0.0, "Other": 0.0, "qty": 0})  # trend
	for b in bags:
		p = flt(b.purity)
		k = karat(p)
		q = int(b.qty or 0)
		pure = flt(b.nett_weight) * p / 100.0
		d = str(b.order_date)
		day[d][k] += pure; day[d]["qty"] += q
		if d == today:
			tbk[k]["orders"] += 1; tbk[k]["qty"] += q; tbk[k]["gross"] += flt(b.gross_weight); tbk[k]["pure"] += pure
			ttot["orders"] += 1; ttot["qty"] += q; ttot["gross"] += flt(b.gross_weight); ttot["pure"] += pure
			ttot["dmd"] += flt(b.dmd_weight); ttot["ps"] += flt(b.ps_weight); ttot["cs"] += flt(b.cs_weight)
			if b.customer:
				tcust[b.customer] += pure

	today_kar = [
		{"karat": k, "orders": tbk[k]["orders"], "qty": tbk[k]["qty"], "gross": round(tbk[k]["gross"], 1), "pure": round(tbk[k]["pure"], 1)}
		for k in ("22K", "18K", "14K", "Other") if tbk[k]["orders"]
	]
	trend = [
		{"date": d, "k22": round(v["22K"], 1), "k18": round(v["18K"], 1), "k14": round(v["14K"], 1), "qty": v["qty"]}
		for d, v in sorted(day.items())
	]
	top = sorted(tcust.items(), key=lambda x: -x[1])[:8]
	return {
		"date": frappe.utils.formatdate(today, "dd MMM yyyy"),
		"as_of": frappe.utils.now_datetime().strftime("%H:%M"),
		"days": days,
		"today": {**{k: (round(v, 1) if isinstance(v, float) else v) for k, v in ttot.items()}, "by_karat": today_kar},
		"trend": trend,
		"top_customers": [{"customer": c, "pure": round(w, 1)} for c, w in top],
	}
